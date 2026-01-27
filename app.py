from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, abort
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func
from models import db, User, Horario, Carrera, Materia, HorarioAcademico, DisponibilidadProfesor, Grupo, init_db, init_upload_dirs, AsignacionProfesorGrupo, Role
from forms import (LoginForm, RegistrationForm, HorarioForm, EliminarHorarioForm, 
                   CarreraForm, ImportarProfesoresForm, FiltrarProfesoresForm, ExportarProfesoresForm,
                   MateriaForm, ImportarMateriasForm, FiltrarMateriasForm, ExportarMateriasForm,
                   GenerarHorariosForm, EditarHorarioAcademicoForm, EliminarHorarioAcademicoForm,
                   DisponibilidadProfesorForm, EditarDisponibilidadProfesorForm, AgregarProfesorForm,
                   EditarUsuarioForm, AsignarMateriasProfesorForm, GrupoForm, AsignarMateriasGrupoForm,
                   CambiarPasswordProfesorForm, ImportarCarrerasForm, ImportarAsignacionesForm)
from utils import (procesar_archivo_profesores, generar_pdf_profesores, procesar_archivo_materias, 
                   generar_pdf_materias, generar_plantilla_csv, procesar_archivo_carreras, 
                   generar_plantilla_csv_carreras, procesar_archivo_asignaciones, 
                   generar_plantilla_csv_asignaciones, calcular_carga_profesor)
from datetime import time, datetime
import os
import pandas as pd
from io import BytesIO, StringIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, A4, landscape 
import threading
import time
import json

# Variable global para tracking del progreso de generación masiva
generacion_progreso = {
    'activo': False,
    'total_grupos': 0,
    'grupos_procesados': 0,
    'grupo_actual': '',
    'codigo_grupo': '',
    'horarios_generados': 0,
    'mensaje': '',
    'completado': False,
    'exito': False,
    'iniciado': False
}

# Lock para acceso thread-safe
progreso_lock = threading.Lock()

import re


# ==========================================
# FUNCIONES AUXILIARES DE CARGA HORARIA
# ==========================================
def obtener_limite_horas_profesor(tipo_profesor):
    """
    Obtiene el límite de horas semanales para un tipo de profesor.
    Los valores se configuran desde el panel de administración.
    """
    from models import ConfiguracionSistema
    
    tipo_lower = (tipo_profesor or '').lower().strip()
    
    if 'tiempo completo' in tipo_lower or 'tiempo_completo' in tipo_lower:
        return ConfiguracionSistema.get_config('horas_tiempo_completo', 40)
    elif 'asignatura' in tipo_lower:
        return ConfiguracionSistema.get_config('horas_asignatura', 20)
    elif 'medio tiempo' in tipo_lower or 'medio_tiempo' in tipo_lower:
        return ConfiguracionSistema.get_config('horas_medio_tiempo', 20)
    else:
        # Tipo no reconocido: usar límite de asignatura como default
        return ConfiguracionSistema.get_config('horas_asignatura', 20)


def obtener_horas_actuales_profesor(profesor_id):
    """
    Calcula las horas actuales asignadas a un profesor.
    """
    from models import HorarioAcademico
    return HorarioAcademico.query.filter_by(profesor_id=profesor_id, activo=True).count()


def validar_carga_horaria_profesor(profesor):
    """
    Valida si un profesor puede recibir más horas de clase.
    Retorna tuple: (puede_asignar, horas_actuales, limite, horas_disponibles)
    """
    from models import ConfiguracionSistema
    
    horas_actuales = obtener_horas_actuales_profesor(profesor.id)
    limite = obtener_limite_horas_profesor(profesor.tipo_profesor)
    limite_absoluto = ConfiguracionSistema.get_config('horas_limite_absoluto', 50)
    
    # El límite efectivo es el menor entre el límite del tipo y el absoluto
    limite_efectivo = min(limite, limite_absoluto)
    horas_disponibles = max(0, limite_efectivo - horas_actuales)
    puede_asignar = horas_disponibles > 0
    
    return (puede_asignar, horas_actuales, limite_efectivo, horas_disponibles)


app = Flask(__name__)

# Configuración de la aplicación
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui_cambiala_en_produccion'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sistema_academico.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Inicializar extensiones
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    """Cargar usuario por ID para Flask-Login"""
    return User.query.get(int(user_id))

# Rutas principales
@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de inicio de sesión"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data):
            if user.activo:
                login_user(user)
                
                # Verificar si el usuario requiere cambio de contraseña
                if user.requiere_cambio_password:
                    flash(f'Bienvenido, {user.get_nombre_completo()}. Por seguridad, debes cambiar tu contraseña temporal.', 'warning')
                    return redirect(url_for('cambiar_password_obligatorio'))
                
                flash(f'¡Bienvenido, {user.get_nombre_completo()}!', 'success')
                
                # Redirigir a la página solicitada o al dashboard
                next_page = request.args.get('next')
                return redirect(next_page) if next_page else redirect(url_for('dashboard'))
            else:
                flash('Tu cuenta está desactivada. Contacta al administrador.', 'error')
        else:
            flash('Usuario o contraseña incorrectos.', 'error')
    
    return render_template('login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Página de registro"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    
    # Obtener horarios para el formulario de disponibilidad
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    if form.validate_on_submit():
        try:
            # Obtener el rol final (considerando tipo de profesor)
            rol_final = form.get_final_rol()
            
            # Para profesores y jefes de carrera, obtener las carreras seleccionadas
            carreras = []
            if rol_final in ['profesor_completo', 'profesor_asignatura', 'jefe_carrera'] and form.carrera.data:
                carreras = Carrera.query.filter(Carrera.id.in_(form.carrera.data)).all()
            
            # Crear nuevo usuario - ahora todos usan la relación many-to-many carreras
            user = User(
                username=form.username.data,
                email=form.email.data,
                password=form.password.data,
                nombre=form.nombre.data,
                apellido=form.apellido.data,
                rol=rol_final,
                telefono=form.telefono.data if form.telefono.data else None,
                carreras=carreras
            )
            
            db.session.add(user)
            db.session.commit()
            
            # Procesar disponibilidad si es profesor
            if rol_final in ['profesor_completo', 'profesor_asignatura']:
                dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
                
                for horario in horarios:
                    for dia in dias:
                        # Verificar si el checkbox fue marcado
                        field_name = f"availability_{horario.id}_{dia}"
                        if request.form.get(field_name):
                            disponibilidad = DisponibilidadProfesor(
                                profesor_id=user.id,
                                horario_id=horario.id,
                                dia_semana=dia,
                                disponible=True
                            )
                            db.session.add(disponibilidad)
                
                db.session.commit()
            
            flash(f'¡Registro exitoso! Bienvenido, {user.get_nombre_completo()}.', 'success')
            
            # Iniciar sesión automáticamente después del registro
            login_user(user)
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al crear la cuenta. Inténtalo de nuevo.', 'error')
            print(f"Error en registro: {e}")
    
    return render_template('register.html', form=form, horarios=horarios)

@app.route('/cambiar-password-obligatorio', methods=['GET', 'POST'])
@login_required
def cambiar_password_obligatorio():
    """Página obligatoria para cambiar contraseña temporal"""
    # Si el usuario ya cambió su contraseña, redirigir al dashboard
    if not current_user.requiere_cambio_password:
        return redirect(url_for('dashboard'))
    
    from forms import CambiarPasswordObligatorioForm
    form = CambiarPasswordObligatorioForm()
    
    if form.validate_on_submit():
        try:
            # Verificar que la contraseña actual sea correcta
            if not current_user.check_password(form.password_actual.data):
                flash('La contraseña actual es incorrecta.', 'error')
                return render_template('cambiar_password_obligatorio.html', form=form)
            
            # Actualizar contraseña
            current_user.password = form.nueva_password.data
            current_user.requiere_cambio_password = False
            current_user.password_temporal = None
            
            db.session.commit()
            
            flash('¡Contraseña actualizada exitosamente! Ahora puedes acceder al sistema.', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al cambiar la contraseña. Inténtalo de nuevo.', 'error')
            print(f"Error en cambiar password obligatorio: {e}")
    
    return render_template('cambiar_password_obligatorio.html', form=form, password_temporal=current_user.password_temporal)

# ==========================================
# FUNCIÓN CENTRAL PARA PROCESAR HORARIOS
# ==========================================
def procesar_horarios(agrupar_por='profesor', carrera_id=None, incluir_ids=False):
    """
    Función centralizada para obtener y procesar los horarios académicos.
    
    :param agrupar_por: 'profesor' o 'grupo'. Define cómo se agruparán los datos.
    :param carrera_id: Opcional. Si se provee un ID, filtra los horarios para esa carrera.
    :param incluir_ids: Si True, incluye los IDs de los horarios para acciones
    :return: Un diccionario con los horarios organizados.
    """
    
    # 1. Consulta base a la base de datos
    query = HorarioAcademico.query.filter_by(activo=True)
    
    # 2. Si se especifica una carrera_id, filtramos los resultados
    if carrera_id:
        query = query.join(Materia).filter(Materia.carrera_id == carrera_id)
        
    asignaciones = query.all()
    
    # 3. Ordenamos los resultados en Python
    asignaciones.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
    
    # 4. Diccionario para almacenar el resultado final
    datos_organizados = {}
    
    # Mapeo de días para asegurar formato correcto
    dias_map = {
        'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
        'jueves': 'Jueves', 'viernes': 'Viernes'
    }

    # 5. Iteramos sobre cada asignación para construir el diccionario
    for a in asignaciones:
        if not all([a.profesor, a.materia, a.horario]):
            continue

        clave_agrupacion = None
        info_clase_html = ""
        
        # Lógica para agrupar por PROFESOR
        if agrupar_por == 'profesor':
            clave_agrupacion = a.profesor.get_nombre_completo()
            if incluir_ids:
                info_clase_html = {
                    'id': a.id,
                    'html': f"{a.materia.nombre}<br><small class='text-muted'>{a.materia.codigo}</small><br>{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
                }
            else:
                info_clase_html = (
                    f"{a.materia.nombre}<br>"
                    f"<small class='text-muted'>{a.materia.codigo}</small><br>"
                    f"{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
                )

        # Lógica para agrupar por GRUPO
        elif agrupar_por == 'grupo':
            # CORREGIDO: Usar el campo 'grupo' del HorarioAcademico en lugar de materia.grupos
            grupo_codigo = a.grupo
            
            if grupo_codigo:
                # Buscar el objeto Grupo por su código para filtrar por carrera si es necesario
                grupo = Grupo.query.filter_by(codigo=grupo_codigo).first()
                
                if grupo and (carrera_id is None or grupo.carrera_id == carrera_id):
                    clave_agrupacion = grupo_codigo
                    # Obtener la hora de inicio como entero para la cuadrícula
                    hora_inicio_int = a.horario.hora_inicio.hour if a.horario.hora_inicio else 7
                    if incluir_ids:
                        info_clase_html = {
                            'id': a.id,
                            'grupo_id': grupo.id,
                            'hora_inicio': hora_inicio_int,
                            'materia': a.materia.nombre,
                            'profesor': a.profesor.get_nombre_completo(),
                            'hora_texto': f"{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}",
                            'html': f"{a.materia.nombre}<br>Prof: {a.profesor.get_nombre_completo()}<br>{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
                        }
                    else:
                        info_clase_html = (
                            f"{a.materia.nombre}<br>"
                            f"Prof: {a.profesor.get_nombre_completo()}<br>"
                            f"{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
                        )

        # Si encontramos una clave válida, la agregamos al diccionario
        if clave_agrupacion:
            if clave_agrupacion not in datos_organizados:
                datos_organizados[clave_agrupacion] = {d: [] for d in dias_map.values()}
            
            dia_correcto = dias_map.get(a.dia_semana.lower())
            if dia_correcto:
                datos_organizados[clave_agrupacion][dia_correcto].append(info_clase_html)

    # 6. Ordenamos el diccionario final si es por grupo para una mejor presentación
    if agrupar_por == 'grupo':
        # Importamos re aquí si no está arriba o usamos la función auxiliar si tenemos
        import re
        def natural_keys(text):
            return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', text)]
            
        datos_ordenados = {k: datos_organizados[k] for k in sorted(datos_organizados.keys(), key=natural_keys)}
        return datos_ordenados
        
    return datos_organizados


# =================================================================
# FUNCIÓN "CEREBRO" PARA OBTENER DATOS DETALLADOS (FORMATO FDA)
# =================================================================

def procesar_horarios_formato_fda(carrera_id=None):
    """
    Obtiene los datos de horarios con todos los detalles necesarios 
    para generar el formato de Carga Horaria (FDA).
    """
    query = HorarioAcademico.query.filter_by(activo=True)
    
    if carrera_id:
        query = query.join(Materia).filter(Materia.carrera_id == carrera_id)
        
    asignaciones = query.all()
    
    horarios_por_profesor = {}
    
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

    for a in asignaciones:
        if not all([a.profesor, a.materia, a.horario, a.materia.carrera]):
            continue

        profesor_nombre = a.profesor.get_nombre_completo()
        
        if profesor_nombre not in horarios_por_profesor:
            # ¡CAMBIO IMPORTANTE! Guardamos la info en un diccionario, no en el objeto.
            info_profesor = {
                'nombre_completo': a.profesor.get_nombre_completo(),
                # Lógica de ejemplo para TC/PA. Debes adaptar 'tipo_contrato' al nombre
                # del campo en tu modelo 'Profesor' o 'Usuario'.
                'es_tc': getattr(a.profesor, 'tipo_contrato', 'PA').upper() == 'TC'
            }
            horarios_por_profesor[profesor_nombre] = {
                'info': info_profesor,
                'clases': []
            }
        
        duracion_horas = (a.horario.hora_fin.hour - a.horario.hora_inicio.hour) + (a.horario.hora_fin.minute - a.horario.hora_inicio.minute) / 60.0
        
        # CORREGIDO: Usar el campo 'grupo' del HorarioAcademico directamente
        grupo_codigo = a.grupo if a.grupo else "N/A"

        dia_correcto = dias_map.get(a.dia_semana.lower())
        if not dia_correcto: continue

        detalle_clase = {
            'clave': a.materia.codigo, 'asignatura': a.materia.nombre, 'grupo': grupo_codigo,
            'dia_raw': a.dia_semana.lower(), 'hora_inicio': a.get_hora_inicio_str(),
            'hora_fin': a.get_hora_fin_str(), 'horas_totales': duracion_horas,
            'carrera': a.materia.carrera.codigo
        }
        horarios_por_profesor[profesor_nombre]['clases'].append(detalle_clase)

    for data in horarios_por_profesor.values():
        data['clases'].sort(key=lambda c: (['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado'].index(c['dia_raw']), c['hora_inicio']))

    return horarios_por_profesor

# =================================================================
# FUNCIÓN PARA GENERAR EL REPORTE FDA EN EXCEL (VERSIÓN CORREGIDA)
# =================================================================
def generar_excel_formato_fda(datos_profesor):
    """
    Crea un archivo Excel con el formato de Carga Horaria (FDA) para un profesor.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Carga Horaria"

    bold_font = Font(bold=True, name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))



    ws.merge_cells('D1:J4'); ws['D1'] = 'Formato de Carga Horaria'; ws['D1'].font = Font(bold=True, size=16, name='Arial'); ws['D1'].alignment = center_align
    ws.merge_cells('K1:L2'); ws['K1'] = 'CÓDIGO'; ws['K1'].font = bold_font; ws['K1'].alignment = center_align; ws['K1'].border = thin_border
    ws.merge_cells('K3:L4'); ws['K3'] = 'FDA-25'; ws['K3'].alignment = center_align; ws['K3'].border = thin_border

    # --- ¡LÍNEAS CORREGIDAS! ---
    # Ahora accedemos a los datos como un diccionario con ['key']
    ws.merge_cells('A6:F6')
    ws['A6'].value = f"NOMBRE DEL PROFESOR: {datos_profesor['info']['nombre_completo']}"
    
    ws.merge_cells('G6:I6')
    ws['G6'].value = f"TC: {'X' if datos_profesor['info']['es_tc'] else ''}   PA: {'X' if not datos_profesor['info']['es_tc'] else ''}"
    # --- FIN DE LA CORRECCIÓN ---

    headers = ['No.', 'CLAVE', 'ASIGNATURA', 'GPO.', 'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'HRS. T.', 'CARRERA']
    ws.append(headers)
    for cell in ws[8]: cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border

    row_num, total_horas = 9, 0
    for i, clase in enumerate(datos_profesor['clases'], 1):
        hora_str = f"{clase['hora_inicio']}\n{clase['hora_fin']}"
        row_data = [i, clase['clave'], clase['asignatura'], clase['grupo'],
            hora_str if clase['dia_raw'] == 'lunes' else '', hora_str if clase['dia_raw'] == 'martes' else '',
            hora_str if clase['dia_raw'] == 'miercoles' else '', hora_str if clase['dia_raw'] == 'jueves' else '',
            hora_str if clase['dia_raw'] == 'viernes' else '', hora_str if clase['dia_raw'] == 'sabado' else '',
            f"{clase['horas_totales']:.1f}", clase['carrera']]
        ws.append(row_data)
        total_horas += clase['horas_totales']
        for cell in ws[row_num]: cell.alignment = center_align; cell.border = thin_border
        ws[f'C{row_num}'].alignment = left_align
        row_num += 1

    ws.merge_cells(f'A{row_num}:J{row_num}'); ws[f'A{row_num}'].value = 'TOTAL DE HORAS FRENTE A GRUPO'; ws[f'A{row_num}'].font = bold_font; ws[f'A{row_num}'].alignment = Alignment(horizontal='right'); ws[f'A{row_num}'].border = thin_border
    ws[f'K{row_num}'].value = f"{total_horas:.1f}"; ws[f'K{row_num}'].font = bold_font; ws[f'K{row_num}'].alignment = center_align; ws[f'K{row_num}'].border = thin_border
    ws[f'L{row_num}'].border = thin_border

    column_widths = {'A': 5, 'B': 15, 'C': 40, 'D': 8, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 8, 'L': 15}
    for col, width in column_widths.items(): ws.column_dimensions[col].width = width

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal para usuarios autenticados"""
    # Obtener estadísticas para mostrar en el dashboard
    user_count = User.query.count()
    admin_count = User.query.filter_by(rol='admin').count()
    jefe_count = User.query.filter_by(rol='jefe_carrera').count()
    profesor_completo_count = User.query.filter_by(rol='profesor_completo').count()
    profesor_asignatura_count = User.query.filter_by(rol='profesor_asignatura').count()
    profesor_count = profesor_completo_count + profesor_asignatura_count
    
    return render_template('dashboard.html',
                         user_count=user_count,
                         admin_count=admin_count,
                         jefe_count=jefe_count,
                         profesor_count=profesor_count)

@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    name = current_user.get_nombre_completo()
    logout_user()
    flash(f'¡Hasta luego, {name}!', 'info')
    return redirect(url_for('index'))

# Rutas para diferentes roles (ejemplos básicos)
@app.route('/admin')
@login_required
def admin_panel():
    """Panel de administración - solo para administradores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    return render_template('admin/panel.html', users=users)

# @app.route('/jefe-carrera')
# @login_required
# def jefe_carrera_panel():
#     """Panel para jefes de carrera"""
#     if not current_user.is_jefe_carrera():
#         flash('No tienes permisos para acceder a esta página.', 'error')
#         return redirect(url_for('dashboard'))
#     
#     # Verificar que el jefe tenga una carrera asignada
#     if not current_user.carreras:
#         flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
#         return redirect(url_for('dashboard'))
#     
#     # Obtener datos específicos de la carrera del jefe
#     profesores = current_user.get_profesores_carrera()
#     materias = current_user.get_materias_carrera()
#     horarios_academicos = current_user.get_horarios_academicos_carrera()
#     
#     return render_template('jefe/panel.html', 
#                          profesores=profesores, 
#                          materias=materias,
#                          horarios_academicos=horarios_academicos,
#                          carrera=current_user.carrera)

# ==========================================
# GESTIÓN DE PROFESORES PARA JEFES DE CARRERA
# ==========================================

@app.route('/jefe-carrera/profesores')
@login_required
def gestionar_profesores_jefe():
    """Gestión de profesores para jefes de carrera (solo sus carreras)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    if not current_user.carreras:
        flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
        return redirect(url_for('dashboard'))
    
    profesores = current_user.get_profesores_carrera()
    # Pasar la primera carrera para compatibilidad, pero el jefe puede tener múltiples
    return render_template('jefe/profesores.html', profesores=profesores, carrera=current_user.carreras[0] if current_user.carreras else None, carreras=current_user.carreras)

@app.route('/jefe-carrera/profesor/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_profesor_jefe(id):
    """Editar profesor para jefes de carrera (solo de sus carreras)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Obtener IDs de carreras del jefe
    jefe_carrera_ids = [c.id for c in current_user.carreras]
    
    # Verificar que el profesor pertenezca a alguna carrera del jefe
    if profesor.is_profesor():
        if not any(c.id in jefe_carrera_ids for c in profesor.carreras):
            flash('No tienes permisos para editar este profesor.', 'error')
            return redirect(url_for('gestionar_profesores_jefe'))
    # Para jefes de carrera: verificar que comparta al menos una carrera
    elif profesor.is_jefe_carrera():
        if not any(c.id in jefe_carrera_ids for c in profesor.carreras):
            flash('No tienes permisos para editar este jefe de carrera.', 'error')
            return redirect(url_for('gestionar_profesores_jefe'))
    
    # Obtener horarios para mostrar en la tabla de disponibilidad
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    form = EditarUsuarioForm(user=profesor)
    
    if form.validate_on_submit():
        profesor.username = form.username.data
        profesor.nombre = form.nombre.data
        profesor.apellido = form.apellido.data
        profesor.email = form.email.data
        profesor.telefono = form.telefono.data
        profesor.rol = form.rol.data
        profesor.activo = form.activo.data
        
        # Procesar disponibilidad horaria para profesores
        if profesor.is_profesor():
            # Desactivar disponibilidades anteriores (mantener historial)
            DisponibilidadProfesor.query.filter_by(
                profesor_id=profesor.id,
                activo=True
            ).update({'activo': False})
            
            # Crear nuevas disponibilidades basadas en los checkboxes MARCADOS
            # Solo guardamos los horarios donde el profesor SÍ está disponible
            dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
            for horario in horarios:
                for dia in dias:
                    field_name = f'disp_{horario.id}_{dia}'
                    # Si el checkbox está marcado (existe en el form), crear disponibilidad
                    if request.form.get(field_name):
                        nueva_disponibilidad = DisponibilidadProfesor(
                            profesor_id=profesor.id,
                            horario_id=horario.id,
                            dia_semana=dia,
                            disponible=True,
                            creado_por=current_user.id
                        )
                        db.session.add(nueva_disponibilidad)
        
        db.session.commit()
        flash('Profesor actualizado exitosamente.', 'success')
        return redirect(url_for('gestionar_profesores_jefe'))
    elif request.method == 'POST':
        # Si hay errores de validación, mostrarlos
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error en {field}: {error}', 'error')
    
    # Pre-llenar el formulario
    if request.method == 'GET':
        form.username.data = profesor.username
        form.nombre.data = profesor.nombre
        form.apellido.data = profesor.apellido
        form.email.data = profesor.email
        form.telefono.data = profesor.telefono
        form.rol.data = profesor.rol
        form.activo.data = profesor.activo
        if profesor.carreras:
            form.carreras.data = [c.id for c in profesor.carreras]
    
    # Cargar disponibilidades actuales del profesor
    disponibilidad_dict = {}
    if profesor.is_profesor():
        disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
            profesor_id=profesor.id,
            activo=True
        ).all()
        
        for disp in disponibilidades_actuales:
            if disp.disponible:
                disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True
    
    return render_template('jefe/editar_profesor.html', 
                         form=form, 
                         profesor=profesor,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict)

@app.route('/jefe-carrera/profesor/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_profesor_jefe(id):
    """Eliminar profesor para jefes de carrera (solo de su carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que el profesor pertenezca a la carrera del jefe
    # Para profesores: verificar si está asignado a la carrera del jefe
    if profesor.is_profesor():
        if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
            flash('No tienes permisos para eliminar este profesor.', 'error')
            return redirect(url_for('gestionar_profesores_jefe'))
    # Para jefes de carrera: verificar que sea de la misma carrera
    elif profesor.is_jefe_carrera():
        if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
            flash('No tienes permisos para eliminar este jefe de carrera.', 'error')
            return redirect(url_for('gestionar_profesores_jefe'))
    
    # Desactivar en lugar de eliminar
    profesor.activo = False
    db.session.commit()
    
    flash('Profesor eliminado exitosamente.', 'success')
    return redirect(url_for('gestionar_profesores_jefe'))

@app.route('/jefe-carrera/profesor/<int:id>/materias', methods=['GET', 'POST'])
@login_required
def gestionar_materias_profesor_jefe(id):
    """Asignar/modificar materias de un profesor (jefe de carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que el profesor pertenezca a la carrera del jefe
    if profesor.is_profesor():
        if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
            flash('No tienes permisos para gestionar este profesor.', 'error')
            return redirect(url_for('gestionar_profesores_jefe'))
    else:
        flash('Este usuario no es un profesor.', 'error')
        return redirect(url_for('gestionar_profesores_jefe'))
    
    from forms import AsignarMateriasProfesorForm
    form = AsignarMateriasProfesorForm(profesor=profesor)
    
    # Filtrar solo materias de la carrera del jefe
    materias_carrera = Materia.query.filter_by(
        carrera_id=current_user.primera_carrera_id,
        activa=True
    ).order_by(Materia.cuatrimestre, Materia.nombre).all()
    
    form.materias.choices = [
        (m.id, f"Cuatri {m.cuatrimestre} - {m.codigo}: {m.nombre}")
        for m in materias_carrera
    ]
    
    if form.validate_on_submit():
        try:
            # Obtener materias seleccionadas
            materias_ids = form.materias.data
            materias_nuevas = Materia.query.filter(Materia.id.in_(materias_ids)).all()
            
            # Actualizar materias del profesor
            profesor.materias = materias_nuevas
            db.session.commit()
            
            flash(f'Materias actualizadas exitosamente para {profesor.get_nombre_completo()}. '
                  f'Total: {len(materias_nuevas)} materias asignadas.', 'success')
            return redirect(url_for('gestionar_profesores_jefe'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al asignar materias: {str(e)}', 'error')
    
    # Precargar materias actuales del profesor
    elif request.method == 'GET':
        form.materias.data = [m.id for m in profesor.materias]
    
    return render_template('jefe/asignar_materias_profesor.html', 
                         form=form, 
                         profesor=profesor,
                         titulo=f"Asignar Materias - {profesor.get_nombre_completo()}")

@app.route('/jefe-carrera/profesor/<int:id>/materias/ver')
@login_required
def ver_materias_profesor_jefe(id):
    """Ver materias asignadas a un profesor (jefe de carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que el profesor pertenezca a la carrera del jefe
    if profesor.is_profesor():
        if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
            flash('No tienes permisos para ver este profesor.', 'error')
            return redirect(url_for('gestionar_profesores_jefe'))
    else:
        flash('Este usuario no es un profesor.', 'error')
        return redirect(url_for('gestionar_profesores_jefe'))
    
    # Obtener materias ordenadas por cuatrimestre
    materias = sorted(profesor.materias, key=lambda m: (m.cuatrimestre, m.nombre))
    
    # Obtener horarios del profesor agrupados por materia
    horarios_por_materia = {}
    for materia in materias:
        horarios = HorarioAcademico.query.filter_by(
            profesor_id=profesor.id,
            materia_id=materia.id,
            activo=True
        ).all()
        horarios_por_materia[materia.id] = horarios
    
    return render_template('jefe/ver_materias_profesor.html',
                         profesor=profesor,
                         materias=materias,
                         horarios_por_materia=horarios_por_materia)

# ==========================================
# GESTIÓN DE DISPONIBILIDAD DE PROFESORES (JEFE DE CARRERA)
# ==========================================

@app.route('/jefe-carrera/profesores/disponibilidad')
@login_required
def disponibilidad_profesores_jefe():
    """Módulo de disponibilidad horaria de profesores para jefes de carrera"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    if not current_user.carreras:
        flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Obtener profesores de la carrera del jefe
    profesores = User.query.filter(
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True,
        User.carreras.any(Carrera.id.in_(current_user.get_carreras_jefe_ids()))
    ).order_by(User.apellido, User.nombre).all()
    
    # Calcular estadísticas de disponibilidad
    total_profesores = len(profesores)
    profesores_con_disponibilidad = 0
    
    for profesor in profesores:
        disponibilidades = DisponibilidadProfesor.query.filter_by(
            profesor_id=profesor.id,
            activo=True
        ).count()
        if disponibilidades > 0:
            profesores_con_disponibilidad += 1
    
    return render_template('jefe/disponibilidad_profesores.html',
                         profesores=profesores,
                         carrera=current_user.carrera,
                         total_profesores=total_profesores,
                         profesores_con_disponibilidad=profesores_con_disponibilidad)

@app.route('/jefe-carrera/profesor/<int:id>/disponibilidad/editar', methods=['GET', 'POST'])
@login_required
def editar_disponibilidad_profesor_jefe(id):
    """Editar disponibilidad horaria de un profesor (jefe de carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que el profesor pertenezca a la carrera del jefe
    if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
        flash('No tienes permisos para editar la disponibilidad de este profesor.', 'error')
        return redirect(url_for('disponibilidad_profesores_jefe'))
    
    # Obtener horarios del sistema
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    if request.method == 'POST':
        try:
            # Procesar disponibilidad
            dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
            
            # Desactivar disponibilidades anteriores
            DisponibilidadProfesor.query.filter_by(
                profesor_id=profesor.id,
                activo=True
            ).update({'activo': False})
            
            # Crear nuevas disponibilidades basadas en los checkboxes marcados
            for horario in horarios:
                for dia in dias:
                    field_name = f"disp_{horario.id}_{dia}"
                    if request.form.get(field_name):
                        nueva_disponibilidad = DisponibilidadProfesor(
                            profesor_id=profesor.id,
                            horario_id=horario.id,
                            dia_semana=dia,
                            disponible=True,
                            creado_por=current_user.id
                        )
                        db.session.add(nueva_disponibilidad)
            
            db.session.commit()
            flash(f'Disponibilidad de {profesor.get_nombre_completo()} actualizada exitosamente.', 'success')
            return redirect(url_for('disponibilidad_profesores_jefe'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar disponibilidad: {str(e)}', 'error')
    
    # Cargar disponibilidades actuales
    disponibilidad_dict = {}
    disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
        profesor_id=profesor.id,
        activo=True
    ).all()
    
    for disp in disponibilidades_actuales:
        if disp.disponible:
            disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True
    
    return render_template('jefe/editar_disponibilidad_profesor.html',
                         profesor=profesor,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict,
                         carrera=current_user.carrera)

@app.route('/jefe-carrera/profesor/<int:id>/disponibilidad/ver')
@login_required
def ver_disponibilidad_profesor_jefe(id):
    """Ver disponibilidad horaria de un profesor (jefe de carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que el profesor pertenezca a la carrera del jefe
    if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
        flash('No tienes permisos para ver la disponibilidad de este profesor.', 'error')
        return redirect(url_for('disponibilidad_profesores_jefe'))
    
    # Obtener horarios del sistema
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    # Cargar disponibilidades actuales
    disponibilidad_dict = {}
    disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
        profesor_id=profesor.id,
        activo=True
    ).all()
    
    for disp in disponibilidades_actuales:
        if disp.disponible:
            disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True
    
    # Calcular total de horas disponibles
    total_horas_disponibles = len(disponibilidades_actuales)
    
    return render_template('jefe/ver_disponibilidad_profesor.html',
                         profesor=profesor,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict,
                         carrera=current_user.carrera,
                         total_horas_disponibles=total_horas_disponibles)

# ==========================================
# ASIGNACIÓN MASIVA DE MATERIAS (JEFE DE CARRERA)
# ==========================================
@app.route('/jefe-carrera/asignacion-masiva-materias', methods=['GET', 'POST'])
@login_required
def asignacion_masiva_materias_jefe():
    """Asignación masiva de materias a múltiples profesores (jefe de carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    if not current_user.carreras:
        flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            asignaciones_marcadas = set(request.form.getlist('asignaciones[]'))
            
            # Obtener filtros para reconstruir el estado original
            cuatrimestre = request.args.get('cuatrimestre', type=int)
            
            # Obtener todos los profesores y materias según filtros
            profesores = User.query.filter(
                User.rol.in_(['profesor_completo', 'profesor_asignatura']),
                User.activo == True,
                User.carreras.any(Carrera.id.in_(current_user.get_carreras_jefe_ids()))
            ).all()
            
            materias_query = Materia.query.filter_by(
                carrera_id=current_user.primera_carrera_id,
                activa=True
            )
            if cuatrimestre:
                materias_query = materias_query.filter_by(cuatrimestre=cuatrimestre)
            materias = materias_query.all()
            
            # Procesar cambios
            contador_asignaciones = 0
            contador_desasignaciones = 0
            errores = []
            
            # Iterar sobre todos los profesores y materias visibles
            for profesor in profesores:
                for materia in materias:
                    clave = f"{profesor.id}-{materia.id}"
                    esta_marcada = clave in asignaciones_marcadas
                    esta_asignada = materia in profesor.materias
                    
                    # Si está marcada y no está asignada -> ASIGNAR
                    if esta_marcada and not esta_asignada:
                        profesor.materias.append(materia)
                        contador_asignaciones += 1
                    
                    # Si NO está marcada y SÍ está asignada -> DESASIGNAR
                    elif not esta_marcada and esta_asignada:
                        profesor.materias.remove(materia)
                        contador_desasignaciones += 1
            
            # Guardar cambios
            db.session.commit()
            
            # Mensaje de éxito
            mensajes = []
            if contador_asignaciones > 0:
                mensajes.append(f'{contador_asignaciones} nueva(s) asignación(es)')
            if contador_desasignaciones > 0:
                mensajes.append(f'{contador_desasignaciones} desasignación(es)')
            
            if mensajes:
                flash(f'Cambios realizados exitosamente: {", ".join(mensajes)}.', 'success')
            else:
                flash('No se realizaron cambios.', 'info')
            
            if errores:
                flash(f'Errores: {", ".join(errores[:5])}', 'warning')
            
            return redirect(url_for('asignacion_masiva_materias_jefe'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al realizar las asignaciones: {str(e)}', 'error')
            return redirect(url_for('asignacion_masiva_materias_jefe'))
    
    # GET - Mostrar formulario
    # Obtener filtros
    cuatrimestre = request.args.get('cuatrimestre', type=int)
    
    # Obtener profesores activos de la carrera del jefe
    profesores = User.query.filter(
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True,
        User.carreras.any(Carrera.id.in_(current_user.get_carreras_jefe_ids()))
    ).order_by(User.apellido, User.nombre).all()
    
    # Obtener materias activas de la carrera del jefe
    materias_query = Materia.query.filter_by(
        carrera_id=current_user.primera_carrera_id,
        activa=True
    )
    
    if cuatrimestre:
        materias_query = materias_query.filter_by(cuatrimestre=cuatrimestre)
    
    materias = materias_query.order_by(Materia.cuatrimestre, Materia.nombre).all()
    
    # Crear matriz de asignaciones actuales
    asignaciones_actuales = {}
    for profesor in profesores:
        asignaciones_actuales[profesor.id] = set(m.id for m in profesor.materias)
    
    # Calcular carga de trabajo de cada profesor
    cargas_profesores = {}
    for profesor in profesores:
        cargas_profesores[profesor.id] = calcular_carga_profesor(profesor)
        
    # Obtener filtro de disponibilidad
    solo_disponibles = request.args.get('solo_disponibles', type=int, default=0)
    
    # Si se solicita filtrar solo profesores disponibles, filtrar
    if solo_disponibles:
        profesores = [p for p in profesores if cargas_profesores[p.id]['puede_mas']]
    
    return render_template('jefe/asignacion_masiva_materias.html',
                         profesores=profesores,
                         materias=materias,
                         asignaciones_actuales=asignaciones_actuales,
                         cargas_profesores=cargas_profesores,
                         filtro_cuatrimestre=cuatrimestre,
                         solo_disponibles=solo_disponibles,
                         carrera=current_user.carrera)

# ==========================================
# GESTIÓN DE MATERIAS PARA JEFES DE CARRERA
# ==========================================

@app.route('/jefe-carrera/materias')
@login_required
def gestionar_materias_jefe():
    """Gestión de materias para jefes de carrera (solo su carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    if not current_user.carreras:
        flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Obtener filtros
    ciclo_str = request.args.get('ciclo', type=str)
    ciclo = int(ciclo_str) if ciclo_str and ciclo_str != '' else None
    cuatrimestre_str = request.args.get('cuatrimestre', type=str)
    cuatrimestre = int(cuatrimestre_str) if cuatrimestre_str and cuatrimestre_str != '' else None
    
    # Query base para materias de la carrera
    query = Materia.query.filter(
        Materia.carrera_id.in_(current_user.get_carreras_jefe_ids()),
        Materia.activa == True
    )
    
    # Filtro por ciclo escolar
    if ciclo:
        if ciclo == 1:
            query = query.filter(Materia.cuatrimestre % 3 == 1)
        elif ciclo == 2:
            query = query.filter(Materia.cuatrimestre % 3 == 2)
        elif ciclo == 3:
            query = query.filter(Materia.cuatrimestre % 3 == 0)
    
    # Filtro por cuatrimestre específico
    if cuatrimestre is not None:
        query = query.filter(Materia.cuatrimestre == cuatrimestre)
    
    materias = query.order_by(Materia.cuatrimestre, Materia.nombre).all()
    
    return render_template('jefe/materias.html', 
                         materias=materias, 
                         carrera=current_user.carrera,
                         filtros_activos={
                             'ciclo': ciclo,
                             'cuatrimestre': cuatrimestre
                         })

@app.route('/jefe-carrera/materia/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_materia_jefe(id):
    """Editar materia para jefes de carrera (solo de su carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    materia = Materia.query.get_or_404(id)
    
    # Verificar que la materia pertenezca a la carrera del jefe
    if not current_user.puede_acceder_carrera(materia.carrera_id):
        flash('No tienes permisos para editar esta materia.', 'error')
        return redirect(url_for('gestionar_materias_jefe'))
    
    form = MateriaForm()
    
    # Pre-llenar el formulario solo en GET
    if request.method == 'GET':
        form.nombre.data = materia.nombre
        form.codigo.data = materia.codigo
        form.cuatrimestre.data = materia.cuatrimestre
        form.creditos.data = materia.creditos
        form.horas_semanales.data = materia.horas_semanales
        form.descripcion.data = materia.descripcion
        form.carrera.data = str(materia.carrera_id)
    
    if form.validate_on_submit():
        materia.nombre = form.nombre.data
        materia.codigo = form.codigo.data.upper()
        materia.descripcion = form.descripcion.data
        materia.cuatrimestre = form.cuatrimestre.data
        materia.creditos = form.creditos.data
        materia.horas_semanales = form.horas_semanales.data
        materia.carrera_id = int(form.carrera.data)
        
        db.session.commit()
        flash('Materia actualizada exitosamente.', 'success')
        return redirect(url_for('gestionar_materias_jefe'))
    
    return render_template('jefe/editar_materia.html', form=form, materia=materia)

# ==========================================
# GESTIÓN DE HORARIOS ACADÉMICOS PARA JEFES DE CARRERA
# ==========================================

@app.route('/jefe-carrera/horarios-academicos')
@login_required
def gestionar_horarios_academicos_jefe():
    """Gestión de horarios académicos para jefes de carrera (solo su carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    if not current_user.carreras:
        flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
        return redirect(url_for('dashboard'))
    
    horarios_academicos = current_user.get_horarios_academicos_carrera()
    
    # Ordenar horarios: primero por día (Lunes=0, ...), luego por orden de horario
    dias_order = {
        'lunes': 0, 'martes': 1, 'miercoles': 2, 'jueves': 3, 'viernes': 4, 'sabado': 5, 'domingo': 6
    }
    # Helper para ordenamiento natural (1MSC vs 10MSC)
    import re
    def natural_sort_key(s):
        """Ordena cadenas que contienen números de forma natural (1, 2, 10)"""
        if not s:
            return []
        return [int(text) if text.isdigit() else text.lower()
                for text in re.split(r'(\d+)', s)]

    # Ordenar horarios: 
    # 1. Por día (Lunes, Martes...)
    # 2. Por grupo (naturalmente: 1MSC, 10MSC...)
    # 3. Por hora de inicio (7:00, 8:00...)
    horarios_academicos.sort(key=lambda x: (
        dias_order.get(x.dia_semana, 100), 
        natural_sort_key(x.grupo),
        x.horario.hora_inicio
    ))

    # Obtener grupos únicos y ordenarlos naturalmente
    grupos_unicos = list(set(h.grupo for h in horarios_academicos))
    grupos_unicos.sort(key=natural_sort_key)

    return render_template('jefe/horarios_academicos.html', 
                         horarios_academicos=horarios_academicos, 
                         grupos_unicos=grupos_unicos,
                         carrera=current_user.carrera)

@app.route('/jefe-carrera/horario-academico/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_horario_academico_jefe(id):
    """Editar horario académico para jefes de carrera (solo de su carrera)"""
    if not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    horario_academico = HorarioAcademico.query.get_or_404(id)
    
    # Verificar que el horario pertenezca a la carrera del jefe
    # Verificar si el profesor del horario está asignado a la carrera del jefe
    profesor = horario_academico.profesor
    if profesor.is_profesor():
        if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
            flash('No tienes permisos para editar este horario académico.', 'error')
            return redirect(url_for('gestionar_horarios_academicos_jefe'))
    elif profesor.is_jefe_carrera():
        if not any(current_user.puede_acceder_carrera(c.id) for c in profesor.carreras):
            flash('No tienes permisos para editar este horario académico.', 'error')
            return redirect(url_for('gestionar_horarios_academicos_jefe'))
        flash('No tienes permisos para editar este horario académico.', 'error')
        return redirect(url_for('gestionar_horarios_academicos_jefe'))
    
    form = EditarHorarioAcademicoForm()
    if form.validate_on_submit():
        horario_academico.horario_id = form.horario_id.data
        horario_academico.dia_semana = form.dia_semana.data
        horario_academico.grupo = form.grupo.data
        
        db.session.commit()
        flash('Horario académico actualizado exitosamente.', 'success')
        return redirect(url_for('gestionar_horarios_academicos_jefe'))
    
    # Pre-llenar el formulario
    form.horario_id.data = horario_academico.horario_id
    form.dia_semana.data = horario_academico.dia_semana
    form.grupo.data = horario_academico.grupo
    
    return render_template('jefe/editar_horario_academico.html', 
                         form=form, 
                         horario_academico=horario_academico)

@app.route('/profesor')
@login_required
def profesor_panel():
    """Panel para profesores"""
    if not current_user.is_profesor():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('profesor/panel.html')

@app.route('/profesor/horarios')
@login_required
def profesor_horarios():
    """Ver horarios asignados al profesor"""
    if not current_user.is_profesor():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener horarios académicos del profesor actual
    horarios = HorarioAcademico.query.filter_by(
        profesor_id=current_user.id,
        activo=True
    ).join(Horario).order_by(
        Horario.orden,
        HorarioAcademico.dia_semana
    ).all()
    
    # Organizar horarios por día de la semana
    dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
    horarios_por_dia = {}
    
    for dia in dias_semana:
        horarios_por_dia[dia] = [h for h in horarios if h.dia_semana == dia]
    
    # Estadísticas del profesor
    total_horas = len(horarios)
    materias_unicas = len(set(h.materia_id for h in horarios))
    
    return render_template('profesor/horarios.html', 
                         horarios_por_dia=horarios_por_dia,
                         total_horas=total_horas,
                         materias_unicas=materias_unicas,
                         dias_semana=dias_semana)

@app.route('/profesor/mis-materias')
@login_required
def profesor_mis_materias():
    """Ver materias asignadas al profesor (solo lectura)"""
    if not current_user.is_profesor():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener materias ordenadas por cuatrimestre
    materias = sorted(current_user.materias, key=lambda m: (m.cuatrimestre, m.nombre))
    
    # Obtener horarios del profesor agrupados por materia
    horarios_por_materia = {}
    for materia in materias:
        horarios = HorarioAcademico.query.filter_by(
            profesor_id=current_user.id,
            materia_id=materia.id,
            activo=True
        ).all()
        horarios_por_materia[materia.id] = horarios
    
    # Estadísticas
    total_materias = len(materias)
    total_horas_clase = sum(len(horarios_por_materia.get(m.id, [])) for m in materias)
    
    # Agrupar materias por cuatrimestre
    materias_por_cuatrimestre = {}
    for materia in materias:
        if materia.cuatrimestre not in materias_por_cuatrimestre:
            materias_por_cuatrimestre[materia.cuatrimestre] = []
        materias_por_cuatrimestre[materia.cuatrimestre].append(materia)
    
    return render_template('profesor/mis_materias.html',
                         materias=materias,
                         horarios_por_materia=horarios_por_materia,
                         total_materias=total_materias,
                         total_horas_clase=total_horas_clase,
                         materias_por_cuatrimestre=materias_por_cuatrimestre)

@app.route('/profesor/disponibilidad', methods=['GET', 'POST'])
@login_required
def profesor_disponibilidad():
    """Gestionar disponibilidad horaria del profesor"""
    if not current_user.is_profesor():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
    
    # Obtener todos los horarios disponibles del sistema (matutinos primero, luego vespertinos)
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    if request.method == 'POST':
        try:
            # Eliminar disponibilidades anteriores (pero mantener inactivas para historial)
            DisponibilidadProfesor.query.filter_by(
                profesor_id=current_user.id,
                activo=True
            ).update({DisponibilidadProfesor.activo: False})
            
            # Procesar los datos del formulario
            disponibilidades_creadas = 0
            for dia in dias_semana:
                for horario in horarios:
                    checkbox_name = f'disponible_{dia}_{horario.id}'
                    # Si el checkbox está marcado, significa que SÍ está disponible
                    disponible = checkbox_name in request.form
                    
                    # Crear registro de disponibilidad
                    nueva_disponibilidad = DisponibilidadProfesor(
                        profesor_id=current_user.id,
                        horario_id=horario.id,
                        dia_semana=dia,
                        disponible=disponible,
                        creado_por=current_user.id
                    )
                    db.session.add(nueva_disponibilidad)
                    disponibilidades_creadas += 1
            
            db.session.commit()
            flash(f'✅ Disponibilidad actualizada correctamente. Se crearon {disponibilidades_creadas} registros.', 'success')
            return redirect(url_for('profesor_disponibilidad'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error al guardar la disponibilidad: {str(e)}', 'error')
            print(f"Error al guardar disponibilidad: {e}")
    
    # GET: Obtener disponibilidades actuales del profesor
    disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
        profesor_id=current_user.id,
        activo=True
    ).all()
    
    # Crear diccionario para acceso rápido
    disponibilidad_dict = {}
    for disp in disponibilidades_actuales:
        key = f"{disp.dia_semana}_{disp.horario_id}"
        disponibilidad_dict[key] = disp.disponible
    
    # Si no hay disponibilidades, asumir que está disponible en todo (para facilitar primer uso)
    tiene_disponibilidades = len(disponibilidades_actuales) > 0
    
    return render_template('profesor/disponibilidad.html',
                         dias_semana=dias_semana,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict,
                         tiene_disponibilidades=tiene_disponibilidades)

# ==================== RUTAS DE GESTIÓN DE GRUPOS (ADMIN) ====================

@app.route('/admin/grupos')
@login_required
def gestionar_grupos():
    """Gestión de grupos - para administradores y jefes de carrera"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Si es admin, mostrar TODOS los grupos (incluso si también es jefe de carrera)
    if current_user.is_admin():
        grupos = Grupo.query.join(Carrera).order_by(
            Carrera.nombre, Grupo.cuatrimestre, Grupo.numero_grupo
        ).all()
        
        # Obtener todas las carreras activas
        carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    # Si es solo jefe de carrera (sin ser admin), mostrar grupos de su carrera
    elif current_user.is_jefe_carrera():
        if not current_user.carreras:
            flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
            return redirect(url_for('dashboard'))
        
        grupos = Grupo.query.filter_by(carrera_id=current_user.primera_carrera_id).join(Carrera).order_by(
            Grupo.cuatrimestre, Grupo.numero_grupo
        ).all()
        
        # Obtener carreras únicas (solo la del jefe)
        carreras = Carrera.query.filter_by(id=current_user.primera_carrera_id).all()

    
    return render_template('admin/grupos.html', grupos=grupos, carreras=carreras)

@app.route('/admin/grupo/nuevo', methods=['GET', 'POST'])
@login_required
def crear_grupo():
    """Crear nuevo grupo - para administradores y jefes de carrera"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = GrupoForm()
    
    # Si es SOLO jefe de carrera (y NO admin), restringir a su carrera
    # Los admins pueden modificar cualquier carrera sin restricciones
    es_solo_jefe = current_user.is_jefe_carrera() and not current_user.is_admin()
    
    if es_solo_jefe:
        if not current_user.carreras:
            flash('No tienes carreras asignadas. Contacta al administrador.', 'warning')
            return redirect(url_for('dashboard'))
        
        # Pre-seleccionar la carrera del jefe y deshabilitar el selector
        form.carrera.data = current_user.primera_carrera_id
    
    if form.validate_on_submit():
        # Verificar que se haya seleccionado una carrera válida
        if form.carrera.data == 0:
            flash('Debe seleccionar una carrera.', 'error')
            return render_template('admin/grupo_form.html', form=form, titulo='Crear Grupo')
        
        # Verificar que jefe de carrera (no admin) solo cree grupos de su carrera
        if es_solo_jefe and not current_user.tiene_carrera(form.carrera.data):
            flash('Solo puedes crear grupos de tu carrera.', 'error')
            return redirect(url_for('gestionar_grupos'))
        
        # Verificar que se haya seleccionado un cuatrimestre válido (0-10 son válidos, -1 no)
        if form.cuatrimestre.data == -1:
            flash('Debe seleccionar un cuatrimestre.', 'error')
            return render_template('admin/grupo_form.html', form=form, titulo='Crear Grupo')
        
        # Verificar si ya existe un grupo con la misma configuración
        grupo_existente = Grupo.query.filter_by(
            numero_grupo=form.numero_grupo.data,
            turno=form.turno.data,
            carrera_id=form.carrera.data,
            cuatrimestre=form.cuatrimestre.data
        ).first()
        
        if grupo_existente:
            flash('Ya existe un grupo con esta configuración.', 'error')
        else:
            grupo = Grupo(
                numero_grupo=form.numero_grupo.data,
                turno=form.turno.data,
                carrera_id=form.carrera.data,
                cuatrimestre=form.cuatrimestre.data,
                creado_por=current_user.id
            )
            db.session.add(grupo)
            db.session.commit()
            flash(f'Grupo {grupo.codigo} creado exitosamente.', 'success')
            return redirect(url_for('gestionar_grupos'))
    
    return render_template('admin/grupo_form.html', form=form, titulo='Crear Grupo', 
                         es_jefe=es_solo_jefe)

@app.route('/admin/grupo/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_grupo(id):
    """Editar grupo existente - para administradores y jefes de carrera"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    grupo = Grupo.query.get_or_404(id)
    
    # Si es SOLO jefe de carrera (y NO admin), restringir a su carrera
    # Los admins pueden modificar cualquier carrera sin restricciones
    es_solo_jefe = current_user.is_jefe_carrera() and not current_user.is_admin()
    
    # Verificar acceso para jefes de carrera (no aplica a admins)
    if es_solo_jefe:
        if not current_user.tiene_carrera(grupo.carrera_id):
            flash('No tienes permisos para editar este grupo.', 'error')
            return redirect(url_for('gestionar_grupos'))
    
    form = GrupoForm()
    
    if form.validate_on_submit():
        # Verificar que se haya seleccionado una carrera válida
        if form.carrera.data == 0:
            flash('Debe seleccionar una carrera.', 'error')
            return render_template('admin/grupo_form.html', form=form, grupo=grupo, titulo=f'Editar Grupo {grupo.codigo}')
        
        # Verificar que jefe de carrera (no admin) solo edite grupos de su carrera
        if es_solo_jefe and not current_user.tiene_carrera(form.carrera.data):
            flash('Solo puedes editar grupos de tu carrera.', 'error')
            return redirect(url_for('gestionar_grupos'))
        
        # Verificar que se haya seleccionado un cuatrimestre válido (0-10 son válidos, -1 no)
        if form.cuatrimestre.data == -1:
            flash('Debe seleccionar un cuatrimestre.', 'error')
            return render_template('admin/grupo_form.html', form=form, grupo=grupo, titulo=f'Editar Grupo {grupo.codigo}')
        
        # Verificar si ya existe otro grupo con la misma configuración
        grupo_existente = Grupo.query.filter(
            Grupo.id != id,
            Grupo.numero_grupo == form.numero_grupo.data,
            Grupo.turno == form.turno.data,
            Grupo.carrera_id == form.carrera.data,
            Grupo.cuatrimestre == form.cuatrimestre.data
        ).first()
        
        if grupo_existente:
            flash('Ya existe otro grupo con esta configuración.', 'error')
        else:
            grupo.numero_grupo = form.numero_grupo.data
            grupo.turno = form.turno.data
            grupo.carrera_id = form.carrera.data
            grupo.cuatrimestre = form.cuatrimestre.data
            grupo.codigo = grupo.generar_codigo()  # Regenerar código
            
            db.session.commit()
            flash(f'Grupo {grupo.codigo} actualizado exitosamente.', 'success')
            return redirect(url_for('gestionar_grupos'))
    
    # Pre-llenar el formulario
    if request.method == 'GET':
        form.numero_grupo.data = grupo.numero_grupo
        form.turno.data = grupo.turno
        form.carrera.data = grupo.carrera_id
        form.cuatrimestre.data = grupo.cuatrimestre
    
    return render_template('admin/grupo_form.html', form=form, grupo=grupo, 
                         titulo=f'Editar Grupo {grupo.codigo}',
                         es_jefe=es_solo_jefe)


@app.route('/admin/grupo/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_grupo(id):
    """Eliminar grupo - para administradores y jefes de carrera"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    grupo = Grupo.query.get_or_404(id)
    
    # Si es jefe de carrera, verificar que el grupo pertenezca a su carrera
    if current_user.is_jefe_carrera():
        if not current_user.tiene_carrera(grupo.carrera_id):
            flash('No tienes permisos para eliminar este grupo.', 'error')
            return redirect(url_for('gestionar_grupos'))
    
    # Verificar si tiene horarios asignados (cuando se implemente la relación)
    # Por ahora solo eliminar
    codigo = grupo.codigo
    db.session.delete(grupo)
    db.session.commit()
    
    flash(f'Grupo {codigo} eliminado exitosamente.', 'success')
    return redirect(url_for('gestionar_grupos'))

@app.route('/admin/grupo/<int:id>/materias', methods=['GET', 'POST'])
@login_required
def gestionar_materias_grupo(id):
    """Asignar materias a un grupo - para administradores y jefes de carrera"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    grupo = Grupo.query.get_or_404(id)
    
    # Si es jefe de carrera, verificar que el grupo pertenezca a su carrera
    if current_user.is_jefe_carrera():
        if not current_user.tiene_carrera(grupo.carrera_id):
            flash('No tienes permisos para gestionar las materias de este grupo.', 'error')
            return redirect(url_for('gestionar_grupos'))
    
    form = AsignarMateriasGrupoForm(grupo=grupo)
    
    if form.validate_on_submit():
        # Limpiar materias actuales y asignar nuevas
        grupo.materias = []
        materias_seleccionadas = Materia.query.filter(Materia.id.in_(form.materias.data)).all()
        grupo.materias = materias_seleccionadas
        
        db.session.commit()
        flash(f'{len(materias_seleccionadas)} materias asignadas al grupo {grupo.codigo} exitosamente.', 'success')
        return redirect(url_for('ver_materias_grupo', id=grupo.id))
    
    # Pre-seleccionar materias ya asignadas
    if request.method == 'GET':
        form.materias.data = [m.id for m in grupo.materias]
    
    return render_template('admin/asignar_materias_grupo.html', form=form, grupo=grupo)

@app.route('/admin/grupo/<int:id>/materias/ver')
@login_required
def ver_materias_grupo(id):
    """Ver materias asignadas a un grupo - para administradores y jefes de carrera"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    grupo = Grupo.query.get_or_404(id)
    
    # Si es jefe de carrera, verificar que el grupo pertenezca a su carrera
    if current_user.is_jefe_carrera():
        if not current_user.tiene_carrera(grupo.carrera_id):
            flash('No tienes permisos para ver las materias de este grupo.', 'error')
            return redirect(url_for('gestionar_grupos'))
    
    return render_template('admin/ver_materias_grupo.html', grupo=grupo)

@app.route('/admin/grupos/asignacion-masiva-materias', methods=['GET', 'POST'])
@login_required
def asignacion_masiva_materias_grupos():
    """Asignación masiva de materias a grupos - vista matricial"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener filtros
    carrera_id = request.args.get('carrera_id', type=int)
    cuatrimestre = request.args.get('cuatrimestre', type=int)
    
    # Si es jefe de carrera, filtrar por su carrera
    if current_user.is_jefe_carrera():
        carrera_id = current_user.primera_carrera_id
    
    # Obtener carreras para el filtro
    if current_user.is_admin():
        carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    else:
        carreras = [current_user.carrera]
    
    grupos = []
    materias = []
    asignaciones_actuales = {}  # {grupo_id: [materia_ids]}
    
    if carrera_id and cuatrimestre:
        # Obtener grupos
        grupos = Grupo.query.filter_by(
            carrera_id=carrera_id,
            cuatrimestre=cuatrimestre,
            activo=True
        ).order_by(Grupo.codigo).all()
        
        # Obtener materias del cuatrimestre
        materias = Materia.query.filter_by(
            carrera_id=carrera_id,
            cuatrimestre=cuatrimestre,
            activa=True
        ).order_by(Materia.codigo).all()
        
        # Obtener asignaciones actuales
        for grupo in grupos:
            asignaciones_actuales[grupo.id] = [m.id for m in grupo.materias]
    
    # Procesar POST (guardar cambios)
    if request.method == 'POST':
        try:
            # Procesar asignaciones desde el formulario
            for grupo in grupos:
                materias_seleccionadas = []
                for materia in materias:
                    checkbox_name = f'asignacion_{grupo.id}_{materia.id}'
                    if request.form.get(checkbox_name) == 'on':
                        materias_seleccionadas.append(materia)
                
                # Actualizar materias del grupo
                grupo.materias = materias_seleccionadas
            
            db.session.commit()
            flash('Asignaciones actualizadas exitosamente.', 'success')
            return redirect(url_for('asignacion_masiva_materias_grupos', carrera_id=carrera_id, cuatrimestre=cuatrimestre))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar asignaciones: {str(e)}', 'error')
    
    return render_template('admin/asignacion_masiva_materias_grupos.html',
                         carreras=carreras,
                         carrera_id=carrera_id,
                         cuatrimestre=cuatrimestre,
                         grupos=grupos,
                         materias=materias,
                         asignaciones_actuales=asignaciones_actuales)


@app.route('/admin/grupo/<int:grupo_id>/asignar-profesores', methods=['GET', 'POST'])
@login_required
def asignar_profesores_grupo(grupo_id):
    """Asignar profesores a las materias de un grupo específico"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    grupo = Grupo.query.get_or_404(grupo_id)
    
    # Verificar que el jefe de carrera puede acceder
    if current_user.is_jefe_carrera() and not current_user.tiene_carrera(grupo.carrera_id):
        flash('No tienes acceso a este grupo.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener materias del grupo
    materias_grupo = grupo.materias
    
    # Obtener profesores disponibles (de la misma carrera)
    profesores = User.query.filter(
        User.carreras.any(id=grupo.carrera_id),
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True
    ).order_by(User.apellido, User.nombre).all()
    
    # Obtener asignaciones actuales
    asignaciones_actuales = {}
    for asig in AsignacionProfesorGrupo.query.filter_by(grupo_id=grupo.id, activo=True).all():
        if asig.materia_id not in asignaciones_actuales:
            asignaciones_actuales[asig.materia_id] = []
        asignaciones_actuales[asig.materia_id].append(asig.profesor_id)
    
    if request.method == 'POST':
        try:
            # Desactivar todas las asignaciones anteriores de este grupo
            AsignacionProfesorGrupo.query.filter_by(
                grupo_id=grupo.id,
                activo=True
            ).update({'activo': False})
            db.session.flush()  # Forzar el UPDATE antes de insertar nuevos
            
            # Procesar nuevas asignaciones
            asignaciones_creadas = 0
            for materia in materias_grupo:
                # El nombre del campo es 'profesor_<materia_id>' (singular)
                profesor_id = request.form.get(f'profesor_{materia.id}')
                
                if profesor_id:
                    # Buscar si ya existe una asignación inactiva para esta combinación
                    asig_existente = AsignacionProfesorGrupo.query.filter_by(
                        profesor_id=int(profesor_id),
                        materia_id=materia.id,
                        grupo_id=grupo.id
                    ).first()
                    
                    if asig_existente:
                        # Reactivar la asignación existente
                        asig_existente.activo = True
                        asig_existente.horas_semanales = materia.horas_semanales
                    else:
                        # Crear nueva asignación
                        nueva_asig = AsignacionProfesorGrupo(
                            profesor_id=int(profesor_id),
                            materia_id=materia.id,
                            grupo_id=grupo.id,
                            horas_semanales=materia.horas_semanales,
                            creado_por=current_user.id
                        )
                        db.session.add(nueva_asig)
                    asignaciones_creadas += 1
            
            db.session.commit()
            flash(f'Se guardaron {asignaciones_creadas} asignaciones para el grupo {grupo.codigo}.', 'success')
            return redirect(url_for('asignar_profesores_grupo', grupo_id=grupo.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar asignaciones: {str(e)}', 'error')
    
    return render_template('admin/asignar_profesores_grupo.html',
                         grupo=grupo,
                         materias=materias_grupo,
                         profesores=profesores,
                         asignaciones_actuales=asignaciones_actuales)

@app.route('/admin/grupos/importar-asignaciones-materias', methods=['GET', 'POST'])
@login_required
def importar_asignaciones_grupos():
    """Importar asignaciones de materias a grupos desde CSV"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    from forms import ImportarAsignacionesGrupoForm
    form = ImportarAsignacionesGrupoForm()
    
    if form.validate_on_submit():
        from utils import procesar_archivo_asignaciones_grupo
        archivo = form.archivo.data
        resultado = procesar_archivo_asignaciones_grupo(archivo)
        
        if resultado['exito']:
            flash(resultado['mensaje'], 'success')
            if resultado['errores']:
                for error in resultado['errores'][:10]:  # Mostrar máximo 10 errores
                    flash(error, 'warning')
        else:
            flash(resultado['mensaje'], 'error')
            for error in resultado['errores'][:5]:
                flash(error, 'error')
        
        return redirect(url_for('asignacion_masiva_materias_grupos'))
    
    return render_template('admin/importar_asignaciones_grupos.html', form=form)

@app.route('/admin/grupos/descargar-plantilla-asignaciones')
@login_required
def descargar_plantilla_asignaciones_grupos():
    """Descargar plantilla CSV para importar asignaciones de materias a grupos"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    from utils import generar_plantilla_csv_asignaciones_grupo
    contenido_csv = generar_plantilla_csv_asignaciones_grupo()
    
    response = make_response(contenido_csv)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=plantilla_asignaciones_grupos.csv'
    
    return response

@app.route('/admin/grupos/exportar-asignaciones-materias')
@login_required
def exportar_asignaciones_grupos():
    """Exportar asignaciones actuales de materias a grupos en CSV"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    carrera_id = request.args.get('carrera_id', type=int)
    cuatrimestre = request.args.get('cuatrimestre', type=int)
    
    # Si es jefe de carrera, forzar su carrera
    if current_user.is_jefe_carrera():
        carrera_id = current_user.primera_carrera_id
    
    from utils import exportar_asignaciones_grupo_csv
    contenido_csv = exportar_asignaciones_grupo_csv(carrera_id, cuatrimestre)
    
    response = make_response(contenido_csv)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    
    filename = 'asignaciones_grupos'
    if carrera_id:
        carrera = Carrera.query.get(carrera_id)
        if carrera:
            filename += f'_{carrera.codigo}'
    if cuatrimestre:
        filename += f'_C{cuatrimestre}'
    filename += '.csv'
    
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@app.route('/admin/grupos/auto-asignar-materias', methods=['POST'])
@login_required
def auto_asignar_materias_grupos():
    """Auto-asignar todas las materias del cuatrimestre a los grupos"""
    if not current_user.is_admin() and not current_user.is_jefe_carrera():
        return jsonify({'exito': False, 'mensaje': 'No tienes permisos'}), 403
    
    carrera_id = request.form.get('carrera_id', type=int)
    cuatrimestre = request.form.get('cuatrimestre', type=int)
    
    # Si es jefe de carrera, forzar su carrera
    if current_user.is_jefe_carrera():
        carrera_id = current_user.primera_carrera_id
    
    if not carrera_id or not cuatrimestre:
        return jsonify({'exito': False, 'mensaje': 'Faltan parámetros'}), 400
    
    from utils import auto_asignar_materias_grupo
    resultado = auto_asignar_materias_grupo(carrera_id, cuatrimestre)
    
    return jsonify(resultado)

# Rutas de gestión de horarios (solo administradores)
@app.route('/admin/horarios')
@login_required
def gestionar_horarios():
    """Gestión de horarios - solo para administradores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener horarios separados por turno y ordenados
    horarios_matutino = Horario.query.filter_by(turno='matutino', activo=True).order_by(Horario.orden).all()
    horarios_vespertino = Horario.query.filter_by(turno='vespertino', activo=True).order_by(Horario.orden).all()
    
    return render_template('admin/horarios.html', 
                         horarios_matutino=horarios_matutino,
                         horarios_vespertino=horarios_vespertino)

@app.route('/admin/horarios/agregar', methods=['GET', 'POST'])
@login_required
def agregar_horario():
    """Agregar nuevo horario - solo para administradores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = HorarioForm()
    
    if form.validate_on_submit():
        try:
            horario = Horario(
                nombre=form.nombre.data,
                turno=form.turno.data,
                hora_inicio=form.hora_inicio.data,
                hora_fin=form.hora_fin.data,
                orden=form.orden.data,
                creado_por=current_user.id
            )
            
            db.session.add(horario)
            db.session.commit()
            
            flash(f'Horario "{horario.nombre}" agregado exitosamente.', 'success')
            return redirect(url_for('gestionar_horarios'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al crear el horario. Inténtalo de nuevo.', 'error')
            print(f"Error en agregar horario: {e}")
    
    return render_template('admin/horario_form.html', form=form, horario=None)

@app.route('/admin/horarios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_horario(id):
    """Editar horario existente - solo para administradores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    horario = Horario.query.get_or_404(id)
    
    if not horario.activo:
        flash('No se puede editar un horario inactivo.', 'error')
        return redirect(url_for('gestionar_horarios'))
    
    form = HorarioForm(obj=horario)
    
    # Agregar ID del horario al formulario para validaciones
    form._horario_id = horario.id
    
    if form.validate_on_submit():
        try:
            horario.nombre = form.nombre.data
            horario.turno = form.turno.data
            horario.hora_inicio = form.hora_inicio.data
            horario.hora_fin = form.hora_fin.data
            horario.orden = form.orden.data
            
            db.session.commit()
            
            flash(f'Horario "{horario.nombre}" actualizado exitosamente.', 'success')
            return redirect(url_for('gestionar_horarios'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar el horario. Inténtalo de nuevo.', 'error')
            print(f"Error en editar horario: {e}")
    
    return render_template('admin/horario_form.html', form=form, horario=horario)

@app.route('/admin/horarios/eliminar/<int:id>', methods=['GET', 'POST'])
@login_required
def eliminar_horario(id):
    """Eliminar horario - solo para administradores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    horario = Horario.query.get_or_404(id)
    
    if not horario.activo:
        flash('Este horario ya está eliminado.', 'warning')
        return redirect(url_for('gestionar_horarios'))
    
    form = EliminarHorarioForm()
    
    if form.validate_on_submit():
        try:
            # Marcar como inactivo en lugar de eliminar físicamente
            horario.activo = False
            db.session.commit()
            
            flash(f'Horario "{horario.nombre}" eliminado exitosamente.', 'success')
            return redirect(url_for('gestionar_horarios'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al eliminar el horario. Inténtalo de nuevo.', 'error')
            print(f"Error en eliminar horario: {e}")
    
    return render_template('admin/eliminar_horario.html', form=form, horario=horario)

# Rutas para gestión de carreras
@app.route('/admin/carreras')
@login_required
def gestionar_carreras():
    """Gestión de carreras (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    
    # Calcular estadísticas
    total_profesores = sum(carrera.get_profesores_count() for carrera in carreras)
    promedio_profesores = total_profesores / len(carreras) if carreras else 0
    
    # Estadísticas de jefes de carrera
    carreras_con_jefes = sum(1 for carrera in carreras if carrera.tiene_jefe_carrera())
    carreras_sin_jefes = len(carreras) - carreras_con_jefes
    total_jefes_carrera = User.query.filter_by(rol='jefe_carrera', activo=True).count()
    
    # Obtener facultades únicas
    facultades = set(carrera.facultad for carrera in carreras if carrera.facultad)
    
    # Contar carreras con profesores
    carreras_con_profesores = sum(1 for carrera in carreras if carrera.get_profesores_count() > 0)
    
    return render_template('admin/carreras.html', 
                         carreras=carreras,
                         total_profesores=total_profesores,
                         promedio_profesores=promedio_profesores,
                         carreras_con_jefes=carreras_con_jefes,
                         carreras_sin_jefes=carreras_sin_jefes,
                         total_jefes_carrera=total_jefes_carrera,
                         facultades=facultades,
                         carreras_con_profesores=carreras_con_profesores)

@app.route('/admin/carreras/test')
@login_required
def test_carreras():
    """Página de prueba para carreras"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    return render_template('admin/carreras_test.html', carreras=carreras)

@app.route('/admin/carreras/nueva', methods=['GET', 'POST'])
@login_required
def nueva_carrera():
    """Crear nueva carrera (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = CarreraForm()
    if form.validate_on_submit():
        try:
            carrera = Carrera(
                nombre=form.nombre.data,
                codigo=form.codigo.data.upper(),
                descripcion=form.descripcion.data,
                facultad=form.facultad.data,
                creada_por=current_user.id
            )
            
            db.session.add(carrera)
            db.session.flush()  # Para obtener el ID de la carrera
            
            # Asignar jefe de carrera si se seleccionó uno
            if form.jefe_carrera_id.data:
                jefe = User.query.get(form.jefe_carrera_id.data)
                if jefe:
                    jefe.rol = 'jefe_carrera'
                    # Agregar esta carrera a las carreras del jefe (relación many-to-many)
                    if carrera not in jefe.carreras:
                        jefe.carreras.append(carrera)
            
            db.session.commit()
            
            mensaje = f'Carrera "{carrera.nombre}" creada exitosamente.'
            if form.jefe_carrera_id.data:
                jefe = User.query.get(form.jefe_carrera_id.data)
                mensaje += f' Jefe de carrera asignado: {jefe.get_nombre_completo()}.'
                
            flash(mensaje, 'success')
            return redirect(url_for('gestionar_carreras'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al crear la carrera. Inténtalo de nuevo.', 'error')
            print(f"Error en nueva carrera: {e}")
    
    return render_template('admin/carrera_form.html', form=form, titulo="Nueva Carrera")

@app.route('/admin/carreras/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_carrera(id):
    """Editar carrera existente (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    carrera = Carrera.query.get_or_404(id)
    if not carrera.activa:
        flash('Esta carrera no está disponible.', 'error')
        return redirect(url_for('gestionar_carreras'))
    
    # Crear formulario sin usar obj=carrera para evitar conflictos
    form = CarreraForm()
    form._carrera_id = carrera.id  # Para validación de código único
    
    # Cargar datos manualmente en GET
    if request.method == 'GET':
        form.nombre.data = carrera.nombre
        form.codigo.data = carrera.codigo
        form.descripcion.data = carrera.descripcion
        form.facultad.data = carrera.facultad
        
        # Cargar el jefe de carrera actual
        jefe_actual = carrera.get_jefe_carrera()
        if jefe_actual:
            form.jefe_carrera_id.data = jefe_actual.id
        else:
            form.jefe_carrera_id.data = 0
    
    if form.validate_on_submit():
        try:
            # Actualizar datos básicos de la carrera
            carrera.nombre = form.nombre.data
            carrera.codigo = form.codigo.data.upper()
            carrera.descripcion = form.descripcion.data
            carrera.facultad = form.facultad.data
            
            # Manejar cambio de jefe de carrera
            nuevo_jefe_id = form.jefe_carrera_id.data
            jefe_anterior = carrera.get_jefe_carrera()
            
            # Remover esta carrera del jefe anterior si existe
            if jefe_anterior and (not nuevo_jefe_id or str(jefe_anterior.id) != str(nuevo_jefe_id)):
                # Quitar esta carrera de las carreras del jefe anterior
                if carrera in jefe_anterior.carreras:
                    jefe_anterior.carreras.remove(carrera)
                # Si ya no tiene carreras asignadas, cambiar a profesor
                if len(jefe_anterior.carreras) == 0:
                    jefe_anterior.rol = 'profesor_completo'
            
            # Asignar nuevo jefe de carrera si se seleccionó uno
            if nuevo_jefe_id:
                nuevo_jefe = User.query.get(nuevo_jefe_id)
                if nuevo_jefe:
                    nuevo_jefe.rol = 'jefe_carrera'
                    # Agregar esta carrera a las carreras del nuevo jefe (si no está)
                    if carrera not in nuevo_jefe.carreras:
                        nuevo_jefe.carreras.append(carrera)
            
            db.session.commit()
            
            mensaje = f'Carrera "{carrera.nombre}" actualizada exitosamente.'
            if nuevo_jefe_id:
                nuevo_jefe = User.query.get(nuevo_jefe_id)
                if nuevo_jefe:
                    mensaje += f' Jefe de carrera: {nuevo_jefe.get_nombre_completo()}.'
                
            flash(mensaje, 'success')
            return redirect(url_for('gestionar_carreras'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar la carrera. Inténtalo de nuevo.', 'error')
            print(f"Error en editar carrera: {e}")
    
    return render_template('admin/carrera_form.html', form=form, carrera=carrera, titulo="Editar Carrera")

@app.route('/admin/carreras/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_carrera(id):
    """Eliminar carrera (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para realizar esta acción.', 'error')
        return redirect(url_for('gestionar_carreras'))
    
    try:
        carrera = Carrera.query.get_or_404(id)
        
        # Verificar si tiene profesores asignados
        profesores_asignados = User.query.filter_by(carrera_id=id).count()
        if profesores_asignados > 0:
            flash(f'No se puede eliminar la carrera "{carrera.nombre}" porque tiene {profesores_asignados} profesores asignados.', 'error')
            return redirect(url_for('gestionar_carreras'))
        
        # Marcar como inactiva
        carrera.activa = False
        db.session.commit()
        
        flash(f'Carrera "{carrera.nombre}" eliminada exitosamente.', 'success')
        return redirect(url_for('gestionar_carreras'))
        
    except Exception as e:
        db.session.rollback()
        flash('Error al eliminar la carrera. Inténtalo de nuevo.', 'error')
        return redirect(url_for('gestionar_carreras'))

@app.route('/admin/carreras/importar', methods=['GET', 'POST'])
@login_required
def importar_carreras():
    """Importar carreras desde archivo CSV (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = ImportarCarrerasForm()
    
    if form.validate_on_submit():
        try:
            archivo = form.archivo.data
            
            # Procesar el archivo
            resultado = procesar_archivo_carreras(archivo, current_user.id)
            
            if resultado['exito']:
                # Mostrar resultados
                flash(resultado['mensaje'], 'success')
                
                # Mostrar carreras creadas/actualizadas
                if resultado['carreras_creadas']:
                    carreras_msg = "Carreras procesadas: "
                    for carrera in resultado['carreras_creadas'][:10]:  # Limitar a 10 para no saturar
                        carreras_msg += f"{carrera['codigo']} ({carrera['accion']}), "
                    flash(carreras_msg.rstrip(', '), 'info')
                
                # Mostrar errores si los hay
                if resultado['errores']:
                    for error in resultado['errores'][:5]:  # Mostrar solo los primeros 5 errores
                        flash(error, 'warning')
                    if len(resultado['errores']) > 5:
                        flash(f"... y {len(resultado['errores']) - 5} errores más.", 'warning')
                
                return redirect(url_for('gestionar_carreras'))
            else:
                flash(resultado['mensaje'], 'error')
                
                # Mostrar errores
                if resultado['errores']:
                    for error in resultado['errores'][:5]:
                        flash(error, 'warning')
                    if len(resultado['errores']) > 5:
                        flash(f"... y {len(resultado['errores']) - 5} errores más.", 'warning')
                
        except Exception as e:
            flash(f'Error al procesar el archivo: {str(e)}', 'error')
    
    return render_template('admin/importar_carreras.html', form=form)

@app.route('/admin/carreras/plantilla')
@login_required
def descargar_plantilla_carreras():
    """Descargar plantilla CSV para importar carreras (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para realizar esta acción.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        plantilla = generar_plantilla_csv_carreras()
        
        response = make_response(plantilla)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename=plantilla_carreras.csv'
        
        return response
        
    except Exception as e:
        flash(f'Error al generar la plantilla: {str(e)}', 'error')
        return redirect(url_for('gestionar_carreras'))

# Rutas para gestión de materias
@app.route('/admin/materias')
@login_required
def gestionar_materias():
    """Gestión de materias (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener filtros
    carrera_id_str = request.args.get('carrera', type=str)
    carrera_id = int(carrera_id_str) if carrera_id_str and carrera_id_str != '0' else None
    ciclo_str = request.args.get('ciclo', type=str)
    ciclo = int(ciclo_str) if ciclo_str and ciclo_str != '' else None
    cuatrimestre_str = request.args.get('cuatrimestre', type=str)
    cuatrimestre = int(cuatrimestre_str) if cuatrimestre_str and cuatrimestre_str != '' else None
    busqueda = request.args.get('busqueda', '').strip()
    
    # Query base para materias
    query = Materia.query.filter_by(activa=True)
    
    # Aplicar filtros
    if carrera_id:
        query = query.filter(Materia.carrera_id == carrera_id)
    
    # Filtro por ciclo escolar
    if ciclo:
        # Ciclo 1: cuatrimestres 1, 4, 7, 10 (cuatrimestre % 3 == 1)
        # Ciclo 2: cuatrimestres 2, 5, 8 (cuatrimestre % 3 == 2)
        # Ciclo 3: cuatrimestres 0, 3, 6, 9 (cuatrimestre % 3 == 0)
        if ciclo == 1:
            query = query.filter(Materia.cuatrimestre % 3 == 1)
        elif ciclo == 2:
            query = query.filter(Materia.cuatrimestre % 3 == 2)
        elif ciclo == 3:
            query = query.filter(Materia.cuatrimestre % 3 == 0)
    
    if cuatrimestre is not None:
        query = query.filter(Materia.cuatrimestre == cuatrimestre)
    
    if busqueda:
        query = query.filter(
            db.or_(
                Materia.nombre.ilike(f'%{busqueda}%'),
                Materia.codigo.ilike(f'%{busqueda}%'),
                Materia.descripcion.ilike(f'%{busqueda}%')
            )
        )
    
    materias = query.order_by(Materia.cuatrimestre, Materia.nombre).all()
    
    # Obtener carreras para el filtro
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    
    # Calcular estadísticas
    total_materias = len(materias)
    total_carreras = len(carreras)
    cuatrimestres_unicos = len(set(materia.cuatrimestre for materia in materias)) if materias else 0
    
    # Crear formularios
    filtrar_form = FiltrarMateriasForm()
    filtrar_form.carrera.choices = [(0, 'Todas las carreras')] + [(c.id, c.nombre) for c in carreras]
    
    # Configurar valores por defecto del formulario
    if carrera_id is not None:
        filtrar_form.carrera.data = carrera_id
    else:
        filtrar_form.carrera.data = 0  # Todas las carreras por defecto
    
    if ciclo is not None:
        filtrar_form.ciclo.data = str(ciclo)
    
    if cuatrimestre is not None:
        filtrar_form.cuatrimestre.data = str(cuatrimestre)
    
    exportar_form = ExportarMateriasForm()
    exportar_form.carrera.choices = [('', 'Todas las carreras')] + [(str(c.id), c.nombre) for c in carreras]
    
    return render_template('admin/materias.html', 
                         materias=materias, 
                         carreras=carreras,
                         filtrar_form=filtrar_form,
                         exportar_form=exportar_form,
                         total_materias=total_materias,
                         total_carreras=total_carreras,
                         cuatrimestres_unicos=cuatrimestres_unicos,
                         filtros_activos={
                             'carrera': carrera_id,
                             'ciclo': ciclo,
                             'cuatrimestre': cuatrimestre,
                             'busqueda': busqueda
                         })

@app.route('/admin/materias/nueva', methods=['GET', 'POST'])
@login_required
def nueva_materia():
    """Crear nueva materia (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = MateriaForm()
    if form.validate_on_submit():
        try:
            materia = Materia(
                nombre=form.nombre.data,
                codigo=form.codigo.data,
                cuatrimestre=form.cuatrimestre.data,
                carrera_id=int(form.carrera.data),
                creditos=form.creditos.data,
                horas_semanales=form.horas_semanales.data,
                descripcion=form.descripcion.data,
                creado_por=current_user.id
            )
            
            db.session.add(materia)
            db.session.commit()
            
            flash(f'Materia "{materia.nombre}" creada exitosamente.', 'success')
            return redirect(url_for('gestionar_materias'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al crear la materia. Inténtalo de nuevo.', 'error')
            print(f"Error en nueva materia: {e}")
    
    return render_template('admin/materia_form.html', form=form, titulo="Nueva Materia")

@app.route('/admin/materias/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_materia(id):
    """Editar materia existente (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    materia = Materia.query.get_or_404(id)
    if not materia.activa:
        flash('Esta materia no está disponible.', 'error')
        return redirect(url_for('gestionar_materias'))
    
    form = MateriaForm()
    
    # Agregar ID de la materia al formulario para validaciones
    form._materia_id = materia.id
    
    # Pre-llenar el formulario solo en GET
    if request.method == 'GET':
        form.nombre.data = materia.nombre
        form.codigo.data = materia.codigo
        form.cuatrimestre.data = materia.cuatrimestre
        form.creditos.data = materia.creditos
        form.horas_semanales.data = materia.horas_semanales
        form.descripcion.data = materia.descripcion
        form.carrera.data = str(materia.carrera_id)
    
    if form.validate_on_submit():
        try:
            materia.nombre = form.nombre.data
            materia.codigo = form.codigo.data
            materia.cuatrimestre = form.cuatrimestre.data
            materia.carrera_id = int(form.carrera.data)
            materia.creditos = form.creditos.data
            materia.horas_semanales = form.horas_semanales.data
            materia.descripcion = form.descripcion.data
            
            db.session.commit()
            
            flash(f'Materia "{materia.nombre}" actualizada exitosamente.', 'success')
            return redirect(url_for('gestionar_materias'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar la materia. Inténtalo de nuevo.', 'error')
            print(f"Error en editar materia: {e}")
    
    return render_template('admin/materia_form.html', form=form, materia=materia, titulo="Editar Materia")

@app.route('/admin/materias/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_materia(id):
    """Eliminar materia (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para realizar esta acción.', 'error')
        return redirect(url_for('gestionar_materias'))
    
    try:
        materia = Materia.query.get_or_404(id)
        
        # Marcar como inactiva
        materia.activa = False
        db.session.commit()
        
        flash(f'Materia "{materia.nombre}" eliminada exitosamente.', 'success')
        return redirect(url_for('gestionar_materias'))
        
    except Exception as e:
        db.session.rollback()
        flash('Error al eliminar la materia. Inténtalo de nuevo.', 'error')
        return redirect(url_for('gestionar_materias'))

@app.route('/admin/materias/importar', methods=['GET', 'POST'])
@login_required
def importar_materias():
    """Importar materias desde archivo CSV/Excel (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = ImportarMateriasForm()
    
    if form.validate_on_submit():
        try:
            archivo = form.archivo.data
            carrera_defecto_id = int(form.carrera_defecto.data) if form.carrera_defecto.data else None
            restar_horas = form.restar_horas.data if form.restar_horas.data else 0
            
            resultado = procesar_archivo_materias(archivo, carrera_defecto_id, restar_horas)
            
            if resultado['exito']:
                mensaje = f"Importación exitosa: {resultado['procesados']} materias procesadas, " \
                         f"{resultado['creados']} nuevas, {resultado['actualizados']} actualizadas."
                
                if restar_horas > 0:
                    mensaje += f" Se restaron {restar_horas} hora(s) a cada materia."
                
                flash(mensaje, 'success')
                
                if resultado['errores']:
                    flash(f"Se encontraron {len(resultado['errores'])} errores durante la importación.", 'warning')
                    return render_template('admin/importar_materias.html', 
                                         form=form, 
                                         resultado=resultado)
            else:
                flash(f"Error en la importación: {resultado['mensaje']}", 'error')
                
        except Exception as e:
            flash('Error al procesar el archivo. Inténtalo de nuevo.', 'error')
            print(f"Error en importar materias: {e}")
    
    return render_template('admin/importar_materias.html', form=form)

@app.route('/admin/materias/exportar')
@login_required
def exportar_materias():
    """Exportar materias a PDF (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Obtener filtros de la URL
        carrera_id = request.args.get('carrera_id', type=int)
        ciclo = request.args.get('ciclo', type=int)
        cuatrimestre = request.args.get('cuatrimestre', type=int)
        
        # Query para materias con filtros
        query = Materia.query.filter_by(activa=True)
        
        if carrera_id:
            query = query.filter(Materia.carrera_id == carrera_id)
        
        if ciclo:
            # Filtrar por ciclo escolar
            cuatrimestres_ciclo = []
            if ciclo == 1:
                cuatrimestres_ciclo = [1, 4, 7, 10]
            elif ciclo == 2:
                cuatrimestres_ciclo = [2, 5, 8]
            elif ciclo == 3:
                cuatrimestres_ciclo = [0, 3, 6, 9]
            
            if cuatrimestres_ciclo:
                query = query.filter(Materia.cuatrimestre.in_(cuatrimestres_ciclo))
        
        if cuatrimestre:
            query = query.filter(Materia.cuatrimestre == cuatrimestre)
        
        materias = query.order_by(Materia.cuatrimestre, Materia.nombre).all()
        
        # Generar PDF
        nombre_carrera = None
        if carrera_id:
            carrera = Carrera.query.get(carrera_id)
            nombre_carrera = carrera.nombre if carrera else None
        
        pdf_content = generar_pdf_materias(materias, nombre_carrera, cuatrimestre, ciclo)
        
        # Crear respuesta con el PDF
        response = make_response(pdf_content)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=materias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        flash('Error al generar el PDF. Inténtalo de nuevo.', 'error')
        print(f"Error en exportar materias: {e}")
        return redirect(url_for('gestionar_materias'))

@app.route('/admin/materias/plantilla-csv')
@login_required
def descargar_plantilla_csv_materias():
    """Descargar plantilla CSV para importar materias (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Crear contenido CSV solo con encabezados (sin ejemplos)
        contenido_csv = """nombre,codigo,cuatrimestre,carrera_codigo,creditos,horas_semanales,descripcion
"""
        
        # Crear respuesta con archivo CSV
        response = make_response(contenido_csv)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename=plantilla_materias.csv'
        
        return response
        
    except Exception as e:
        flash('Error al generar la plantilla. Inténtalo de nuevo.', 'error')
        print(f"Error en descargar plantilla CSV: {e}")
        return redirect(url_for('importar_materias'))

# Rutas para gestión de profesores
@app.route('/admin/profesores')
@login_required
def gestionar_profesores():
    """Gestión de profesores (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener todos los profesores activos (sin filtros para estadísticas)
    todos_profesores = User.query.filter(
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True
    ).all()
    
    # Calcular estadísticas
    total_profesores = len(todos_profesores)
    profesores_completos = len([p for p in todos_profesores if p.rol == 'profesor_completo'])
    profesores_asignatura = len([p for p in todos_profesores if p.rol == 'profesor_asignatura'])
    
    # Obtener carreras para el filtro
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    total_carreras = len(carreras)
    
    # Para la tabla mostramos todos los profesores (el filtrado se hace en JavaScript)
    profesores = todos_profesores
    
    # Crear formularios
    filtrar_form = FiltrarProfesoresForm()
    filtrar_form.carrera.choices = [(0, 'Todas las carreras')] + [(c.id, c.nombre) for c in carreras]
    
    exportar_form = ExportarProfesoresForm()
    
    return render_template('admin/profesores.html', 
                         profesores=profesores, 
                         carreras=carreras,
                         filtrar_form=filtrar_form,
                         exportar_form=exportar_form,
                         total_profesores=total_profesores,
                         profesores_completos=profesores_completos,
                         profesores_asignatura=profesores_asignatura,
                         total_carreras=total_carreras)

@app.route('/admin/profesores/importar', methods=['GET', 'POST'])
@login_required
def importar_profesores():
    """Importar profesores desde archivo CSV/Excel (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = ImportarProfesoresForm()
    
    if form.validate_on_submit():
        try:
            archivo = form.archivo.data
            resultado = procesar_archivo_profesores(archivo)
            
            if resultado['exito']:
                # Crear mensaje con usuarios creados
                mensaje_usuarios = ""
                if resultado['usuarios_creados']:
                    mensaje_usuarios = f" {len(resultado['usuarios_creados'])} usuarios nuevos con contraseñas generadas."
                
                flash(f"Importación exitosa: {resultado['procesados']} usuarios procesados, "
                     f"{resultado['creados']} nuevos, {resultado['actualizados']} actualizados.{mensaje_usuarios}", 'success')
                
                if resultado['errores']:
                    flash(f"Se encontraron {len(resultado['errores'])} errores durante la importación.", 'warning')
                
                # Si hay usuarios creados, mostrar la tabla con contraseñas
                if resultado['usuarios_creados']:
                    return render_template('admin/importar_profesores.html', 
                                         form=form, 
                                         resultado=resultado,
                                         mostrar_passwords=True)
                
                return render_template('admin/importar_profesores.html', 
                                     form=form, 
                                     resultado=resultado)
            else:
                flash(f"Error en la importación: {resultado['mensaje']}", 'error')
                
        except Exception as e:
            flash('Error al procesar el archivo. Inténtalo de nuevo.', 'error')
            print(f"Error en importar profesores: {e}")
    
    return render_template('admin/importar_profesores.html', form=form)

@app.route('/admin/profesores/plantilla-csv')
@login_required
def descargar_plantilla_csv_profesores():
    """Descargar plantilla CSV para importar profesores (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        return generar_plantilla_csv()
    except Exception as e:
        flash('Error al generar la plantilla CSV. Inténtalo de nuevo.', 'error')
        print(f"Error en descargar plantilla CSV profesores: {e}")
        return redirect(url_for('importar_profesores'))

@app.route('/admin/profesores/exportar')
@login_required
def exportar_profesores():
    """Exportar profesores a PDF (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Obtener filtros de la URL
        carrera_id = request.args.get('carrera_id', type=int)
        incluir_contacto = request.args.get('incluir_contacto', 'si') == 'si'
        
        # Generar PDF usando la función de utils
        from utils import generar_pdf_profesores
        pdf_buffer = generar_pdf_profesores(carrera_id=carrera_id, incluir_contacto=incluir_contacto)
        
        # Crear nombre del archivo
        if carrera_id:
            carrera = Carrera.query.get(carrera_id)
            nombre_archivo = f'profesores_{carrera.codigo if carrera else "carrera"}_{datetime.now().strftime("%Y%m%d")}.pdf'
        else:
            nombre_archivo = f'profesores_{datetime.now().strftime("%Y%m%d")}.pdf'
        
        return send_file(
            BytesIO(pdf_buffer),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        flash('Error al generar el PDF. Inténtalo de nuevo.', 'error')
        print(f"Error en exportar profesores: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('gestionar_profesores'))

@app.route('/admin/profesores/<int:id>/toggle-estado', methods=['POST'])
@login_required
def toggle_estado_profesor(id):
    """Activar/desactivar profesor (solo admin)"""
    if not current_user.is_admin():
        return jsonify({'error': 'No autorizado'}), 403
    
    try:
        profesor = User.query.get_or_404(id)
        
        if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
            return jsonify({'error': 'Usuario no es profesor'}), 400
        
        profesor.activo = not profesor.activo
        db.session.commit()
        
        estado = 'activado' if profesor.activo else 'desactivado'
        return jsonify({
            'message': f'Profesor {profesor.get_nombre_completo()} {estado} exitosamente.',
            'nuevo_estado': profesor.activo
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error al cambiar estado del profesor.'}), 500

@app.route('/admin/profesores/agregar', methods=['GET', 'POST'])
@login_required
def agregar_profesor():
    """Agregar profesor manualmente (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = AgregarProfesorForm()
    
    # Obtener horarios para mostrar en la tabla
    from models import Horario
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    if form.validate_on_submit():
        try:
            # Crear nuevo profesor
            nuevo_profesor = User(
                username=form.username.data,
                email=form.email.data,
                password=form.password.data,
                nombre=form.nombre.data,
                apellido=form.apellido.data,
                rol=form.tipo_profesor.data,
                telefono=form.telefono.data
            )
            
            db.session.add(nuevo_profesor)
            db.session.flush()  # Obtener el ID del profesor sin hacer commit
            
            # Asignar carreras (relación many-to-many para profesores - ahora soporta múltiples)
            if form.carreras.data:
                from models import Carrera
                carreras_seleccionadas = Carrera.query.filter(Carrera.id.in_(form.carreras.data)).all()
                nuevo_profesor.carreras = carreras_seleccionadas
            
            # Guardar disponibilidades
            disponibilidades_data = form.get_disponibilidades_data()
            for disp_data in disponibilidades_data:
                disponibilidad = DisponibilidadProfesor(
                    profesor_id=nuevo_profesor.id,
                    horario_id=disp_data['horario_id'],
                    dia_semana=disp_data['dia_semana'],
                    disponible=disp_data['disponible'],
                    creado_por=current_user.id
                )
                db.session.add(disponibilidad)
            
            db.session.commit()
            
            flash(f'Profesor {nuevo_profesor.get_nombre_completo()} creado exitosamente con su disponibilidad horaria.', 'success')
            return redirect(url_for('gestionar_profesores'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error al crear el profesor. Inténtalo de nuevo.', 'error')
            print(f"Error al crear profesor: {e}")
    
    return render_template('admin/profesor_form.html', form=form, titulo="Agregar Profesor", horarios=horarios)

# ==========================================
# GESTIÓN DE MATERIAS DE PROFESORES
# ==========================================
@app.route('/admin/profesores/<int:id>/materias', methods=['GET', 'POST'])
@login_required
def gestionar_materias_profesor(id):
    """Asignar/modificar materias de un profesor (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
        flash('Este usuario no es un profesor.', 'error')
        return redirect(url_for('gestionar_profesores'))
    
    from forms import AsignarMateriasProfesorForm
    form = AsignarMateriasProfesorForm(profesor=profesor)
    
    if form.validate_on_submit():
        try:
            # Obtener materias seleccionadas
            materias_ids = form.materias.data
            materias_nuevas = Materia.query.filter(Materia.id.in_(materias_ids)).all()
            
            # Actualizar materias del profesor
            profesor.materias = materias_nuevas
            db.session.commit()
            
            flash(f'Materias actualizadas exitosamente para {profesor.get_nombre_completo()}. '
                  f'Total: {len(materias_nuevas)} materias asignadas.', 'success')
            return redirect(url_for('gestionar_profesores'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al asignar materias: {str(e)}', 'error')
    
    # Precargar materias actuales del profesor
    elif request.method == 'GET':
        form.materias.data = [m.id for m in profesor.materias]
    
    return render_template('admin/asignar_materias_profesor.html', 
                         form=form, 
                         profesor=profesor,
                         titulo=f"Asignar Materias - {profesor.get_nombre_completo()}")

@app.route('/admin/profesores/<int:id>/materias/ver')
@login_required
def ver_materias_profesor(id):
    """Ver materias asignadas a un profesor (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
        flash('Este usuario no es un profesor.', 'error')
        return redirect(url_for('gestionar_profesores'))
    
    # Obtener materias ordenadas por cuatrimestre
    materias = sorted(profesor.materias, key=lambda m: (m.cuatrimestre, m.nombre))
    
    # Obtener horarios del profesor agrupados por materia
    horarios_por_materia = {}
    for materia in materias:
        horarios = HorarioAcademico.query.filter_by(
            profesor_id=profesor.id,
            materia_id=materia.id,
            activo=True
        ).all()
        horarios_por_materia[materia.id] = horarios
    
    return render_template('admin/ver_materias_profesor.html',
                         profesor=profesor,
                         materias=materias,
                         horarios_por_materia=horarios_por_materia)

@app.route('/admin/profesores/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_profesor(id):
    """Editar profesor (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
        flash('Este usuario no es un profesor.', 'error')
        return redirect(url_for('gestionar_profesores'))
    
    from forms import EditarUsuarioForm
    form = EditarUsuarioForm(user=profesor)
    
    # Obtener horarios para mostrar en la tabla de disponibilidad
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    if form.validate_on_submit():
        try:
            # Obtener el rol final (considerando tipo de profesor)
            rol_final = form.get_final_rol()
            
            # Actualizar datos del profesor
            profesor.username = form.username.data
            profesor.email = form.email.data
            profesor.nombre = form.nombre.data
            profesor.apellido = form.apellido.data
            profesor.rol = rol_final
            profesor.telefono = form.telefono.data
            profesor.activo = form.activo.data
            
            # Actualizar carreras (many-to-many)
            from models import Carrera
            if form.carreras.data:
                carreras_seleccionadas = Carrera.query.filter(Carrera.id.in_(form.carreras.data)).all()
                profesor.carreras = carreras_seleccionadas
            else:
                profesor.carreras = []
            
            # Limpiar carrera_id si existía (por si antes era jefe de carrera)
            profesor.carrera_id = None
            
            # Procesar disponibilidad horaria
            # Desactivar disponibilidades anteriores (mantener historial)
            DisponibilidadProfesor.query.filter_by(
                profesor_id=profesor.id,
                activo=True
            ).update({'activo': False})
            
            # Crear nuevas disponibilidades basadas en el formulario
            dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
            for horario in horarios:
                for dia in dias:
                    field_name = f'disp_{horario.id}_{dia}'
                    disponible = request.form.get(field_name) == 'True'
                    
                    nueva_disponibilidad = DisponibilidadProfesor(
                        profesor_id=profesor.id,
                        horario_id=horario.id,
                        dia_semana=dia,
                        disponible=disponible,
                        creado_por=current_user.id
                    )
                    db.session.add(nueva_disponibilidad)
            
            db.session.commit()
            flash(f'Profesor {profesor.get_nombre_completo()} actualizado exitosamente.', 'success')
            return redirect(url_for('gestionar_profesores'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar profesor: {str(e)}', 'error')
    
    # Llenar formulario con datos actuales
    elif request.method == 'GET':
        form.username.data = profesor.username
        form.email.data = profesor.email
        form.nombre.data = profesor.nombre
        form.apellido.data = profesor.apellido
        form.telefono.data = profesor.telefono
        form.activo.data = profesor.activo
        
        # Configurar rol y tipo de profesor
        form.rol.data = 'profesor'
        form.tipo_profesor.data = profesor.rol
        # Cargar carreras del profesor
        form.carreras.data = [c.id for c in profesor.carreras]
    
    # Cargar disponibilidades actuales del profesor
    disponibilidad_dict = {}
    disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
        profesor_id=profesor.id,
        activo=True
    ).all()
    
    for disp in disponibilidades_actuales:
        if disp.disponible:
            disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True
    
    return render_template('admin/usuario_form.html', 
                         form=form, 
                         titulo=f"Editar Profesor - {profesor.get_nombre_completo()}", 
                         usuario=profesor,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict)

@app.route('/admin/profesores/<int:id>/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password_profesor(id):
    """Cambiar contraseña de un profesor (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
        flash('Este usuario no es un profesor.', 'error')
        return redirect(url_for('gestionar_profesores'))
    
    form = CambiarPasswordProfesorForm()
    
    if form.validate_on_submit():
        try:
            # Cambiar la contraseña del profesor
            profesor.password = form.nueva_password.data
            db.session.commit()
            
            flash(f'Contraseña actualizada exitosamente para {profesor.get_nombre_completo()}.', 'success')
            return redirect(url_for('gestionar_profesores'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al cambiar la contraseña: {str(e)}', 'error')
    
    return render_template('admin/cambiar_password_profesor.html',
                         form=form,
                         profesor=profesor,
                         titulo=f"Cambiar Contraseña - {profesor.get_nombre_completo()}")

# ==========================================
# GESTIÓN DE DISPONIBILIDAD DE PROFESORES (ADMIN)
# ==========================================

@app.route('/admin/profesores/disponibilidad')
@login_required
def admin_disponibilidad_profesores():
    """Módulo de disponibilidad horaria de profesores para administradores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener filtros opcionales
    carrera_id = request.args.get('carrera', type=int)
    rol_filtro = request.args.get('rol', type=str)
    
    # Construir query base
    query = User.query.filter(
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True
    )
    
    # Aplicar filtros
    if carrera_id:
        query = query.filter(User.carreras.any(id=carrera_id))
    if rol_filtro:
        query = query.filter(User.rol == rol_filtro)
    
    profesores = query.order_by(User.apellido, User.nombre).all()
    
    # Calcular estadísticas de disponibilidad
    total_profesores = len(profesores)
    profesores_con_disponibilidad = 0
    
    for profesor in profesores:
        disponibilidades = DisponibilidadProfesor.query.filter_by(
            profesor_id=profesor.id,
            activo=True
        ).count()
        if disponibilidades > 0:
            profesores_con_disponibilidad += 1
    
    # Obtener todas las carreras para el filtro
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    
    return render_template('admin/admin_disponibilidad_profesores.html',
                         profesores=profesores,
                         carreras=carreras,
                         carrera_id=carrera_id,
                         rol_filtro=rol_filtro,
                         total_profesores=total_profesores,
                         profesores_con_disponibilidad=profesores_con_disponibilidad)

@app.route('/admin/profesor/<int:id>/disponibilidad/editar', methods=['GET', 'POST'])
@login_required
def admin_editar_disponibilidad_profesor(id):
    """Editar disponibilidad horaria de un profesor (administrador)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que sea un profesor
    if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
        flash('El usuario seleccionado no es un profesor.', 'error')
        return redirect(url_for('admin_disponibilidad_profesores'))
    
    # Obtener horarios del sistema
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    if request.method == 'POST':
        try:
            # Procesar disponibilidad
            dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
            
            # Desactivar disponibilidades anteriores
            DisponibilidadProfesor.query.filter_by(
                profesor_id=profesor.id,
                activo=True
            ).update({'activo': False})
            
            # Crear nuevas disponibilidades basadas en los checkboxes marcados
            for horario in horarios:
                for dia in dias:
                    field_name = f"disp_{horario.id}_{dia}"
                    if request.form.get(field_name):
                        nueva_disponibilidad = DisponibilidadProfesor(
                            profesor_id=profesor.id,
                            horario_id=horario.id,
                            dia_semana=dia,
                            disponible=True,
                            creado_por=current_user.id
                        )
                        db.session.add(nueva_disponibilidad)
            
            db.session.commit()
            flash(f'Disponibilidad de {profesor.get_nombre_completo()} actualizada exitosamente.', 'success')
            return redirect(url_for('admin_disponibilidad_profesores'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar disponibilidad: {str(e)}', 'error')
    
    # Cargar disponibilidades actuales
    disponibilidad_dict = {}
    disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
        profesor_id=profesor.id,
        activo=True
    ).all()
    
    for disp in disponibilidades_actuales:
        if disp.disponible:
            disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True
    
    return render_template('admin/admin_editar_disponibilidad_profesor.html',
                         profesor=profesor,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict)

@app.route('/admin/profesor/<int:id>/disponibilidad/ver')
@login_required
def admin_ver_disponibilidad_profesor(id):
    """Ver disponibilidad horaria de un profesor (administrador)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    profesor = User.query.get_or_404(id)
    
    # Verificar que sea un profesor
    if profesor.rol not in ['profesor_completo', 'profesor_asignatura']:
        flash('El usuario seleccionado no es un profesor.', 'error')
        return redirect(url_for('admin_disponibilidad_profesores'))
    
    # Obtener horarios del sistema
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()
    
    # Cargar disponibilidades actuales
    disponibilidad_dict = {}
    disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
        profesor_id=profesor.id,
        activo=True
    ).all()
    
    for disp in disponibilidades_actuales:
        if disp.disponible:
            disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True
    
    # Calcular total de horas disponibles
    total_horas_disponibles = len(disponibilidades_actuales)
    
    return render_template('admin/admin_ver_disponibilidad_profesor.html',
                         profesor=profesor,
                         horarios=horarios,
                         disponibilidad_dict=disponibilidad_dict,
                         total_horas_disponibles=total_horas_disponibles)

# ==========================================
# ASIGNACIÓN MASIVA DE MATERIAS
# ==========================================
@app.route('/admin/asignacion-masiva-materias', methods=['GET', 'POST'])
@login_required
def asignacion_masiva_materias():
    """Asignación masiva de materias a múltiples profesores"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            asignaciones_marcadas = set(request.form.getlist('asignaciones[]'))
            
            # Obtener filtros para reconstruir el estado original
            carrera_id = request.args.get('carrera', type=int)
            cuatrimestre = request.args.get('cuatrimestre', type=int)
            
            # Obtener todos los profesores y materias según filtros
            profesores = User.query.filter(
                User.rol.in_(['profesor_completo', 'profesor_asignatura']),
                User.activo == True
            ).all()
            
            materias_query = Materia.query.filter_by(activa=True)
            if carrera_id:
                materias_query = materias_query.filter_by(carrera_id=carrera_id)
            if cuatrimestre:
                materias_query = materias_query.filter_by(cuatrimestre=cuatrimestre)
            materias = materias_query.all()
            
            # Procesar cambios
            contador_asignaciones = 0
            contador_desasignaciones = 0
            errores = []
            
            # Iterar sobre todos los profesores y materias visibles
            for profesor in profesores:
                for materia in materias:
                    clave = f"{profesor.id}-{materia.id}"
                    esta_marcada = clave in asignaciones_marcadas
                    esta_asignada = materia in profesor.materias
                    
                    # Si está marcada y no está asignada -> ASIGNAR
                    if esta_marcada and not esta_asignada:
                        profesor.materias.append(materia)
                        contador_asignaciones += 1
                    
                    # Si NO está marcada y SÍ está asignada -> DESASIGNAR
                    elif not esta_marcada and esta_asignada:
                        profesor.materias.remove(materia)
                        contador_desasignaciones += 1
            
            # Guardar cambios
            db.session.commit()
            
            # Mensaje de éxito
            mensajes = []
            if contador_asignaciones > 0:
                mensajes.append(f'{contador_asignaciones} nueva(s) asignación(es)')
            if contador_desasignaciones > 0:
                mensajes.append(f'{contador_desasignaciones} desasignación(es)')
            
            if mensajes:
                flash(f'Cambios realizados exitosamente: {", ".join(mensajes)}.', 'success')
            else:
                flash('No se realizaron cambios.', 'info')
            
            if errores:
                flash(f'Errores: {", ".join(errores)}', 'warning')
            
            return redirect(url_for('asignacion_masiva_materias'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al realizar las asignaciones: {str(e)}', 'error')
            return redirect(url_for('asignacion_masiva_materias'))
    
    # GET - Mostrar formulario
    # Obtener filtros
    carrera_id = request.args.get('carrera', type=int)
    cuatrimestre = request.args.get('cuatrimestre', type=int)
    solo_disponibles = request.args.get('solo_disponibles', type=int, default=0)
    
    # Obtener todos los profesores activos
    profesores_query = User.query.filter(
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True
    )
    
    # Si se filtra por carrera, filtrar profesores de esa carrera
    if carrera_id:
        profesores_query = profesores_query.filter(User.carreras.any(id=carrera_id))
    
    profesores = profesores_query.order_by(User.apellido, User.nombre).all()
    
    # Obtener materias activas
    materias_query = Materia.query.filter_by(activa=True)
    
    if carrera_id:
        materias_query = materias_query.filter_by(carrera_id=carrera_id)
    
    if cuatrimestre:
        materias_query = materias_query.filter_by(cuatrimestre=cuatrimestre)
    
    materias = materias_query.order_by(Materia.cuatrimestre, Materia.nombre).all()
    
    # Obtener carreras para filtros
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    
    # Crear matriz de asignaciones actuales
    asignaciones_actuales = {}
    for profesor in profesores:
        asignaciones_actuales[profesor.id] = set(m.id for m in profesor.materias)
    
    # Calcular carga de trabajo de cada profesor
    cargas_profesores = {}
    for profesor in profesores:
        cargas_profesores[profesor.id] = calcular_carga_profesor(profesor)
    
    # Si se solicita filtrar solo profesores disponibles, filtrar
    if solo_disponibles:
        profesores = [p for p in profesores if cargas_profesores[p.id]['puede_mas']]
    
    return render_template('admin/asignacion_masiva_materias.html',
                         profesores=profesores,
                         materias=materias,
                         carreras=carreras,
                         asignaciones_actuales=asignaciones_actuales,
                         cargas_profesores=cargas_profesores,
                         filtro_carrera=carrera_id,
                         filtro_cuatrimestre=cuatrimestre,
                         solo_disponibles=solo_disponibles)

@app.route('/admin/asignacion-masiva-materias/importar', methods=['GET', 'POST'])
@login_required
def importar_asignaciones_masivas():
    """Importar asignaciones masivas desde archivo CSV (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    form = ImportarAsignacionesForm()
    
    if form.validate_on_submit():
        try:
            archivo = form.archivo.data
            
            # Procesar el archivo
            resultado = procesar_archivo_asignaciones(archivo)
            
            if resultado['exito'] or resultado['asignados'] > 0:
                # Mostrar resultados
                flash(resultado['mensaje'], 'success')
                
                # Mostrar asignaciones realizadas
                if resultado['asignaciones_realizadas']:
                    asignaciones_msg = "Asignaciones realizadas: "
                    for asig in resultado['asignaciones_realizadas'][:10]:  # Limitar a 10
                        asignaciones_msg += f"{asig['profesor']} → {asig['materia_codigo']}, "
                    flash(asignaciones_msg.rstrip(', '), 'info')
                    
                    if len(resultado['asignaciones_realizadas']) > 10:
                        flash(f"... y {len(resultado['asignaciones_realizadas']) - 10} asignaciones más.", 'info')
                
                # Mostrar errores si los hay
                if resultado['errores']:
                    for error in resultado['errores'][:5]:  # Mostrar solo los primeros 5
                        flash(error, 'warning')
                    if len(resultado['errores']) > 5:
                        flash(f"... y {len(resultado['errores']) - 5} errores más.", 'warning')
                
                return redirect(url_for('asignacion_masiva_materias'))
            else:
                flash(resultado['mensaje'], 'warning')
                
                # Mostrar errores
                if resultado['errores']:
                    for error in resultado['errores'][:5]:
                        flash(error, 'warning')
                    if len(resultado['errores']) > 5:
                        flash(f"... y {len(resultado['errores']) - 5} errores más.", 'warning')
                
        except Exception as e:
            flash(f'Error al procesar el archivo: {str(e)}', 'error')
    
    return render_template('admin/importar_asignaciones.html', form=form)

@app.route('/admin/asignacion-masiva-materias/plantilla')
@login_required
def descargar_plantilla_asignaciones():
    """Descargar plantilla CSV para importar asignaciones (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para realizar esta acción.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        plantilla = generar_plantilla_csv_asignaciones()
        
        response = make_response(plantilla)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename=plantilla_asignaciones.csv'
        
        return response
        
    except Exception as e:
        flash(f'Error al generar la plantilla: {str(e)}', 'error')
        return redirect(url_for('asignacion_masiva_materias'))

@app.route('/admin/asignacion-masiva-materias/exportar-actuales')
@login_required
def exportar_asignaciones_actuales():
    """Exportar asignaciones actuales a CSV para editar y reimportar (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para realizar esta acción.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Obtener filtros
        carrera_id = request.args.get('carrera', type=int)
        cuatrimestre = request.args.get('cuatrimestre', type=int)
        
        # Construir consulta
        profesores_query = User.query.filter(
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.activo == True
        )
        
        if carrera_id:
            profesores_query = profesores_query.filter(User.carreras.any(id=carrera_id))
        
        profesores = profesores_query.order_by(User.apellido, User.nombre).all()
        
        # Generar CSV
        csv_content = "profesor_email,materia_codigo\n"
        
        for profesor in profesores:
            for materia in profesor.materias:
                # Aplicar filtros de cuatrimestre si corresponde
                if cuatrimestre and materia.cuatrimestre != cuatrimestre:
                    continue
                if carrera_id and materia.carrera_id != carrera_id:
                    continue
                    
                csv_content += f"{profesor.email},{materia.codigo}\n"
        
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        
        # Nombre del archivo con filtros aplicados
        filename = 'asignaciones_actuales'
        if carrera_id:
            carrera = Carrera.query.get(carrera_id)
            if carrera:
                filename += f'_{carrera.codigo}'
        if cuatrimestre:
            filename += f'_cuatri{cuatrimestre}'
        filename += '.csv'
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error al exportar asignaciones: {str(e)}', 'error')
        return redirect(url_for('asignacion_masiva_materias'))

@app.route('/admin/asignacion-masiva-materias/auto-asignar', methods=['POST'])
@login_required
def auto_asignar_por_carrera():
    """Auto-asignar materias de una carrera a profesores de esa carrera (solo admin)"""
    if not current_user.is_admin():
        return jsonify({'exito': False, 'mensaje': 'No tienes permisos'}), 403
    
    try:
        carrera_id = request.form.get('carrera_id', type=int)
        cuatrimestre = request.form.get('cuatrimestre', type=int)
        
        if not carrera_id:
            return jsonify({'exito': False, 'mensaje': 'Debe especificar una carrera'}), 400
        
        carrera = Carrera.query.get_or_404(carrera_id)
        
        # Obtener profesores de la carrera
        profesores = User.query.filter(
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.activo == True,
            User.carreras.any(id=carrera_id)
        ).all()
        
        if not profesores:
            return jsonify({'exito': False, 'mensaje': f'No hay profesores activos en {carrera.nombre}'}), 400
        
        # Obtener materias de la carrera
        materias_query = Materia.query.filter_by(carrera_id=carrera_id, activa=True)
        if cuatrimestre:
            materias_query = materias_query.filter_by(cuatrimestre=cuatrimestre)
        materias = materias_query.all()
        
        if not materias:
            return jsonify({'exito': False, 'mensaje': 'No hay materias activas para asignar'}), 400
        
        # Estrategia: distribuir materias equitativamente
        # Prioridad a profesores de tiempo completo
        profesores_tc = [p for p in profesores if p.rol == 'profesor_completo']
        profesores_pa = [p for p in profesores if p.rol == 'profesor_asignatura']
        
        # Combinar listas (TC primero)
        profesores_ordenados = profesores_tc + profesores_pa
        
        contador_asignaciones = 0
        indice_profesor = 0
        
        for materia in materias:
            # Buscar profesor que no tenga esta materia y tenga capacidad
            asignado = False
            intentos = 0
            
            while not asignado and intentos < len(profesores_ordenados):
                profesor = profesores_ordenados[indice_profesor]
                carga = calcular_carga_profesor(profesor)
                
                # Si el profesor no tiene la materia y tiene capacidad
                if materia not in profesor.materias and carga['puede_mas']:
                    profesor.materias.append(materia)
                    contador_asignaciones += 1
                    asignado = True
                
                # Avanzar al siguiente profesor (round-robin)
                indice_profesor = (indice_profesor + 1) % len(profesores_ordenados)
                intentos += 1
        
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'mensaje': f'{contador_asignaciones} materias asignadas automáticamente en {carrera.nombre}',
            'asignaciones': contador_asignaciones
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'exito': False, 'mensaje': f'Error: {str(e)}'}), 500

# Manejo de errores
@app.errorhandler(404)
def not_found_error(error):
    """Página de error 404"""
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """Página de error 500"""
    db.session.rollback()
    return render_template('errors/500.html'), 500

# Contexto de plantillas
@app.context_processor
def inject_user_counts():
    """Inyectar conteos de usuarios en todas las plantillas"""
    if current_user.is_authenticated and current_user.is_admin():
        return dict(
            total_users=User.query.count(),
            total_admins=User.query.filter_by(rol='admin').count(),
            total_jefes=User.query.filter_by(rol='jefe_carrera').count(),
            total_profesores=User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).count()
        )
    return dict()

# Rutas para gestión de horarios académicos
@app.route('/admin/horarios-academicos')
@login_required
def gestionar_horarios_academicos():
    """Gestión de horarios académicos generados - Vista por Grupos"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    # Obtener filtros
    carrera_id = request.args.get('carrera', type=int)

    # Query base
    query = HorarioAcademico.query.filter_by(activo=True)

    if carrera_id:
        # Filtrar por carrera a través de la materia
        query = query.join(HorarioAcademico.materia).filter(Materia.carrera_id == carrera_id)

    # Obtener todos los horarios
    horarios_academicos = query.join(HorarioAcademico.horario).all()

    # Agrupar horarios por grupo usando el campo 'grupo' de HorarioAcademico
    # CORREGIDO: Antes se iteraba sobre materia.grupos lo que causaba duplicados
    grupos_con_horarios = {}
    
    for horario in horarios_academicos:
        # Usar el campo 'grupo' del HorarioAcademico (código del grupo)
        grupo_codigo = horario.grupo
        
        if not grupo_codigo:
            continue
        
        # Buscar el objeto Grupo por su código
        grupo = Grupo.query.filter_by(codigo=grupo_codigo).first()
        
        if not grupo:
            continue
        
        # Filtrar por carrera si se especifica
        if carrera_id and grupo.carrera_id != carrera_id:
            continue
        
        grupo_key = grupo.id
        
        if grupo_key not in grupos_con_horarios:
            grupos_con_horarios[grupo_key] = {
                'grupo': grupo,
                'horarios': []
            }
        
        grupos_con_horarios[grupo_key]['horarios'].append(horario)
    
    # Ordenar los horarios de cada grupo por día y hora
    for grupo_data in grupos_con_horarios.values():
        grupo_data['horarios'].sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
    
    # Convertir a lista ordenada por código de grupo
    grupos_lista = sorted(grupos_con_horarios.values(), key=lambda x: x['grupo'].codigo)

    # Obtener datos para filtros
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()

    # Estadísticas
    total_grupos = len(grupos_lista)
    total_horarios = sum(len(g['horarios']) for g in grupos_lista)
    profesores_unicos = len(set(h.profesor_id for g in grupos_lista for h in g['horarios']))
    materias_unicas = len(set(h.materia_id for g in grupos_lista for h in g['horarios']))

    return render_template('admin/horarios_academicos.html',
                         grupos_con_horarios=grupos_lista,
                         carreras=carreras,
                         total_grupos=total_grupos,
                         total_horarios=total_horarios,
                         profesores_unicos=profesores_unicos,
                         materias_unicas=materias_unicas,
                         filtro_carrera=carrera_id)

@app.route('/admin/horarios-academicos/generar-masivo', methods=['GET', 'POST'])
@login_required
def generar_horarios_masivo():
    """Generar horarios para múltiples grupos simultáneamente (equilibrado)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    from generador_horarios import generar_horarios_masivos
    from datetime import datetime
    
    # Obtener todos los grupos activos organizados por carrera y cuatrimestre
    grupos = Grupo.query.filter_by(activo=True).order_by(
        Grupo.carrera_id, 
        Grupo.cuatrimestre, 
        Grupo.codigo
    ).all()
    
    # Organizar grupos por carrera y cuatrimestre
    grupos_organizados = {}
    for grupo in grupos:
        carrera_nombre = grupo.get_carrera_nombre()
        if carrera_nombre not in grupos_organizados:
            grupos_organizados[carrera_nombre] = {}
        
        cuatri = grupo.cuatrimestre
        if cuatri not in grupos_organizados[carrera_nombre]:
            grupos_organizados[carrera_nombre][cuatri] = []
        
        grupos_organizados[carrera_nombre][cuatri].append(grupo)
    
    resultado = None
    
    if request.method == 'POST':
        # Obtener grupos seleccionados
        grupos_ids = request.form.getlist('grupos_ids[]')
        
        if not grupos_ids:
            flash('❌ Debe seleccionar al menos un grupo', 'error')
            return render_template('admin/generar_horarios_masivo.html', 
                                 grupos_organizados=grupos_organizados,
                                 resultado=resultado)
        
        # Convertir a enteros
        grupos_ids = [int(gid) for gid in grupos_ids]
        
        # Obtener configuración
        version_nombre = request.form.get('version_nombre', '').strip()
        dias_config = request.form.get('dias_semana', 'lunes_viernes')
        
        if not version_nombre:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            version_nombre = f"Generación Masiva {timestamp}"
        
        # Configurar días
        if dias_config == 'lunes_viernes':
            dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
        elif dias_config == 'lunes_sabado':
            dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
        else:
            dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
        
        # Calcular período académico
        año_actual = datetime.now().year
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        periodo_academico = f"{año_actual}-{año_actual}_{timestamp}"
        
        # Generar horarios masivos
        print(f"🚀 Iniciando generación masiva para {len(grupos_ids)} grupos...")
        # Inicializar progreso
        global generacion_progreso
        generacion_progreso = {
            'activo': True,
            'total_grupos': len(grupos_ids),
            'grupos_procesados': 0,
            'grupo_actual': 'Iniciando...',
            'horarios_generados': 0,
            'mensaje': 'Iniciando generación...',
            'completado': False,
            'exito': False
        }
        
        # Generar horarios masivos con callback de progreso
        resultado = generar_horarios_masivos_con_progreso(
            grupos_ids=grupos_ids,
            periodo_academico=periodo_academico,
            version_nombre=version_nombre,
            creado_por=current_user.id,
            dias_semana=dias_semana
        )
        
        # Marcar como completado
        generacion_progreso['completado'] = True
        generacion_progreso['exito'] = resultado['exito']
        generacion_progreso['mensaje'] = resultado['mensaje']
        generacion_progreso['activo'] = False
        
        if resultado['exito']:
            flash(f"✅ {resultado['mensaje']}", 'success')
        else:
            flash(f"❌ {resultado['mensaje']}", 'error')
    
    return render_template('admin/generar_horarios_masivo.html',
                         grupos_organizados=grupos_organizados,
                         resultado=resultado)

def generar_horarios_masivos_con_progreso(grupos_ids, periodo_academico, version_nombre, creado_por, dias_semana):
    """Wrapper que genera horarios actualizando el progreso global"""
    global generacion_progreso, progreso_lock
    from generador_horarios_mejorado import GeneradorHorariosMejorado
    from models import Grupo, db
    
    resultados = {
        'exito': True,
        'mensaje': '',
        'grupos_procesados': 0,
        'grupos_fallidos': 0,
        'horarios_generados': 0,
        'algoritmo': 'Secuencial con Progreso'
    }
    
    # Ordenar grupos por complejidad (los más simples primero)
    grupos_ordenados = grupos_ids
    total = len(grupos_ordenados)
    
    for i, grupo_id in enumerate(grupos_ordenados, 1):
        grupo = Grupo.query.get(grupo_id)
        if not grupo:
            continue
        
        # Actualizar progreso con lock
        with progreso_lock:
            generacion_progreso['grupo_actual'] = i  # Solo el número actual
            generacion_progreso['codigo_grupo'] = grupo.codigo
            generacion_progreso['mensaje'] = f'Procesando grupo {grupo.codigo}...'
        
        try:
            db.session.expire_all()
            
            generador = GeneradorHorariosMejorado(
                grupos_ids=[grupo_id],
                periodo_academico=periodo_academico,
                version_nombre=f"{version_nombre or 'Secuencial'} - {grupo.codigo}",
                creado_por=creado_por,
                dias_semana=dias_semana or ['lunes', 'martes', 'miercoles', 'jueves', 'viernes'],
                tiempo_limite=120
            )
            
            resultado = generador.generar()
            
            if resultado['exito']:
                db.session.commit()
                resultados['grupos_procesados'] += 1
                resultados['horarios_generados'] += resultado['horarios_generados']
                
                # Actualizar progreso con lock
                with progreso_lock:
                    generacion_progreso['grupos_procesados'] = resultados['grupos_procesados']
                    generacion_progreso['horarios_generados'] = resultados['horarios_generados']
            else:
                resultados['grupos_fallidos'] += 1
                
        except Exception as e:
            db.session.rollback()
            resultados['grupos_fallidos'] += 1
            print(f"Error en grupo {grupo.codigo}: {e}")
    
    # Resumen final
    if resultados['grupos_procesados'] > 0:
        if resultados['grupos_fallidos'] == 0:
            resultados['mensaje'] = f"✅ Generación exitosa: {resultados['grupos_procesados']} grupos, {resultados['horarios_generados']} horarios"
        else:
            resultados['exito'] = False
            resultados['mensaje'] = f"⚠️ Generación parcial: {resultados['grupos_procesados']} exitosos, {resultados['grupos_fallidos']} fallidos"
    else:
        resultados['exito'] = False
        resultados['mensaje'] = "❌ No se pudo generar ningún horario"
    
    return resultados

@app.route('/api/generacion-progreso')
@login_required
def api_generacion_progreso():
    """API endpoint para consultar el progreso de generación masiva"""
    global generacion_progreso
    with progreso_lock:
        return jsonify(generacion_progreso.copy())

@app.route('/api/iniciar-generacion-masiva', methods=['POST'])
@login_required
def api_iniciar_generacion_masiva():
    """API endpoint para iniciar generación masiva en segundo plano"""
    global generacion_progreso, progreso_lock
    
    if not current_user.is_admin():
        return jsonify({'error': 'No tienes permisos'}), 403
    
    data = request.get_json()
    grupos_ids = data.get('grupos_ids', [])
    version_nombre = data.get('version_nombre', '')
    dias_config = data.get('dias_semana', 'lunes_viernes')
    
    if not grupos_ids:
        return jsonify({'error': 'Selecciona al menos un grupo'}), 400
    
    # Convertir a enteros
    grupos_ids = [int(gid) for gid in grupos_ids]
    
    # IMPORTANTE: Guardar el user_id ANTES de crear el thread
    # porque current_user no está disponible en threads separados
    user_id = current_user.id
    
    # Configurar días
    if dias_config == 'lunes_viernes':
        dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
    elif dias_config == 'lunes_sabado':
        dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
    else:
        dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
    
    # Calcular período académico
    año_actual = datetime.now().year
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    periodo_academico = f"{año_actual}-{año_actual}_{timestamp}"
    
    if not version_nombre:
        version_nombre = f"Generación Masiva {timestamp}"
    
    # Inicializar progreso
    with progreso_lock:
        generacion_progreso = {
            'activo': True,
            'total_grupos': len(grupos_ids),
            'grupos_procesados': 0,
            'grupo_actual': 0,
            'codigo_grupo': 'Iniciando...',
            'horarios_generados': 0,
            'mensaje': 'Iniciando generación...',
            'completado': False,
            'exito': False
        }
    
    # Función para ejecutar en thread
    def ejecutar_generacion():
        global generacion_progreso, progreso_lock
        with app.app_context():
            resultado = generar_horarios_masivos_con_progreso(
                grupos_ids=grupos_ids,
                periodo_academico=periodo_academico,
                version_nombre=version_nombre,
                creado_por=user_id,  # Usar la variable capturada, no current_user
                dias_semana=dias_semana
            )
            
            # Marcar como completado
            with progreso_lock:
                generacion_progreso['completado'] = True
                generacion_progreso['exito'] = resultado['exito']
                generacion_progreso['mensaje'] = resultado['mensaje']
                generacion_progreso['activo'] = False
    
    # Iniciar thread
    thread = threading.Thread(target=ejecutar_generacion)
    thread.daemon = True
    thread.start()
    
    return jsonify({'mensaje': 'Generación iniciada', 'total_grupos': len(grupos_ids)})

@app.route('/admin/horarios-academicos/generar', methods=['GET', 'POST'])
@login_required
def generar_horarios_academicos():
    """Generar horarios académicos automáticamente"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    form = GenerarHorariosForm()
    resultado = None

    # Cargar grupos activos
    grupos = Grupo.query.filter_by(activo=True).order_by(Grupo.codigo).all()
    form.grupo_id.choices = [(0, 'Seleccione un grupo')] + [
        (g.id, f"{g.codigo} - {g.get_carrera_nombre()} - Cuatri {g.cuatrimestre} - {g.get_turno_display()}") 
        for g in grupos
    ]

    if form.validate_on_submit():
        from generador_horarios import generar_horarios_automaticos
        from datetime import datetime

        # Obtener el grupo seleccionado
        grupo_id = form.grupo_id.data
        
        if grupo_id == 0:
            flash('Debe seleccionar un grupo válido.', 'error')
            return render_template('admin/generar_horarios.html', form=form, resultado=resultado)
        
        grupo = Grupo.query.get(grupo_id)
        
        if not grupo:
            flash('Grupo no encontrado.', 'error')
            return render_template('admin/generar_horarios.html', form=form, resultado=resultado)

        # Calcular período académico y versión
        año_actual = datetime.now().year
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Si el usuario proporcionó un nombre de versión, usarlo; si no, generar automáticamente
        if form.version_nombre.data and form.version_nombre.data.strip():
            version_nombre = form.version_nombre.data.strip()
        else:
            version_nombre = f"Generación {timestamp}"
        
        # El periodo_academico sigue siendo único para evitar conflictos
        periodo_academico = f"{año_actual}-{año_actual}_{timestamp}"

        # Preparar días de la semana
        dias_semana = []
        if form.dias_semana.data == 'lunes_viernes':
            dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
        elif form.dias_semana.data == 'lunes_sabado':
            dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
        else:  # personalizado
            if form.lunes.data == 'si':
                dias_semana.append('lunes')
            if form.martes.data == 'si':
                dias_semana.append('martes')
            if form.miercoles.data == 'si':
                dias_semana.append('miercoles')
            if form.jueves.data == 'si':
                dias_semana.append('jueves')
            if form.viernes.data == 'si':
                dias_semana.append('viernes')
            if form.sabado.data == 'si':
                dias_semana.append('sabado')

        # Generar horarios usando el nuevo enfoque basado en grupos
        resultado = generar_horarios_automaticos(
            grupo_id=grupo_id,
            dias_semana=dias_semana,
            periodo_academico=periodo_academico,
            version_nombre=version_nombre,
            creado_por=current_user.id
        )

        if resultado['exito']:
            # No redirigir automáticamente, mostrar resultados en la página
            pass
        else:
            flash(resultado['mensaje'], 'error')

    return render_template('admin/generar_horarios.html', form=form, resultado=resultado)

@app.route('/admin/horarios-academicos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_horario_academico(id):
    """Editar un horario académico"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    horario_academico = HorarioAcademico.query.get_or_404(id)

    form = EditarHorarioAcademicoForm()

    # Cargar opciones para los select
    profesores = User.query.filter(
        User.rol.in_(['profesor_completo', 'profesor_asignatura']),
        User.activo == True
    ).order_by(User.nombre, User.apellido).all()

    horarios = Horario.query.filter_by(activo=True).order_by(Horario.orden).all()

    form.profesor_id.choices = [(str(p.id), p.get_nombre_completo()) for p in profesores]
    form.horario_id.choices = [(str(h.id), f"{h.nombre} ({h.get_hora_inicio_str()}-{h.get_hora_fin_str()})") for h in horarios]

    if form.validate_on_submit():
        horario_academico.profesor_id = int(form.profesor_id.data)
        horario_academico.horario_id = int(form.horario_id.data)
        horario_academico.dia_semana = form.dia_semana.data
        horario_academico.grupo = form.grupo.data
        # El periodo_academico se calcula automáticamente en el modelo

        db.session.commit()
        flash('Horario académico actualizado exitosamente.', 'success')
        return redirect(url_for('gestionar_horarios_academicos'))

    # Pre-llenar formulario
    form.profesor_id.data = str(horario_academico.profesor_id)
    form.horario_id.data = str(horario_academico.horario_id)
    form.dia_semana.data = horario_academico.dia_semana
    form.grupo.data = horario_academico.grupo

    return render_template('admin/editar_horario_academico.html',
                         form=form,
                         horario_academico=horario_academico)

@app.route('/admin/horarios-academicos/<int:id>/eliminar', methods=['GET', 'POST'])
@login_required
def eliminar_horario_academico(id):
    """Eliminar un horario académico"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    horario_academico = HorarioAcademico.query.get_or_404(id)

    form = EliminarHorarioAcademicoForm()

    if form.validate_on_submit():
        if form.confirmacion.data == 'ELIMINAR':
            horario_academico.activo = False
            db.session.commit()
            flash('Horario académico eliminado exitosamente.', 'success')
            return redirect(url_for('gestionar_horarios_academicos'))
        else:
            flash('Confirmación incorrecta.', 'error')

    return render_template('admin/eliminar_horario_academico.html',
                         form=form,
                         horario_academico=horario_academico)

@app.route('/admin/horarios-academicos/grupo/<int:grupo_id>/eliminar', methods=['GET'])
@login_required
def eliminar_horarios_grupo(grupo_id):
    """Eliminar todos los horarios de un grupo"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    grupo = Grupo.query.get_or_404(grupo_id)
    
    # Obtener los IDs de todas las materias del grupo
    materias_ids = [materia.id for materia in grupo.materias]
    
    if materias_ids:
        # Eliminar todos los horarios de esas materias
        num_eliminados = HorarioAcademico.query.filter(
            HorarioAcademico.materia_id.in_(materias_ids)
        ).delete(synchronize_session=False)
        
        db.session.commit()
        
        flash(f'Se eliminaron {num_eliminados} horarios del grupo {grupo.codigo} exitosamente.', 'success')
    else:
        flash(f'El grupo {grupo.codigo} no tiene materias asignadas.', 'info')
    
    return redirect(url_for('gestionar_horarios_academicos'))

@app.route('/admin/horarios-academicos/<int:id>/eliminar-rapido', methods=['POST'])
@login_required
def eliminar_horario_rapido(id):
    """Eliminar un horario académico de forma rápida (sin confirmación de formulario)"""
    if not current_user.is_admin():
        flash('No tienes permisos para realizar esta acción.', 'error')
        return redirect(url_for('dashboard'))

    horario_academico = HorarioAcademico.query.get_or_404(id)
    
    # Guardar información para el mensaje
    materia_nombre = horario_academico.get_materia_nombre()
    dia = horario_academico.get_dia_display()
    hora = f"{horario_academico.get_hora_inicio_str()} - {horario_academico.get_hora_fin_str()}"
    
    # Eliminar el horario
    db.session.delete(horario_academico)
    db.session.commit()
    
    flash(f'Horario eliminado: {materia_nombre} - {dia} {hora}', 'success')
    
    # Determinar de dónde vino la solicitud para redirigir correctamente
    referer = request.referrer
    if referer and 'grupos' in referer:
        return redirect(url_for('admin_horario_grupos'))
    elif referer and 'profesores' in referer:
        return redirect(url_for('admin_horario_profesores'))
    else:
        return redirect(url_for('gestionar_horarios_academicos'))

# ==========================================
# GESTIÓN DE USUARIOS (SOLO ADMINISTRADORES)
# ==========================================

@app.route('/admin/usuarios')
@login_required
def gestionar_usuarios():
    """Gestión de usuarios (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    # Obtener filtros
    rol_filter = request.args.get('rol', '')
    activo_filter = request.args.get('activo', '')
    busqueda = request.args.get('busqueda', '')

    # Construir consulta base
    query = User.query

    # Aplicar filtros
    if rol_filter:
        if rol_filter == 'profesor':
            query = query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura']))
        else:
            query = query.filter_by(rol=rol_filter)

    if activo_filter:
        if activo_filter == 'activos':
            query = query.filter_by(activo=True)
        elif activo_filter == 'inactivos':
            query = query.filter_by(activo=False)

    if busqueda:
        query = query.filter(
            db.or_(
                User.nombre.contains(busqueda),
                User.apellido.contains(busqueda),
                User.username.contains(busqueda),
                User.email.contains(busqueda)
            )
        )

    usuarios = query.order_by(User.nombre, User.apellido).all()

    # Estadísticas
    total_usuarios = len(usuarios)
    usuarios_activos = sum(1 for u in usuarios if u.activo)
    usuarios_inactivos = total_usuarios - usuarios_activos

    # Conteo por roles
    roles_count = {}
    for usuario in usuarios:
        rol_display = usuario.get_rol_display()
        roles_count[rol_display] = roles_count.get(rol_display, 0) + 1

    return render_template('admin/usuarios.html',
                         usuarios=usuarios,
                         total_usuarios=total_usuarios,
                         usuarios_activos=usuarios_activos,
                         usuarios_inactivos=usuarios_inactivos,
                         roles_count=roles_count,
                         filtros_activos={
                             'rol': rol_filter,
                             'activo': activo_filter,
                             'busqueda': busqueda
                         })

@app.route('/admin/usuario/nuevo', methods=['GET', 'POST'])
@login_required
def agregar_usuario():
    """Agregar nuevo usuario (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    from forms import AgregarUsuarioForm
    form = AgregarUsuarioForm()
    
    # Obtener horarios para mostrar en la tabla de disponibilidad
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()

    if form.validate_on_submit():
        # Obtener roles seleccionados del formulario
        roles_seleccionados = form.get_roles_list()
        
        # Obtener el rol principal para el campo legacy
        rol_principal = form.get_primary_rol()
        
        # Crear nuevo usuario
        nuevo_usuario = User(
            username=form.username.data,
            email=form.email.data,
            password=form.password.data,
            nombre=form.nombre.data,
            apellido=form.apellido.data,
            rol=rol_principal,  # Campo legacy
            telefono=form.telefono.data
        )
        
        # Asignar estado activo después de la creación
        nuevo_usuario.activo = form.activo.data
        
        # Asignar carreras si el usuario tiene roles que requieren carrera
        requiere_carrera = any(r in roles_seleccionados for r in ['jefe_carrera', 'profesor_completo', 'profesor_asignatura'])
        if requiere_carrera and form.carreras.data:
            from models import Carrera
            carreras_seleccionadas = Carrera.query.filter(Carrera.id.in_(form.carreras.data)).all()
            nuevo_usuario.carreras = carreras_seleccionadas

        try:
            db.session.add(nuevo_usuario)
            db.session.flush()  # Para obtener el ID del nuevo usuario
            
            # Agregar todos los roles seleccionados a la relación many-to-many
            from models import Role
            for rol_nombre in roles_seleccionados:
                nuevo_usuario.add_role(rol_nombre)
            
            # Procesar disponibilidad horaria si el usuario es profesor
            es_profesor = 'profesor_completo' in roles_seleccionados or 'profesor_asignatura' in roles_seleccionados
            if es_profesor:
                disponibilidades_data = form.get_disponibilidades_data()
                
                for disp_data in disponibilidades_data:
                    nueva_disponibilidad = DisponibilidadProfesor(
                        profesor_id=nuevo_usuario.id,
                        horario_id=disp_data['horario_id'],
                        dia_semana=disp_data['dia_semana'],
                        disponible=disp_data['disponible'],
                        creado_por=current_user.id
                    )
                    db.session.add(nueva_disponibilidad)
            
            db.session.commit()
            
            # Construir mensaje con todos los roles
            roles_display = [Role.ROLES_DISPLAY.get(r, r) for r in roles_seleccionados]
            flash(f'Usuario {nuevo_usuario.get_nombre_completo()} creado exitosamente con roles: {", ".join(roles_display)}.', 'success')
            return redirect(url_for('gestionar_usuarios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear usuario: {str(e)}', 'error')

    return render_template('admin/usuario_form.html', form=form, titulo="Agregar Usuario", usuario=None, horarios=horarios, disponibilidad_dict={})

@app.route('/admin/usuario/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_usuario(id):
    """Editar usuario existente (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    usuario = User.query.get_or_404(id)
    from forms import EditarUsuarioForm
    form = EditarUsuarioForm(user=usuario)
    
    # Obtener horarios para mostrar en la tabla de disponibilidad
    horarios = Horario.query.filter_by(activo=True).order_by(Horario.turno, Horario.orden).all()

    if form.validate_on_submit():
        # Obtener roles seleccionados del formulario
        roles_seleccionados = form.get_roles_list()
        
        # Obtener el rol principal para el campo legacy
        rol_principal = form.get_primary_rol()
        
        # Actualizar usuario
        usuario.username = form.username.data
        usuario.email = form.email.data
        usuario.nombre = form.nombre.data
        usuario.apellido = form.apellido.data
        usuario.rol = rol_principal  # Campo legacy
        usuario.telefono = form.telefono.data
        usuario.activo = form.activo.data
        
        # Limpiar roles anteriores y agregar los nuevos
        from models import Role
        usuario.roles.clear()  # Limpiar roles actuales
        for rol_nombre in roles_seleccionados:
            usuario.add_role(rol_nombre)
        
        # Asignar carreras si el usuario tiene roles que requieren carrera
        requiere_carrera = any(r in roles_seleccionados for r in ['jefe_carrera', 'profesor_completo', 'profesor_asignatura'])
        if requiere_carrera:
            from models import Carrera
            if form.carreras.data:
                carreras_seleccionadas = Carrera.query.filter(Carrera.id.in_(form.carreras.data)).all()
                usuario.carreras = carreras_seleccionadas
            else:
                usuario.carreras = []
        else:
            # Si no tiene roles que requieren carrera, limpiar carreras
            usuario.carreras = []
        
        # Procesar disponibilidad horaria si el usuario es profesor
        es_profesor = 'profesor_completo' in roles_seleccionados or 'profesor_asignatura' in roles_seleccionados
        if es_profesor:
            disponibilidades_data = form.get_disponibilidades_data()
            
            # Desactivar disponibilidades anteriores (mantener historial)
            DisponibilidadProfesor.query.filter_by(
                profesor_id=usuario.id,
                activo=True
            ).update({'activo': False})
            
            # Crear nuevas disponibilidades
            for disp_data in disponibilidades_data:
                nueva_disponibilidad = DisponibilidadProfesor(
                    profesor_id=usuario.id,
                    horario_id=disp_data['horario_id'],
                    dia_semana=disp_data['dia_semana'],
                    disponible=disp_data['disponible'],
                    creado_por=current_user.id
                )
                db.session.add(nueva_disponibilidad)

        try:
            db.session.commit()
            
            # Construir mensaje con todos los roles
            roles_display = [Role.ROLES_DISPLAY.get(r, r) for r in roles_seleccionados]
            flash(f'Usuario {usuario.get_nombre_completo()} actualizado exitosamente con roles: {", ".join(roles_display)}.', 'success')
            return redirect(url_for('gestionar_usuarios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar usuario: {str(e)}', 'error')

    # Llenar formulario con datos actuales
    elif request.method == 'GET':
        form.username.data = usuario.username
        form.email.data = usuario.email
        form.nombre.data = usuario.nombre
        form.apellido.data = usuario.apellido
        form.telefono.data = usuario.telefono
        form.activo.data = usuario.activo
        
        # Configurar los roles seleccionados (nuevo sistema)
        form.roles_seleccionados.data = usuario.get_roles_list()
        
        # Cargar carreras
        form.carreras.data = [c.id for c in usuario.carreras]
        
        # Mantener compatibilidad con el campo legacy
        form.rol.data = usuario.rol
    
    # Cargar disponibilidades actuales del profesor
    disponibilidad_dict = {}
    roles_actuales = usuario.get_roles_list()
    es_profesor_actual = 'profesor_completo' in roles_actuales or 'profesor_asignatura' in roles_actuales
    if es_profesor_actual:
        disponibilidades_actuales = DisponibilidadProfesor.query.filter_by(
            profesor_id=usuario.id,
            activo=True
        ).all()
        
        for disp in disponibilidades_actuales:
            if disp.disponible:
                disponibilidad_dict[(disp.horario_id, disp.dia_semana)] = True

    return render_template('admin/usuario_form.html', form=form, titulo="Editar Usuario", usuario=usuario, horarios=horarios, disponibilidad_dict=disponibilidad_dict)

@app.route('/admin/usuario/<int:id>/eliminar', methods=['GET', 'POST'])
@login_required
def eliminar_usuario(id):
    """Eliminar usuario (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    usuario = User.query.get_or_404(id)

    # No permitir eliminar al propio usuario
    if usuario.id == current_user.id:
        flash('No puedes eliminar tu propio usuario.', 'error')
        return redirect(url_for('gestionar_usuarios'))

    from forms import EliminarUsuarioForm
    form = EliminarUsuarioForm()

    if form.validate_on_submit():
        if form.confirmacion.data == 'ELIMINAR':
            try:
                # Eliminar físicamente el usuario y sus relaciones
                # Primero eliminar las relaciones que podrían causar problemas
                
                # Eliminar asignaciones profesor-grupo donde el usuario es profesor
                from models import HorarioAcademico, DisponibilidadProfesor, AsignacionProfesorGrupo
                AsignacionProfesorGrupo.query.filter_by(profesor_id=usuario.id).delete()
                AsignacionProfesorGrupo.query.filter_by(creado_por=usuario.id).delete()
                
                # Eliminar horarios académicos donde el usuario es profesor
                HorarioAcademico.query.filter_by(profesor_id=usuario.id).delete()
                HorarioAcademico.query.filter_by(creado_por=usuario.id).delete()
                
                # Eliminar disponibilidades del profesor
                DisponibilidadProfesor.query.filter_by(profesor_id=usuario.id).delete()
                DisponibilidadProfesor.query.filter_by(creado_por=usuario.id).delete()
                
                # Eliminar el usuario físicamente
                db.session.delete(usuario)
                db.session.commit()
                
                flash(f'Usuario {usuario.get_nombre_completo()} eliminado permanentemente.', 'success')
                return redirect(url_for('gestionar_usuarios'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error al eliminar usuario: {str(e)}', 'error')
        else:
            flash('Confirmación incorrecta.', 'error')

    return render_template('admin/eliminar_usuario.html', form=form, usuario=usuario)

@app.route('/admin/usuario/<int:id>/activar', methods=['POST'])
@login_required
def activar_usuario(id):
    """Activar usuario (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    usuario = User.query.get_or_404(id)

    try:
        usuario.activo = True
        db.session.commit()
        flash(f'Usuario {usuario.get_nombre_completo()} activado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al activar usuario: {str(e)}', 'error')

    return redirect(url_for('gestionar_usuarios'))

@app.route('/admin/usuario/<int:id>/desactivar', methods=['POST'])
@login_required
def desactivar_usuario(id):
    """Desactivar usuario (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    usuario = User.query.get_or_404(id)

    # No permitir desactivar al propio usuario
    if usuario.id == current_user.id:
        flash('No puedes desactivar tu propio usuario.', 'error')
        return redirect(url_for('dashboard'))

    try:
        usuario.activo = False
        db.session.commit()
        flash(f'Usuario {usuario.get_nombre_completo()} desactivado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al desactivar usuario: {str(e)}', 'error')

    return redirect(url_for('gestionar_usuarios'))

@app.route('/admin/usuario/<int:id>/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password_usuario(id):
    """Cambiar contraseña de un usuario (solo admin)"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = User.query.get_or_404(id)
    
    form = CambiarPasswordProfesorForm()
    
    if form.validate_on_submit():
        try:
            # Cambiar la contraseña del usuario
            usuario.password = form.nueva_password.data
            db.session.commit()
            
            flash(f'Contraseña actualizada exitosamente para {usuario.get_nombre_completo()}.', 'success')
            return redirect(url_for('gestionar_usuarios'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al cambiar la contraseña: {str(e)}', 'error')
    
    return render_template('admin/cambiar_password_usuario.html',
                         form=form,
                         usuario=usuario,
                         titulo=f"Cambiar Contraseña - {usuario.get_nombre_completo()}")

# Reportes del Sistema
@app.route('/admin/reportes')
@login_required
def reportes_sistema():
    """Página de reportes del sistema"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    # Estadísticas generales
    total_usuarios = User.query.count()
    total_profesores = User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).count()
    total_carreras = Carrera.query.filter_by(activa=True).count()
    total_materias = Materia.query.filter_by(activa=True).count()
    total_horarios_academicos = HorarioAcademico.query.filter_by(activo=True).count()

    # Estadísticas por carrera
    carreras_stats = []
    carreras = Carrera.query.filter_by(activa=True).all()
    for carrera in carreras:
        profesores_carrera = User.query.filter(
            User.carrera_id == carrera.id,
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.activo == True
        ).count()

        materias_carrera = Materia.query.filter(
            Materia.carrera_id == carrera.id,
            Materia.activa == True
        ).count()

        horarios_carrera = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id == carrera.id,
            HorarioAcademico.activo == True
        ).count()

        carreras_stats.append({
            'carrera': carrera,
            'profesores': profesores_carrera,
            'materias': materias_carrera,
            'horarios': horarios_carrera
        })

    return render_template('admin/reportes.html',
                         total_usuarios=total_usuarios,
                         total_profesores=total_profesores,
                         total_carreras=total_carreras,
                         total_materias=total_materias,
                         total_horarios_academicos=total_horarios_academicos,
                         carreras_stats=carreras_stats)

# Configuración del Sistema
@app.route('/admin/configuracion')
@login_required
def configuracion_sistema():
    """Página de configuración del sistema"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    from models import ConfiguracionSistema
    
    # Obtener configuración de carga horaria de profesores
    config_horas = {
        'tiempo_completo': ConfiguracionSistema.get_config('horas_tiempo_completo', 40),
        'asignatura': ConfiguracionSistema.get_config('horas_asignatura', 20),
        'medio_tiempo': ConfiguracionSistema.get_config('horas_medio_tiempo', 20),
        'limite_absoluto': ConfiguracionSistema.get_config('horas_limite_absoluto', 50)
    }

    return render_template('admin/configuracion.html', config_horas=config_horas)


# API para configuración de carga horaria de profesores
@app.route('/admin/configuracion/carga-horaria', methods=['POST'])
@login_required
def guardar_configuracion_carga_horaria():
    """Guardar configuración de límites de carga horaria por tipo de profesor"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        from models import ConfiguracionSistema
        data = request.get_json()
        
        # Validar que el límite absoluto no sea excedido
        limite_absoluto = int(data.get('limite_absoluto', 50))
        tiempo_completo = int(data.get('tiempo_completo', 40))
        asignatura = int(data.get('asignatura', 20))
        medio_tiempo = int(data.get('medio_tiempo', 20))
        
        # Validar límite absoluto entre 20 y 60
        if limite_absoluto < 20 or limite_absoluto > 60:
            return jsonify({'success': False, 'message': 'El límite absoluto debe estar entre 20 y 60 horas'}), 400
        
        # Validar que ningún tipo exceda el límite absoluto
        if tiempo_completo > limite_absoluto:
            return jsonify({'success': False, 'message': 'Las horas de Tiempo Completo no pueden exceder el límite absoluto'}), 400
        if asignatura > limite_absoluto:
            return jsonify({'success': False, 'message': 'Las horas de Asignatura no pueden exceder el límite absoluto'}), 400
        if medio_tiempo > limite_absoluto:
            return jsonify({'success': False, 'message': 'Las horas de Medio Tiempo no pueden exceder el límite absoluto'}), 400
        
        # Guardar configuraciones
        ConfiguracionSistema.set_config(
            'horas_tiempo_completo', tiempo_completo,
            tipo='int', descripcion='Horas máximas semanales para profesores de tiempo completo', 
            categoria='profesores'
        )
        ConfiguracionSistema.set_config(
            'horas_asignatura', asignatura,
            tipo='int', descripcion='Horas máximas semanales para profesores por asignatura', 
            categoria='profesores'
        )
        ConfiguracionSistema.set_config(
            'horas_medio_tiempo', medio_tiempo,
            tipo='int', descripcion='Horas máximas semanales para profesores de medio tiempo', 
            categoria='profesores'
        )
        ConfiguracionSistema.set_config(
            'horas_limite_absoluto', limite_absoluto,
            tipo='int', descripcion='Límite absoluto de horas semanales para cualquier profesor', 
            categoria='profesores'
        )
        
        return jsonify({
            'success': True, 
            'message': 'Configuración de carga horaria guardada exitosamente'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al guardar: {str(e)}'}), 500


# API para obtener la carga horaria de un profesor
@app.route('/api/profesor/<int:profesor_id>/carga-horaria')
@login_required
def api_carga_horaria_profesor(profesor_id):
    """Obtener información de carga horaria de un profesor"""
    try:
        profesor = User.query.get_or_404(profesor_id)
        puede_asignar, horas_actuales, limite, horas_disponibles = validar_carga_horaria_profesor(profesor)
        
        return jsonify({
            'success': True,
            'profesor': profesor.get_nombre_completo(),
            'tipo_profesor': profesor.tipo_profesor or 'No especificado',
            'horas_actuales': horas_actuales,
            'limite_horas': limite,
            'horas_disponibles': horas_disponibles,
            'puede_asignar': puede_asignar
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# API para configuración de base de datos
@app.route('/admin/configuracion/database', methods=['POST'])
@login_required
def guardar_configuracion_database():
    """Guardar configuración de base de datos"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        data = request.get_json()

        # Guardar configuración de base de datos
        from models import ConfiguracionSistema

        # Configuración general
        ConfiguracionSistema.set_config(
            'db_type', data.get('db_type', 'sqlite'),
            tipo='string', descripcion='Tipo de base de datos', categoria='database'
        )

        # Configuración específica por tipo de BD
        if data['db_type'] == 'sqlite':
            ConfiguracionSistema.set_config(
                'sqlite_path', data.get('sqlite_path', 'instance/sistema_academico.db'),
                tipo='string', descripcion='Ruta del archivo SQLite', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'sqlite_backup_freq', data.get('sqlite_backup_freq', 'daily'),
                tipo='string', descripcion='Frecuencia de backup SQLite', categoria='backup'
            )
        elif data['db_type'] == 'postgresql':
            ConfiguracionSistema.set_config(
                'pg_host', data.get('host', 'localhost'),
                tipo='string', descripcion='Host PostgreSQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'pg_port', data.get('port', '5432'),
                tipo='string', descripcion='Puerto PostgreSQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'pg_database', data.get('database', 'sistema_academico'),
                tipo='string', descripcion='Base de datos PostgreSQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'pg_username', data.get('username', 'postgres'),
                tipo='string', descripcion='Usuario PostgreSQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'pg_password', data.get('password', ''),
                tipo='string', descripcion='Contraseña PostgreSQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'pg_ssl', data.get('ssl', 'require'),
                tipo='string', descripcion='SSL PostgreSQL', categoria='database'
            )
        elif data['db_type'] == 'mysql':
            ConfiguracionSistema.set_config(
                'mysql_host', data.get('host', 'localhost'),
                tipo='string', descripcion='Host MySQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mysql_port', data.get('port', '3306'),
                tipo='string', descripcion='Puerto MySQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mysql_database', data.get('database', 'sistema_academico'),
                tipo='string', descripcion='Base de datos MySQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mysql_username', data.get('username', 'root'),
                tipo='string', descripcion='Usuario MySQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mysql_password', data.get('password', ''),
                tipo='string', descripcion='Contraseña MySQL', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mysql_ssl', data.get('ssl', 'true'),
                tipo='string', descripcion='SSL MySQL', categoria='database'
            )
        elif data['db_type'] == 'mariadb':
            ConfiguracionSistema.set_config(
                'mariadb_host', data.get('host', 'localhost'),
                tipo='string', descripcion='Host MariaDB', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mariadb_port', data.get('port', '3306'),
                tipo='string', descripcion='Puerto MariaDB', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mariadb_database', data.get('database', 'sistema_academico'),
                tipo='string', descripcion='Base de datos MariaDB', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mariadb_username', data.get('username', 'root'),
                tipo='string', descripcion='Usuario MariaDB', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mariadb_password', data.get('password', ''),
                tipo='string', descripcion='Contraseña MariaDB', categoria='database'
            )
            ConfiguracionSistema.set_config(
                'mariadb_ssl', data.get('ssl', 'true'),
                tipo='string', descripcion='SSL MariaDB', categoria='database'
            )

        # Configuración de backups
        ConfiguracionSistema.set_config(
            'backup_frequency', data.get('backup_frequency', 'daily'),
            tipo='string', descripcion='Frecuencia de backup automático', categoria='backup'
        )
        ConfiguracionSistema.set_config(
            'backup_retention', data.get('backup_retention', '30'),
            tipo='int', descripcion='Días de retención de backups', categoria='backup'
        )
        ConfiguracionSistema.set_config(
            'backup_location', data.get('backup_location', 'backups/'),
            tipo='string', descripcion='Ubicación de backups', categoria='backup'
        )

        return jsonify({'success': True, 'message': 'Configuración guardada exitosamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al guardar configuración: {str(e)}'}), 500

# API para configuración de horarios
@app.route('/admin/configuracion/horarios', methods=['POST'])
@login_required
def guardar_configuracion_horarios():
    """Guardar configuración de horarios"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        data = request.get_json()
        from models import ConfiguracionSistema

        # Guardar configuración de horarios
        ConfiguracionSistema.set_config(
            'horas_max_dia', data.get('horas_max_dia', '8'),
            tipo='int', descripcion='Máximo de horas de clase por día', categoria='horarios'
        )
        ConfiguracionSistema.set_config(
            'dias_clase', data.get('dias_clase', '5'),
            tipo='int', descripcion='Número de días con clases por semana', categoria='horarios'
        )
        ConfiguracionSistema.set_config(
            'duracion_clase', data.get('duracion_clase', '50'),
            tipo='int', descripcion='Duración de cada clase en minutos', categoria='horarios'
        )
        ConfiguracionSistema.set_config(
            'tiempo_entre_clases', data.get('tiempo_entre_clases', '10'),
            tipo='int', descripcion='Tiempo de descanso entre clases en minutos', categoria='horarios'
        )

        return jsonify({'success': True, 'message': 'Configuración de horarios guardada exitosamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al guardar configuración: {str(e)}'}), 500

# API para obtener configuración de horarios
@app.route('/admin/configuracion/horarios', methods=['GET'])
@login_required
def obtener_configuracion_horarios():
    """Obtener configuración de horarios"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        from models import ConfiguracionSistema

        config = {
            'horas_max_dia': ConfiguracionSistema.get_config('horas_max_dia', '8'),
            'dias_clase': ConfiguracionSistema.get_config('dias_clase', '5'),
            'duracion_clase': ConfiguracionSistema.get_config('duracion_clase', '50'),
            'tiempo_entre_clases': ConfiguracionSistema.get_config('tiempo_entre_clases', '10')
        }

        return jsonify({'success': True, 'data': config})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al obtener configuración: {str(e)}'}), 500

# API para crear backup
@app.route('/admin/configuracion/backup', methods=['POST'])
@login_required
def crear_backup():
    """Crear backup de la base de datos"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        from models import BackupHistory
        import shutil
        import hashlib
        from datetime import datetime

        # Crear directorio de backups si no existe
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        # Generar nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_{timestamp}.db'
        filepath = os.path.join(backup_dir, filename)

        # Copiar archivo de base de datos
        db_path = 'instance/sistema_academico.db'
        if os.path.exists(db_path):
            shutil.copy2(db_path, filepath)

            # Calcular tamaño y checksum
            file_size = os.path.getsize(filepath)
            with open(filepath, 'rb') as f:
                checksum = hashlib.sha256(f.read()).hexdigest()

            # Registrar en historial
            backup = BackupHistory(
                filename=filename,
                tipo='manual',
                tamano=file_size,
                ruta_archivo=filepath,
                usuario_id=current_user.id
            )
            backup.checksum = checksum
            db.session.add(backup)
            db.session.commit()

            return jsonify({
                'success': True,
                'message': 'Backup creado exitosamente',
                'filename': filename,
                'size': file_size
            })
        else:
            return jsonify({'success': False, 'message': 'Archivo de base de datos no encontrado'}), 404

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al crear backup: {str(e)}'}), 500

# API para obtener historial de backups
@app.route('/admin/configuracion/backups', methods=['GET'])
@login_required
def obtener_historial_backups():
    """Obtener historial de backups"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        from models import BackupHistory

        backups = BackupHistory.query.order_by(BackupHistory.fecha_creacion.desc()).limit(10).all()

        backup_list = []
        for backup in backups:
            backup_list.append({
                'id': backup.id,
                'filename': backup.filename,
                'tipo': backup.tipo,
                'tamano': backup.get_tamano_formateado(),
                'fecha': backup.get_fecha_formateada(),
                'estado': backup.estado,
                'usuario': backup.usuario.nombre + ' ' + backup.usuario.apellido if backup.usuario else 'Sistema'
            })

        return jsonify({'success': True, 'backups': backup_list})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al obtener historial: {str(e)}'}), 500

# API para descargar backup
@app.route('/admin/configuracion/backup/<filename>', methods=['GET'])
@login_required
def descargar_backup(filename):
    """Descargar archivo de backup"""
    if not current_user.is_admin():
        flash('No tienes permisos para esta acción.', 'error')
        return redirect(url_for('dashboard'))

    try:
        from models import BackupHistory

        # Verificar que el backup existe en la base de datos
        backup = BackupHistory.query.filter_by(filename=filename).first()
        if not backup:
            flash('Backup no encontrado.', 'error')
            return redirect(url_for('configuracion_sistema'))

        filepath = os.path.join('backups', filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            flash('Archivo de backup no encontrado en el servidor.', 'error')
            return redirect(url_for('configuracion_sistema'))

    except Exception as e:
        flash(f'Error al descargar backup: {str(e)}', 'error')
        return redirect(url_for('configuracion_sistema'))

# API para optimizar base de datos
@app.route('/admin/configuracion/optimize', methods=['POST'])
@login_required
def optimizar_base_datos():
    """Optimizar base de datos"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        # Ejecutar comandos de optimización para SQLite
        db.session.execute('VACUUM')
        db.session.execute('ANALYZE')
        db.session.commit()

        return jsonify({'success': True, 'message': 'Base de datos optimizada exitosamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al optimizar base de datos: {str(e)}'}), 500

# API para reiniciar sistema (simulado)
@app.route('/admin/configuracion/restart', methods=['POST'])
@login_required
def reiniciar_sistema():
    """Reiniciar sistema (simulado)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        # En un entorno real, aquí se reiniciaría el servidor
        # Por ahora solo devolvemos éxito
        return jsonify({'success': True, 'message': 'Sistema reiniciado exitosamente'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al reiniciar sistema: {str(e)}'}), 500

# API para probar conexión a base de datos
@app.route('/admin/configuracion/test-connection', methods=['POST'])
@login_required
def probar_conexion():
    """Probar conexión a base de datos"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'No tienes permisos para esta acción'}), 403

    try:
        data = request.get_json()
        db_type = data.get('db_type', 'sqlite')

        if db_type == 'sqlite':
            # Para SQLite solo verificar que el archivo existe
            db_path = data.get('sqlite_path', 'instance/sistema_academico.db')
            if os.path.exists(db_path):
                return jsonify({'success': True, 'message': 'Conexión a SQLite exitosa'})
            else:
                return jsonify({'success': False, 'message': 'Archivo de base de datos SQLite no encontrado'})

        elif db_type == 'postgresql':
            # Probar conexión PostgreSQL
            try:
                import psycopg2  # type: ignore[import-untyped]
            except ImportError:
                return jsonify({
                    'success': False, 
                    'message': 'Driver PostgreSQL no instalado. Instale con: pip install psycopg2-binary'
                })
            
            conn = psycopg2.connect(
                host=data.get('host', 'localhost'),
                port=data.get('port', '5432'),
                database=data.get('database', 'sistema_academico'),
                user=data.get('username', 'postgres'),
                password=data.get('password', ''),
                sslmode=data.get('ssl', 'require')
            )
            conn.close()
            return jsonify({'success': True, 'message': 'Conexión a PostgreSQL exitosa'})

        elif db_type in ['mysql', 'mariadb']:
            # Probar conexión MySQL/MariaDB
            try:
                import mysql.connector  # type: ignore[import-untyped]
            except ImportError:
                return jsonify({
                    'success': False, 
                    'message': 'Driver MySQL no instalado. Instale con: pip install mysql-connector-python'
                })
            
            conn = mysql.connector.connect(
                host=data.get('host', 'localhost'),
                port=int(data.get('port', '3306')),
                database=data.get('database', 'sistema_academico'),
                user=data.get('username', 'root'),
                password=data.get('password', ''),
                ssl_disabled=(data.get('ssl', 'true') == 'false')
            )
            conn.close()
            return jsonify({'success': True, 'message': f'Conexión a {db_type.title()} exitosa'})

        else:
            return jsonify({'success': False, 'message': 'Tipo de base de datos no soportado'})

    except ImportError as e:
        return jsonify({'success': False, 'message': f'Driver no instalado: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error de conexión: {str(e)}'})

# Subir imagen de perfil
@app.route('/subir_imagen_perfil', methods=['POST'])
@login_required
def subir_imagen_perfil():
    """Subir imagen de perfil del usuario"""
    if 'imagen_perfil' not in request.files:
        flash('No se encontró el archivo de imagen.', 'error')
        return redirect(url_for('dashboard'))

    file = request.files['imagen_perfil']
    
    if file.filename == '':
        flash('No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('dashboard'))

    if file and allowed_file(file.filename):
        # Generar nombre único para el archivo
        filename = secure_filename(f"{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        
        # Crear directorio si no existe
        upload_dir = os.path.join('static', 'uploads', 'perfiles')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Guardar archivo
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        # Actualizar usuario en base de datos
        # Primero eliminar imagen anterior si existe
        if current_user.imagen_perfil:
            old_file_path = os.path.join('static', 'uploads', 'perfiles', current_user.imagen_perfil)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        # Actualizar con nueva imagen
        current_user.imagen_perfil = filename
        db.session.commit()
        
        flash('Imagen de perfil actualizada exitosamente.', 'success')
    else:
        flash('Tipo de archivo no permitido. Solo se permiten imágenes (PNG, JPG, JPEG, GIF).', 'error')

    return redirect(url_for('dashboard'))

# Eliminar imagen de perfil
@app.route('/eliminar_imagen_perfil', methods=['POST'])
@login_required
def eliminar_imagen_perfil():
    """Eliminar imagen de perfil del usuario"""
    if current_user.imagen_perfil:
        # Eliminar archivo físico
        file_path = os.path.join('static', 'uploads', 'perfiles', current_user.imagen_perfil)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Actualizar base de datos
        current_user.imagen_perfil = None
        db.session.commit()
        
        flash('Imagen de perfil eliminada exitosamente.', 'success')
    else:
        flash('No hay imagen de perfil para eliminar.', 'warning')
    
    return redirect(url_for('dashboard'))

# Función auxiliar para verificar tipos de archivo permitidos
def allowed_file(filename):
    """Verificar si el archivo tiene una extensión permitida"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Exportar Reportes - PDF
@app.route('/admin/reportes/exportar/pdf')
@login_required
def exportar_reportes_pdf():
    """Exportar reportes del sistema a PDF"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    # Obtener datos para el reporte
    total_usuarios = User.query.count()
    total_profesores = User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).count()
    total_carreras = Carrera.query.filter_by(activa=True).count()
    total_materias = Materia.query.filter_by(activa=True).count()
    total_horarios_academicos = HorarioAcademico.query.filter_by(activo=True).count()

    # Datos por carrera
    carreras = Carrera.query.filter_by(activa=True).all()
    carreras_data = []
    for carrera in carreras:
        profesores_carrera = User.query.filter(
            User.carrera_id == carrera.id,
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.activo == True
        ).count()

        materias_carrera = Materia.query.filter(
            Materia.carrera_id == carrera.id,
            Materia.activa == True
        ).count()

        horarios_carrera = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id == carrera.id,
            HorarioAcademico.activo == True
        ).count()

        cobertura = (horarios_carrera / materias_carrera * 100) if materias_carrera > 0 else 0

        carreras_data.append([
            carrera.nombre,
            str(profesores_carrera),
            str(materias_carrera),
            str(horarios_carrera),
            ".1f"
        ])

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )

    # Título
    title = Paragraph("Reporte del Sistema Académico", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Fecha de generación
    from datetime import datetime
    fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=10, alignment=2)
    fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style)
    elements.append(fecha)
    elements.append(Spacer(1, 20))

    # Estadísticas Generales
    elements.append(Paragraph("Estadísticas Generales", styles['Heading2']))
    elements.append(Spacer(1, 12))

    general_data = [
        ['Métrica', 'Valor'],
        ['Total de Usuarios', str(total_usuarios)],
        ['Total de Profesores', str(total_profesores)],
        ['Carreras Activas', str(total_carreras)],
        ['Materias Activas', str(total_materias)],
        ['Horarios Asignados', str(total_horarios_academicos)],
        ['Promedio Horarios por Profesor', ".1f"]
    ]

    general_table = Table(general_data, colWidths=[3*inch, 2*inch])
    general_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(general_table)
    elements.append(Spacer(1, 20))

    # Estadísticas por Carrera
    elements.append(Paragraph("Estadísticas por Carrera", styles['Heading2']))
    elements.append(Spacer(1, 12))

    if carreras_data:
        carrera_headers = [['Carrera', 'Profesores', 'Materias', 'Horarios', 'Cobertura (%)']]
        carrera_table_data = carrera_headers + carreras_data

        carrera_table = Table(carrera_table_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
        carrera_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(carrera_table)
    else:
        elements.append(Paragraph("No hay datos de carreras disponibles.", styles['Normal']))

    # Generar PDF
    doc.build(elements)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_sistema_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
        mimetype='application/pdf'
    )

# Exportar Reportes - Excel
@app.route('/admin/reportes/exportar/excel')
@login_required
def exportar_reportes_excel():
    """Exportar reportes del sistema a Excel"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    # Crear datos para Excel
    data_general = {
        'Métrica': [
            'Total de Usuarios',
            'Total de Profesores',
            'Carreras Activas',
            'Materias Activas',
            'Horarios Asignados',
            'Promedio Horarios por Profesor'
        ],
        'Valor': [
            User.query.count(),
            User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).count(),
            Carrera.query.filter_by(activa=True).count(),
            Materia.query.filter_by(activa=True).count(),
            HorarioAcademico.query.filter_by(activo=True).count(),
            ".1f"
        ]
    }

    # Datos por carrera
    carreras = Carrera.query.filter_by(activa=True).all()
    carreras_data = []
    for carrera in carreras:
        profesores_carrera = User.query.filter(
            User.carrera_id == carrera.id,
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.activo == True
        ).count()

        materias_carrera = Materia.query.filter(
            Materia.carrera_id == carrera.id,
            Materia.activa == True
        ).count()

        horarios_carrera = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id == carrera.id,
            HorarioAcademico.activo == True
        ).count()

        cobertura = (horarios_carrera / materias_carrera * 100) if materias_carrera > 0 else 0

        carreras_data.append({
            'Carrera': carrera.nombre,
            'Profesores': profesores_carrera,
            'Materias': materias_carrera,
            'Horarios': horarios_carrera,
            'Cobertura (%)': round(cobertura, 1)
        })

    # Crear Excel con múltiples hojas
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Hoja de estadísticas generales
        df_general = pd.DataFrame(data_general)
        df_general.to_excel(writer, sheet_name='Estadísticas Generales', index=False)

        # Hoja de estadísticas por carrera
        if carreras_data:
            df_carreras = pd.DataFrame(carreras_data)
            df_carreras.to_excel(writer, sheet_name='Estadísticas por Carrera', index=False)

        # Hoja de resumen
        resumen_data = {
            'Información': [
                'Fecha de Generación',
                'Usuario que generó',
                'Total de Registros Analizados'
            ],
            'Valor': [
                datetime.now().strftime('%d/%m/%Y %H:%M'),
                current_user.get_nombre_completo(),
                len(carreras_data) if carreras_data else 0
            ]
        }
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_sistema_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Exportar Reportes - CSV
@app.route('/admin/reportes/exportar/csv')
@login_required
def exportar_reportes_csv():
    """Exportar reportes del sistema a CSV"""
    if not current_user.is_admin():
        flash('No tienes permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))

    # Crear datos para CSV usando StringIO
    buffer = StringIO()

    # Escribir estadísticas generales
    writer = csv.writer(buffer)

    # Encabezado
    writer.writerow(['Reporte del Sistema Académico'])
    writer.writerow(['Generado el', datetime.now().strftime('%d/%m/%Y %H:%M')])
    writer.writerow(['Usuario', current_user.get_nombre_completo()])
    writer.writerow([])

    # Estadísticas Generales
    writer.writerow(['ESTADÍSTICAS GENERALES'])
    writer.writerow(['Métrica', 'Valor'])
    writer.writerow(['Total de Usuarios', User.query.count()])
    writer.writerow(['Total de Profesores', User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).count()])
    writer.writerow(['Carreras Activas', Carrera.query.filter_by(activa=True).count()])
    writer.writerow(['Materias Activas', Materia.query.filter_by(activa=True).count()])
    writer.writerow(['Horarios Asignados', HorarioAcademico.query.filter_by(activo=True).count()])
    writer.writerow(['Promedio Horarios por Profesor', ".1f"])
    writer.writerow([])

    # Estadísticas por Carrera
    writer.writerow(['ESTADÍSTICAS POR CARRERA'])
    writer.writerow(['Carrera', 'Profesores', 'Materias', 'Horarios', 'Cobertura (%)'])

    carreras = Carrera.query.filter_by(activa=True).all()
    for carrera in carreras:
        profesores_carrera = User.query.filter(
            User.carrera_id == carrera.id,
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.activo == True
        ).count()

        materias_carrera = Materia.query.filter(
            Materia.carrera_id == carrera.id,
            Materia.activa == True
        ).count()

        horarios_carrera = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id == carrera.id,
            HorarioAcademico.activo == True
        ).count()

        cobertura = (horarios_carrera / materias_carrera * 100) if materias_carrera > 0 else 0

        writer.writerow([
            carrera.nombre,
            profesores_carrera,
            materias_carrera,
            horarios_carrera,
            f"{cobertura:.1f}"
        ])

    # Convertir StringIO a BytesIO para send_file
    buffer.seek(0)
    byte_buffer = BytesIO(buffer.getvalue().encode('utf-8-sig'))  # utf-8-sig para BOM (Excel compatibility)
    byte_buffer.seek(0)
    
    return send_file(
        byte_buffer,
        as_attachment=True,
        download_name=f'reporte_sistema_{datetime.now().strftime("%Y%m%d_%H%M")}.csv',
        mimetype='text/csv'
    )


# ==========================================
# FUNCIONES AUXILIARES PARA HORARIOS
# ==========================================

def obtener_dia_correcto(dia_semana):
    """Convierte el día de la semana de minúsculas sin tilde a formato con tilde"""
    dias_map = {
        'lunes': 'Lunes',
        'martes': 'Martes',
        'miercoles': 'Miércoles',
        'jueves': 'Jueves',
        'viernes': 'Viernes',
        'sabado': 'Sábado',
        'domingo': 'Domingo'
    }
    return dias_map.get(dia_semana, dia_semana.capitalize())


# ==========================================
# VISTA DE HORARIOS POR PROFESOR PARA ADMIN
# ==========================================
@app.route('/admin/horarios/profesores')
@login_required
def admin_horario_profesores():
    if not current_user.is_admin(): abort(403)
    
    horarios_data = procesar_horarios(agrupar_por='profesor', incluir_ids=True)
    return render_template('admin/admin_horario_profesores.html', horarios_data=horarios_data)

@app.route('/admin/horarios/grupos')
@login_required
def admin_horario_grupos():
    if not current_user.is_admin(): abort(403)
    
    # Obtener filtros de la query string
    carrera_id = request.args.get('carrera_id', type=int)
    cuatrimestre = request.args.get('cuatrimestre', type=int)
    
    horarios_data = procesar_horarios(agrupar_por='grupo', carrera_id=carrera_id, incluir_ids=True)
    
    # Ordenar los grupos por cuatrimestre (ascendente)
    # El código del grupo tiene formato: [cuatrimestre][turno][carrera][numero]
    # Ejemplo: 1MSC1 = cuatrimestre 1, 7VSC2 = cuatrimestre 7
    def extraer_cuatrimestre(grupo_codigo):
        """Extraer el cuatrimestre del código del grupo"""
        try:
            # Obtener dígitos iniciales
            num_str = ''
            for char in grupo_codigo:
                if char.isdigit():
                    num_str += char
                else:
                    break
            return int(num_str) if num_str else 99
        except:
            return 99
    
    # Filtrar por cuatrimestre si se especifica
    if cuatrimestre is not None:
        horarios_data = {
            k: v for k, v in horarios_data.items() 
            if extraer_cuatrimestre(k) == cuatrimestre
        }
    
    # Ordenar por cuatrimestre ascendente, luego por código
    horarios_ordenados = dict(sorted(
        horarios_data.items(), 
        key=lambda x: (extraer_cuatrimestre(x[0]), x[0])
    ))
    
    # Obtener carreras y cuatrimestres para filtros
    carreras = Carrera.query.filter_by(activa=True).order_by(Carrera.nombre).all()
    cuatrimestres = list(range(0, 11))  # 0 a 10
    
    return render_template('admin/admin_horario_grupos.html', 
                         horarios_data=horarios_ordenados,
                         carreras=carreras,
                         cuatrimestres=cuatrimestres,
                         filtro_carrera=carrera_id,
                         filtro_cuatrimestre=cuatrimestre)

# ==========================================
# EXPORTAR HORARIOS POR PROFESOR (WORD EXCEL)
# ==========================================

@app.route('/admin/horarios/profesores/exportar/excel')
@login_required
def exportar_horarios_profesor_excel():
    if not current_user.is_admin():
        abort(403)
    
    # Obtener horarios ordenados por día de semana y hora
    asignaciones = HorarioAcademico.query.filter_by(activo=True).all()
    
    # Ordenar por día de semana (Lunes a Viernes) y hora
    asignaciones.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
    
    horarios_por_profesor = {}
    
    # Mapeo de días en minúsculas a días con tilde
    dias_map = {
        'lunes': 'Lunes',
        'martes': 'Martes',
        'miercoles': 'Miércoles',
        'jueves': 'Jueves',
        'viernes': 'Viernes',
        'sabado': 'Sábado',
        'domingo': 'Domingo'
    }
    
    for a in asignaciones:
        if not a.profesor or not a.materia or not a.horario:
            continue 

        profesor_nombre = a.profesor.get_nombre_completo()
        if profesor_nombre not in horarios_por_profesor:
            horarios_por_profesor[profesor_nombre] = {'Lunes': [], 'Martes': [], 'Miércoles': [], 'Jueves': [], 'Viernes': []}
        
        info = f"{a.materia.nombre}\n{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
        
        # Usar el mapeo para obtener el día con tilde
        dia_correcto = dias_map.get(a.dia_semana, a.dia_semana.capitalize())
        if dia_correcto in horarios_por_profesor[profesor_nombre]:
            horarios_por_profesor[profesor_nombre][dia_correcto].append(info)
       


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horarios por Profesor"
    headers = ["Profesor", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
    for profesor, dias in horarios_por_profesor.items():
        ws.append([profesor] + ["\n\n".join(dias.get(dia, [])) for dia in headers[1:]])
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = 30
        for cell in col_cells: cell.alignment = Alignment(wrap_text=True, vertical='top')
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='horarios_por_profesor.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==========================================
# EXPORTAR HORARIOS POR PROFESOR (PDF)
# ==========================================

@app.route('/admin/horarios/profesores/exportar/pdf')
@login_required
def exportar_horarios_profesor_pdf():
    if not current_user.is_admin():
        abort(403)

    # Obtener horarios ordenados por día de semana y hora
    asignaciones = HorarioAcademico.query.filter_by(activo=True).all()
    
    # Ordenar por día de semana (Lunes a Viernes) y hora
    asignaciones.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
    
    horarios_por_profesor = {}
    
    # Mapeo de días en minúsculas a días con tilde
    dias_map = {
        'lunes': 'Lunes',
        'martes': 'Martes',
        'miercoles': 'Miércoles',
        'jueves': 'Jueves',
        'viernes': 'Viernes',
        'sabado': 'Sábado',
        'domingo': 'Domingo'
    }
    
    for a in asignaciones:
        if not a.profesor or not a.materia or not a.horario:
            continue

        profesor_nombre = a.profesor.get_nombre_completo()
        if profesor_nombre not in horarios_por_profesor:
            horarios_por_profesor[profesor_nombre] = {'Lunes': [], 'Martes': [], 'Miércoles': [], 'Jueves': [], 'Viernes': []}

        info = f"{a.materia.nombre}<br/>{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
        
        # Usar el mapeo para obtener el día con tilde
        dia_correcto = dias_map.get(a.dia_semana, a.dia_semana.capitalize())
        if dia_correcto in horarios_por_profesor[profesor_nombre]:
            horarios_por_profesor[profesor_nombre][dia_correcto].append(info)
       

    styles = getSampleStyleSheet(); styleN = styles['BodyText']
    data = [["Profesor", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]]
    for profesor, dias in horarios_por_profesor.items():
        row_data = [Paragraph(profesor, styleN)]
        for dia in data[0][1:]: row_data.append(Paragraph("<br/><br/>".join(dias.get(dia, [])), styleN))
        data.append(row_data)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    table = Table(data, hAlign='CENTER', colWidths=[1.5*inch]*6)
    style = TableStyle([('BACKGROUND', (0,0), (-1,0), colors.grey), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), ('GRID', (0,0), (-1,-1), 1, colors.black)])
    table.setStyle(style); doc.build([table]); buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='horarios_por_profesor.pdf', mimetype='application/pdf')

# ==========================================
# EXPORTAR HORARIOS POR GRUPO (EXCEL)
# ==========================================

@app.route('/admin/horarios/grupos/exportar/excel')
@login_required
def exportar_horarios_grupo_excel():
    if not current_user.is_admin():
        abort(403)
    
    # Obtener horarios ordenados por día de semana y hora
    asignaciones = HorarioAcademico.query.filter_by(activo=True).all()
    
    # Ordenar por día de semana (Lunes a Viernes) y hora
    asignaciones.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
    
    horarios_por_grupo = {}
    
    # Mapeo de días en minúsculas a días con tilde
    dias_map = {
        'lunes': 'Lunes',
        'martes': 'Martes',
        'miercoles': 'Miércoles',
        'jueves': 'Jueves',
        'viernes': 'Viernes',
        'sabado': 'Sábado',
        'domingo': 'Domingo'
    }
    
    for a in asignaciones:
        if not a.profesor or not a.materia or not a.horario:
            continue

        # CORREGIDO: Usar el campo 'grupo' del HorarioAcademico directamente
        grupo_nombre = a.grupo
        
        if grupo_nombre:
            if grupo_nombre not in horarios_por_grupo:
                horarios_por_grupo[grupo_nombre] = {'Lunes': [], 'Martes': [], 'Miércoles': [], 'Jueves': [], 'Viernes': []}
            
            info = f"{a.materia.nombre}\nProf: {a.profesor.get_nombre_completo()}\n{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
            
            # Usar el mapeo para obtener el día con tilde
            dia_correcto = dias_map.get(a.dia_semana, a.dia_semana.capitalize())
            if dia_correcto in horarios_por_grupo[grupo_nombre]:
                horarios_por_grupo[grupo_nombre][dia_correcto].append(info)
       

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Horarios por Grupo"
    headers = ["Grupo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
    for grupo, dias in horarios_por_grupo.items():
        ws.append([grupo] + ["\n\n".join(dias.get(dia, [])) for dia in headers[1:]])
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = 30
        for cell in col_cells: cell.alignment = Alignment(wrap_text=True, vertical='top')
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='horarios_por_grupo.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==========================================
# EXPORTAR HORARIOS POR GRUPO (PDF)
# ==========================================

@app.route('/admin/horarios/grupos/exportar/pdf')
@login_required
def exportar_horarios_grupo_pdf():
    if not current_user.is_admin():
        abort(403)
    
    # Obtener todos los grupos únicos con horarios
    grupos_con_horarios = db.session.query(HorarioAcademico.grupo).filter_by(activo=True).distinct().all()
    grupos_unicos = sorted([g[0] for g in grupos_con_horarios if g[0]])
    
    if not grupos_unicos:
        flash('No hay horarios para exportar.', 'warning')
        return redirect(url_for('admin_horario_grupos'))
    
    # Mapeo de días
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)
    
    for grupo in grupos_unicos:
        # Obtener horarios del grupo
        asignaciones = HorarioAcademico.query.filter_by(grupo=grupo, activo=True).all()
        
        # Determinar turno y sábado
        es_vespertino = 'V' in grupo[1:3]
        hora_inicio_turno = 13 if es_vespertino else 7
        hora_fin_turno = 21 if es_vespertino else 14
        tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
        
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        if tiene_sabado:
            dias_semana.append('Sábado')
        
        # Título del grupo
        elements.append(Paragraph(f"Horario: {grupo}", title_style))
        elements.append(Spacer(1, 10))
        
        # Crear tabla de cuadrícula
        data = [['Hora'] + dias_semana]
        
        for hora in range(hora_inicio_turno, hora_fin_turno):
            hora_str = f"{hora:02d}:00"
            row = [hora_str]
            
            for dia in dias_semana:
                # Buscar clase para esta hora y día
                contenido = ""
                for a in asignaciones:
                    dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                    if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                        contenido = f"{a.materia.nombre}<br/>Prof: {a.profesor.get_nombre_completo()}"
                        break
                row.append(Paragraph(contenido, cell_style))
            
            data.append(row)
        
        # Calcular anchos de columna
        col_widths = [0.8*inch] + [1.4*inch] * len(dias_semana)
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.118, 0.235, 0.447)),  # #1e3c72
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.97, 0.97, 0.98)),  # #f8f9fa
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.87, 0.89, 0.90)),  # #dee2e6
            ('ROWHEIGHTS', (0, 1), (-1, -1), 40),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 30))
    
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='horarios_por_grupo.pdf', mimetype='application/pdf')

# ==========================================
# EXPORTAR HORARIO INDIVIDUAL POR GRUPO (NUEVA RUTA)
# ==========================================
@app.route('/admin/horarios/grupos/exportar-individual/<nombre_grupo>')
@login_required
def exportar_horario_individual_excel(nombre_grupo):
    if not current_user.is_admin():
        abort(403)

    # 1. Obtener horarios del grupo específico
    asignaciones = HorarioAcademico.query.filter_by(grupo=nombre_grupo, activo=True).all()

    if not asignaciones:
        flash(f'No se encontró el horario para el grupo "{nombre_grupo}".', 'danger')
        return redirect(url_for('admin_horario_grupos'))

    # 2. Determinar turno (Matutino o Vespertino) y si tiene sábado
    es_vespertino = 'V' in nombre_grupo[1:3]
    hora_inicio_turno = 13 if es_vespertino else 7
    hora_fin_turno = 21 if es_vespertino else 14
    
    tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
    
    # 3. Crear estructura de datos por hora
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    if tiene_sabado:
        dias_semana.append('Sábado')
    
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

    # 4. Crear workbook con openpyxl para mejor formato
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Horario {nombre_grupo}'
    
    # Estilos
    header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    hora_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    hora_font = Font(bold=True, size=10)
    cell_border = Border(
        left=Side(style='thin', color='dee2e6'),
        right=Side(style='thin', color='dee2e6'),
        top=Side(style='thin', color='dee2e6'),
        bottom=Side(style='thin', color='dee2e6')
    )
    
    # Encabezados
    headers = ['Hora'] + dias_semana
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
    
    # Filas por hora
    row_num = 2
    for hora in range(hora_inicio_turno, hora_fin_turno):
        hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
        
        # Celda de hora
        cell = ws.cell(row=row_num, column=1, value=hora_str)
        cell.fill = hora_fill
        cell.font = hora_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
        
        # Celdas de cada día
        for col, dia in enumerate(dias_semana, 2):
            # Buscar clase para esta hora y día
            clase_encontrada = None
            for a in asignaciones:
                dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                    clase_encontrada = a
                    break
            
            if clase_encontrada:
                contenido = f"{clase_encontrada.materia.nombre}\nProf: {clase_encontrada.profesor.get_nombre_completo()}"
            else:
                contenido = ""
            
            cell = ws.cell(row=row_num, column=col, value=contenido)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
        
        row_num += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 30
    
    # Ajustar altura de filas
    for row in range(2, row_num):
        ws.row_dimensions[row].height = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'horario_{nombre_grupo.replace(" ", "_")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==========================================
# EXPORTAR PDF INDIVIDUAL POR GRUPO (ADMIN)
# ==========================================
@app.route('/admin/horarios/grupos/exportar-pdf/<nombre_grupo>')
@login_required
def exportar_horario_individual_pdf(nombre_grupo):
    if not current_user.is_admin():
        abort(403)

    # 1. Obtener horarios del grupo específico
    asignaciones = HorarioAcademico.query.filter_by(grupo=nombre_grupo, activo=True).all()

    if not asignaciones:
        flash(f'No se encontró el horario para el grupo "{nombre_grupo}".', 'danger')
        return redirect(url_for('admin_horario_grupos'))

    # 2. Determinar turno y si tiene sábado
    es_vespertino = 'V' in nombre_grupo[1:3]
    hora_inicio_turno = 13 if es_vespertino else 7
    hora_fin_turno = 21 if es_vespertino else 14
    
    tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
    
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    if tiene_sabado:
        dias_semana.append('Sábado')
    
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

    # 3. Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
    
    # Título
    elements.append(Paragraph(f"Horario del Grupo: {nombre_grupo}", title_style))
    
    # Crear tabla de cuadrícula
    data = [['Hora'] + dias_semana]
    
    for hora in range(hora_inicio_turno, hora_fin_turno):
        hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
        row = [hora_str]
        
        for dia in dias_semana:
            contenido = ""
            for a in asignaciones:
                dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                    contenido = f"{a.materia.nombre}<br/><font size='7'>Prof: {a.profesor.get_nombre_completo()}</font>"
                    break
            row.append(Paragraph(contenido, cell_style))
        
        data.append(row)
    
    # Calcular anchos de columna
    col_widths = [1.0*inch] + [1.4*inch] * len(dias_semana)
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.118, 0.235, 0.447)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.97, 0.97, 0.98)),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (0, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.87, 0.89, 0.90)),
        ('ROWHEIGHTS', (0, 1), (-1, -1), 45),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'horario_{nombre_grupo}.pdf', mimetype='application/pdf')


# ===================================================================
# RUTAS PARA JEFES DE CARRERA 
# ===================================================================

@app.route('/jefe/horarios/profesores')
@login_required
def jefe_ver_horarios_profesores():
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera:
        flash("No tienes una carrera asignada.", "warning")
        return redirect(url_for('dashboard'))

    horarios_data = procesar_horarios(agrupar_por='profesor', carrera_id=id_carrera)
    return render_template('jefe/jefe_horario_profesores.html', horarios_data=horarios_data)


@app.route('/jefe/horarios/grupos')
@login_required
def jefe_ver_horarios_grupos():
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera:
        flash("No tienes una carrera asignada.", "warning")
        return redirect(url_for('dashboard'))

    horarios_data = procesar_horarios(agrupar_por='grupo', carrera_id=id_carrera)
    return render_template('jefe/jefe_horario_grupos.html', horarios_data=horarios_data)


@app.route('/jefe/horarios/grupos/exportar-individual/<nombre_grupo>')
@login_required
def exportar_jefe_horario_grupo_excel(nombre_grupo):
    if not current_user.is_jefe_carrera():
        abort(403)
        
    id_carrera = current_user.primera_carrera_id
    
    # 1. Obtener horarios del grupo específico filtrando por carrera
    asignaciones = HorarioAcademico.query.join(Materia).filter(
        HorarioAcademico.grupo == nombre_grupo,
        HorarioAcademico.activo == True,
        Materia.carrera_id == id_carrera
    ).all()
    
    if not asignaciones:
        flash(f'El grupo {nombre_grupo} no existe o no pertenece a tu carrera.', 'error')
        return redirect(url_for('jefe_ver_horarios_grupos'))

    # 2. Determinar turno (Matutino o Vespertino) y si tiene sábado
    es_vespertino = 'V' in nombre_grupo[1:3]
    hora_inicio_turno = 13 if es_vespertino else 7
    hora_fin_turno = 21 if es_vespertino else 14
    
    tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
    
    # 3. Crear estructura de datos por hora
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    if tiene_sabado:
        dias_semana.append('Sábado')
    
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

    # 4. Crear workbook con openpyxl para mejor formato
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Horario {nombre_grupo}'
    
    # Estilos
    header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    hora_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    hora_font = Font(bold=True, size=10)
    cell_border = Border(
        left=Side(style='thin', color='dee2e6'),
        right=Side(style='thin', color='dee2e6'),
        top=Side(style='thin', color='dee2e6'),
        bottom=Side(style='thin', color='dee2e6')
    )
    
    # Encabezados
    headers = ['Hora'] + dias_semana
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
    
    # Filas por hora
    row_num = 2
    for hora in range(hora_inicio_turno, hora_fin_turno):
        hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
        
        # Celda de hora
        cell = ws.cell(row=row_num, column=1, value=hora_str)
        cell.fill = hora_fill
        cell.font = hora_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
        
        # Celdas de cada día
        for col, dia in enumerate(dias_semana, 2):
            # Buscar clase para esta hora y día
            clase_encontrada = None
            for a in asignaciones:
                dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                    clase_encontrada = a
                    break
            
            if clase_encontrada:
                contenido = f"{clase_encontrada.materia.nombre}\nProf: {clase_encontrada.profesor.get_nombre_completo()}"
            else:
                contenido = ""
            
            cell = ws.cell(row=row_num, column=col, value=contenido)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
        
        row_num += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 30
    
    # Ajustar altura de filas
    for row in range(2, row_num):
        ws.row_dimensions[row].height = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'horario_{nombre_grupo.replace(" ", "_")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==========================================
# EXPORTAR PDF INDIVIDUAL POR GRUPO (JEFE)
# ==========================================
@app.route('/jefe/horarios/grupos/exportar-pdf/<nombre_grupo>')
@login_required
def exportar_jefe_horario_grupo_pdf(nombre_grupo):
    if not current_user.is_jefe_carrera():
        abort(403)
        
    id_carrera = current_user.primera_carrera_id
    
    # 1. Obtener horarios del grupo específico filtrando por carrera
    asignaciones = HorarioAcademico.query.join(Materia).filter(
        HorarioAcademico.grupo == nombre_grupo,
        HorarioAcademico.activo == True,
        Materia.carrera_id == id_carrera
    ).all()
    
    if not asignaciones:
        flash(f'El grupo {nombre_grupo} no existe o no pertenece a tu carrera.', 'error')
        return redirect(url_for('jefe_ver_horarios_grupos'))

    # 2. Determinar turno y si tiene sábado
    es_vespertino = 'V' in nombre_grupo[1:3]
    hora_inicio_turno = 13 if es_vespertino else 7
    hora_fin_turno = 21 if es_vespertino else 14
    
    tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
    
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    if tiene_sabado:
        dias_semana.append('Sábado')
    
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

    # 3. Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
    
    # Título
    elements.append(Paragraph(f"Horario del Grupo: {nombre_grupo}", title_style))
    
    # Crear tabla de cuadrícula
    data = [['Hora'] + dias_semana]
    
    for hora in range(hora_inicio_turno, hora_fin_turno):
        hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
        row = [hora_str]
        
        for dia in dias_semana:
            contenido = ""
            for a in asignaciones:
                dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                    contenido = f"{a.materia.nombre}<br/><font size='7'>Prof: {a.profesor.get_nombre_completo()}</font>"
                    break
            row.append(Paragraph(contenido, cell_style))
        
        data.append(row)
    
    # Calcular anchos de columna
    col_widths = [1.0*inch] + [1.4*inch] * len(dias_semana)
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.118, 0.235, 0.447)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.97, 0.97, 0.98)),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (0, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.87, 0.89, 0.90)),
        ('ROWHEIGHTS', (0, 1), (-1, -1), 45),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'horario_{nombre_grupo}.pdf', mimetype='application/pdf')


# --- Rutas de Exportación para Jefes ---

@app.route('/jefe/horarios/profesores/exportar/excel')
@login_required
def exportar_jefe_horarios_profesor_excel():
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera: return redirect(url_for('dashboard'))
    
    try:
        # Obtener horarios ordenados por día de semana y hora
        asignaciones = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id==id_carrera, 
            HorarioAcademico.activo==True
        ).all()
        
        # Ordenar por día de semana (Lunes a Viernes) y hora
        asignaciones.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
        
        horarios_por_profesor = {}
        for a in asignaciones:
            if not all([a.profesor, a.materia, a.horario]): continue
            profesor_nombre = a.profesor.get_nombre_completo()
            if profesor_nombre not in horarios_por_profesor:
                horarios_por_profesor[profesor_nombre] = {'Lunes': [], 'Martes': [], 'Miércoles': [], 'Jueves': [], 'Viernes': []}
            info = f"{a.materia.nombre}\n{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
            dia_correcto = obtener_dia_correcto(a.dia_semana)
            if dia_correcto in horarios_por_profesor[profesor_nombre]:
                 horarios_por_profesor[profesor_nombre][dia_correcto].append(info)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Horarios Profesores {current_user.carrera.codigo}"
        headers = ["Profesor", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
        ws.append(headers)
        for cell in ws[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
        for profesor, dias in horarios_por_profesor.items():
            ws.append([profesor] + ["\n\n".join(dias.get(dia, [])) for dia in headers[1:]])
        for col_cells in ws.columns:
            ws.column_dimensions[col_cells[0].column_letter].width = 30
            for cell in col_cells: cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'horarios_profesores_{current_user.carrera.codigo}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f"Error al generar el archivo Excel: {e}", "danger")
        return redirect(url_for('jefe/jefe_ver_horarios_profesores'))

@app.route('/jefe/horarios/profesores/exportar/pdf')
@login_required
def exportar_jefe_horarios_profesor_pdf():
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera: return redirect(url_for('dashboard'))
    
    try:
        # Obtener horarios ordenados por día de semana y hora
        asignaciones = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id==id_carrera,
            HorarioAcademico.activo==True
        ).all()
        
        # Ordenar por día de semana (Lunes a Viernes) y hora
        asignaciones.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))
        
        horarios_por_profesor = {}
        for a in asignaciones:
            if not all([a.profesor, a.materia, a.horario]): continue
            profesor_nombre = a.profesor.get_nombre_completo()
            if profesor_nombre not in horarios_por_profesor:
                horarios_por_profesor[profesor_nombre] = {'Lunes': [], 'Martes': [], 'Miércoles': [], 'Jueves': [], 'Viernes': []}
            info = f"{a.materia.nombre}<br/>{a.get_hora_inicio_str()} - {a.get_hora_fin_str()}"
            dia_correcto = obtener_dia_correcto(a.dia_semana)
            if dia_correcto in horarios_por_profesor[profesor_nombre]:
                 horarios_por_profesor[profesor_nombre][dia_correcto].append(info)
    
        styles=getSampleStyleSheet(); styleN = styles['BodyText']
        data = [["Profesor", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]]
        for profesor, dias in horarios_por_profesor.items():
            row_data = [Paragraph(profesor, styleN)]
            for dia in data[0][1:]: row_data.append(Paragraph("<br/><br/>".join(dias.get(dia, [])), styleN))
            data.append(row_data)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        table = Table(data, hAlign='CENTER', colWidths=[1.5*inch]*6)
        style = TableStyle([('BACKGROUND', (0,0), (-1,0), colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('GRID', (0,0), (-1,-1), 1, colors.black)])
        table.setStyle(style); doc.build([table]); buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'horarios_profesores_{current_user.carrera.codigo}.pdf', mimetype='application/pdf')
    except Exception as e:
        flash(f"Error al generar el archivo PDF: {e}", "danger")
        return redirect(url_for('jefe/jefe_ver_horarios_profesores'))
        
@app.route('/jefe/horarios/grupos/exportar/excel')
@login_required
def exportar_jefe_horarios_grupo_excel():
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera: return redirect(url_for('dashboard'))
    
    try:
        # Obtener horarios de la carrera
        asignaciones = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id==id_carrera,
            HorarioAcademico.activo==True
        ).all()
        
        # Obtener grupos únicos
        grupos_unicos = sorted(list(set([a.grupo for a in asignaciones if a.grupo])))
        
        if not grupos_unicos:
            flash('No hay horarios para exportar.', 'warning')
            return redirect(url_for('jefe_ver_horarios_grupos'))
        
        # Mapeo de días
        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        hora_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
        hora_font = Font(bold=True, size=10)
        cell_border = Border(
            left=Side(style='thin', color='dee2e6'),
            right=Side(style='thin', color='dee2e6'),
            top=Side(style='thin', color='dee2e6'),
            bottom=Side(style='thin', color='dee2e6')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for i, grupo in enumerate(grupos_unicos):
            # Verificar que el grupo pertenece a la carrera
            grupo_obj = Grupo.query.filter_by(codigo=grupo).first()
            if not grupo_obj or grupo_obj.carrera_id != id_carrera:
                continue
            
            # Crear hoja para cada grupo
            if i == 0:
                ws = wb.active
                ws.title = grupo[:31]  # Excel limita nombres de hojas a 31 caracteres
            else:
                ws = wb.create_sheet(title=grupo[:31])
            
            # Obtener horarios del grupo
            horarios_grupo = [a for a in asignaciones if a.grupo == grupo]
            
            # Determinar turno y sábado
            es_vespertino = 'V' in grupo[1:3]
            hora_inicio_turno = 13 if es_vespertino else 7
            hora_fin_turno = 21 if es_vespertino else 14
            tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in horarios_grupo)
            
            dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            if tiene_sabado:
                dias_semana.append('Sábado')
            
            # Encabezados
            headers = ['Hora'] + dias_semana
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = cell_border
            
            # Filas por hora
            row_num = 2
            for hora in range(hora_inicio_turno, hora_fin_turno):
                hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
                
                # Celda de hora
                cell = ws.cell(row=row_num, column=1, value=hora_str)
                cell.fill = hora_fill
                cell.font = hora_font
                cell.alignment = center_align
                cell.border = cell_border
                
                # Celdas de cada día
                for col, dia in enumerate(dias_semana, 2):
                    contenido = ""
                    for a in horarios_grupo:
                        dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                        if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                            contenido = f"{a.materia.nombre}\nProf: {a.profesor.get_nombre_completo()}"
                            break
                    
                    cell = ws.cell(row=row_num, column=col, value=contenido)
                    cell.alignment = center_align
                    cell.border = cell_border
                
                row_num += 1
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 15
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col_letter].width = 25
            
            # Ajustar altura de filas
            for row in range(2, row_num):
                ws.row_dimensions[row].height = 45
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'horarios_grupos_{current_user.carrera.codigo}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f"Error al generar el archivo Excel: {e}", "danger")
        return redirect(url_for('jefe_ver_horarios_grupos'))

@app.route('/jefe/horarios/grupos/exportar/pdf')
@login_required
def exportar_jefe_horarios_grupo_pdf():
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera: return redirect(url_for('dashboard'))
    
    try:
        # Obtener grupos únicos de la carrera con horarios
        asignaciones = HorarioAcademico.query.join(Materia).filter(
            Materia.carrera_id==id_carrera,
            HorarioAcademico.activo==True
        ).all()
        
        grupos_unicos = sorted(list(set([a.grupo for a in asignaciones if a.grupo])))
        
        if not grupos_unicos:
            flash('No hay horarios para exportar.', 'warning')
            return redirect(url_for('jefe_ver_horarios_grupos'))
        
        # Mapeo de días
        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}
        
        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)
        
        for grupo in grupos_unicos:
            # Verificar que el grupo pertenece a la carrera
            grupo_obj = Grupo.query.filter_by(codigo=grupo).first()
            if not grupo_obj or grupo_obj.carrera_id != id_carrera:
                continue
            
            # Obtener horarios del grupo
            horarios_grupo = [a for a in asignaciones if a.grupo == grupo]
            
            # Determinar turno y sábado
            es_vespertino = 'V' in grupo[1:3]
            hora_inicio_turno = 13 if es_vespertino else 7
            hora_fin_turno = 21 if es_vespertino else 14
            tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in horarios_grupo)
            
            dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            if tiene_sabado:
                dias_semana.append('Sábado')
            
            # Título del grupo
            elements.append(Paragraph(f"Horario: {grupo}", title_style))
            elements.append(Spacer(1, 10))
            
            # Crear tabla de cuadrícula
            data = [['Hora'] + dias_semana]
            
            for hora in range(hora_inicio_turno, hora_fin_turno):
                hora_str = f"{hora:02d}:00"
                row = [hora_str]
                
                for dia in dias_semana:
                    # Buscar clase para esta hora y día
                    contenido = ""
                    for a in horarios_grupo:
                        dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                        if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                            contenido = f"{a.materia.nombre}<br/>Prof: {a.profesor.get_nombre_completo()}"
                            break
                    row.append(Paragraph(contenido, cell_style))
                
                data.append(row)
            
            # Calcular anchos de columna
            col_widths = [0.8*inch] + [1.4*inch] * len(dias_semana)
            
            table = Table(data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.118, 0.235, 0.447)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.97, 0.97, 0.98)),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.87, 0.89, 0.90)),
                ('ROWHEIGHTS', (0, 1), (-1, -1), 40),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 30))
        
        doc.build(elements)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'horarios_grupos_{current_user.carrera.codigo}.pdf', mimetype='application/pdf')
    except Exception as e:
        flash(f"Error al generar el archivo PDF: {e}", "danger")
        return redirect(url_for('jefe_ver_horarios_grupos'))
    
# =================================================================
# RUTAS DE EXPORTACIÓN EN FORMATO FDA (ADMIN)
# =================================================================
# =================================================================
# RUTA DE EXPORTACIÓN FDA PARA ADMIN (CORREGIDA)
# =================================================================
@app.route('/admin/horarios/profesores/exportar/formato-fda/<profesor_nombre>')
@login_required
def exportar_admin_fda_profesor(profesor_nombre):
    if not current_user.is_admin(): abort(403)

    try:
        horarios_data = procesar_horarios_formato_fda()
        datos_profesor = horarios_data.get(profesor_nombre)

        if not datos_profesor:
            flash(f'No se encontraron datos para el profesor {profesor_nombre}.', 'danger')
            return redirect(url_for('admin_horario_profesores'))
            
        buffer = generar_excel_formato_fda(datos_profesor)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'carga_horaria_{profesor_nombre.replace(" ", "_")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f"Ocurrió un error al generar el reporte: {e}", "danger")
        return redirect(url_for('admin_horario_profesores'))
# =================================================================
# RUTAS DE EXPORTACIÓN EN FORMATO FDA (JEFE DE CARRERA)
# =================================================================
@app.route('/jefe/horarios/profesores/exportar/formato-fda/<profesor_nombre>')
@login_required
def exportar_jefe_fda_profesor(profesor_nombre):
    if not current_user.is_jefe_carrera(): abort(403)
    id_carrera = current_user.primera_carrera_id
    if not id_carrera:
        flash("No tienes una carrera asignada.", "warning")
        return redirect(url_for('dashboard'))

    try:
        horarios_data = procesar_horarios_formato_fda(carrera_id=id_carrera)
        datos_profesor = horarios_data.get(profesor_nombre)

        if not datos_profesor:
            flash(f'No se encontraron datos para {profesor_nombre} en tu carrera.', 'danger')
            return redirect(url_for('jefe_ver_horarios_profesores'))
            
        buffer = generar_excel_formato_fda(datos_profesor)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'carga_horaria_{profesor_nombre.replace(" ", "_")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f"Ocurrió un error al generar el reporte: {e}", "danger")
        return redirect(url_for('jefe_ver_horarios_profesores'))
    

def _generar_excel_horario_profesor(profesor_nombre):
    """
    Función auxiliar interna que genera y retorna el archivo Excel del horario
    con formato de cuadrícula donde cada hora tiene su propia fila.
    """
    try:
        # =====================================================================
        # 1. OBTENER DATOS REALES DEL PROFESOR
        # =====================================================================
        profesor = User.query.filter(
            (User.nombre + ' ' + User.apellido_paterno + ' ' + func.coalesce(User.apellido_materno, '')) == profesor_nombre
        ).first()
        
        if not profesor:
            # Intentar buscar solo con nombre y apellido paterno
            profesor = User.query.filter(
                (User.nombre + ' ' + User.apellido_paterno) == profesor_nombre
            ).first()
        
        if not profesor:
            return "Profesor no encontrado", 404

        # Obtener horarios reales del profesor
        asignaciones = HorarioAcademico.query.filter_by(profesor_id=profesor.id, activo=True).all()
        
        # Determinar si tiene sábado
        tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
        
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        if tiene_sabado:
            dias_semana.append('Sábado')
        
        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

        # =====================================================================
        # 2. CREACIÓN DEL EXCEL CON FORMATO DE CUADRÍCULA
        # =====================================================================
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horario Profesor"
        
        # Estilos
        header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        title_font = Font(bold=True, size=16)
        hora_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
        hora_font = Font(bold=True, size=10)
        cell_border = Border(
            left=Side(style='thin', color='dee2e6'),
            right=Side(style='thin', color='dee2e6'),
            top=Side(style='thin', color='dee2e6'),
            bottom=Side(style='thin', color='dee2e6')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f'Horario de {profesor_nombre}'
        title_cell.font = title_font
        title_cell.alignment = center_align
        
        # Información del profesor
        ws['A3'] = 'Tipo:'
        ws['B3'] = profesor.tipo_profesor or 'No especificado'
        
        # Encabezados de la cuadrícula (fila 5)
        headers = ['Hora'] + dias_semana
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = cell_border
        
        # Filas por hora (7:00 a 21:00)
        row_num = 6
        for hora in range(7, 21):
            hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
            
            # Celda de hora
            cell = ws.cell(row=row_num, column=1, value=hora_str)
            cell.fill = hora_fill
            cell.font = hora_font
            cell.alignment = center_align
            cell.border = cell_border
            
            # Celdas de cada día
            for col, dia in enumerate(dias_semana, 2):
                # Buscar clase para esta hora y día
                clase_encontrada = None
                for a in asignaciones:
                    dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                    if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                        clase_encontrada = a
                        break
                
                if clase_encontrada:
                    contenido = f"{clase_encontrada.materia.nombre}\nGrupo: {clase_encontrada.grupo}"
                else:
                    contenido = ""
                
                cell = ws.cell(row=row_num, column=col, value=contenido)
                cell.alignment = center_align
                cell.border = cell_border
            
            row_num += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 25
        
        # Ajustar altura de filas
        for row in range(6, row_num):
            ws.row_dimensions[row].height = 45
        
        # Calcular total de horas
        total_horas = len(asignaciones)
        ws.cell(row=row_num + 1, column=1, value='Total de horas:').font = Font(bold=True)
        ws.cell(row=row_num + 1, column=2, value=total_horas).font = Font(bold=True)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Horario_{profesor.nombre.replace(' ','_')}_{profesor.apellido_paterno.replace(' ','_')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Ocurrió un error al generar el archivo Excel: {e}", 500


# 2. RUTA ORIGINAL (PARA ADMIN), AHORA SIMPLEMENTE LLAMA A LA FUNCIÓN AUXILIAR
@app.route('/exportar/horario-excel/<profesor_nombre>')
# @login_required
def exportar_excel_profesor(profesor_nombre):
    return _generar_excel_horario_profesor(profesor_nombre)


# 3. NUEVA RUTA (PARA JEFE DE CARRERA), LLAMA A LA MISMA FUNCIÓN AUXILIAR
@app.route('/exportar/horario-jefe-excel/<profesor_nombre>')
# @login_required
def exportar_jefe_excel_profesor(profesor_nombre):
    # Aquí podrías agregar lógica de permisos si es necesario,
    # por ejemplo, verificar que el jefe de carrera solo pueda
    # exportar horarios de profesores de su carrera.
    return _generar_excel_horario_profesor(profesor_nombre)


# =====================================================================
# FUNCIÓN AUXILIAR PARA GENERAR PDF DE HORARIO DE PROFESOR
# =====================================================================
def _generar_pdf_horario_profesor(profesor_nombre):
    """
    Función auxiliar interna que genera y retorna el archivo PDF del horario
    con formato de cuadrícula donde cada hora tiene su propia fila.
    """
    try:
        # 1. OBTENER DATOS REALES DEL PROFESOR
        profesor = User.query.filter(
            (User.nombre + ' ' + User.apellido_paterno + ' ' + func.coalesce(User.apellido_materno, '')) == profesor_nombre
        ).first()
        
        if not profesor:
            profesor = User.query.filter(
                (User.nombre + ' ' + User.apellido_paterno) == profesor_nombre
            ).first()
        
        if not profesor:
            return "Profesor no encontrado", 404

        # Obtener horarios reales del profesor
        asignaciones = HorarioAcademico.query.filter_by(profesor_id=profesor.id, activo=True).all()
        
        # Determinar si tiene sábado
        tiene_sabado = any(a.dia_semana.lower() == 'sabado' for a in asignaciones)
        
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        if tiene_sabado:
            dias_semana.append('Sábado')
        
        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles', 
                    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

        # 2. CREAR PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=15)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
        
        # Título
        elements.append(Paragraph(f"Horario de: {profesor_nombre}", title_style))
        
        tipo_profesor = profesor.tipo_profesor or 'No especificado'
        elements.append(Paragraph(f"Tipo: {tipo_profesor}", subtitle_style))
        
        # Crear tabla de cuadrícula
        data = [['Hora'] + dias_semana]
        
        for hora in range(7, 21):
            hora_str = f"{hora:02d}:00 - {hora+1:02d}:00"
            row = [hora_str]
            
            for dia in dias_semana:
                contenido = ""
                for a in asignaciones:
                    dia_correcto = dias_map.get(a.dia_semana.lower(), '')
                    if dia_correcto == dia and a.horario.hora_inicio.hour == hora:
                        contenido = f"{a.materia.nombre}<br/><font size='7'>Grupo: {a.grupo}</font>"
                        break
                row.append(Paragraph(contenido, cell_style))
            
            data.append(row)
        
        # Calcular anchos de columna
        col_widths = [1.0*inch] + [1.4*inch] * len(dias_semana)
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.118, 0.235, 0.447)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.97, 0.97, 0.98)),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.87, 0.89, 0.90)),
            ('ROWHEIGHTS', (0, 1), (-1, -1), 45),
        ]))
        
        elements.append(table)
        
        # Total de horas
        total_horas = len(asignaciones)
        elements.append(Paragraph(f"<br/><b>Total de horas asignadas: {total_horas}</b>", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"Horario_{profesor.nombre.replace(' ','_')}_{profesor.apellido_paterno.replace(' ','_')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Ocurrió un error al generar el archivo PDF: {e}", 500


# 4. RUTA PDF PARA ADMIN (HORARIO DE PROFESOR)
@app.route('/exportar/horario-pdf/<profesor_nombre>')
# @login_required
def exportar_pdf_profesor(profesor_nombre):
    return _generar_pdf_horario_profesor(profesor_nombre)


# 5. RUTA PDF PARA JEFE DE CARRERA (HORARIO DE PROFESOR)
@app.route('/exportar/horario-jefe-pdf/<profesor_nombre>')
# @login_required
def exportar_jefe_pdf_profesor(profesor_nombre):
    return _generar_pdf_horario_profesor(profesor_nombre)


with app.app_context():
    init_db()
    init_upload_dirs()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)